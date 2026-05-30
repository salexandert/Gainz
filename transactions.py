import os
import zipfile
from openpyxl import load_workbook
import pandas as pd
from conversion import Conversion
from transaction import Buy, Sell, Send, Receive
from assets import Asset
from time import strftime

# Define the base directory for the project
basedir = os.path.dirname(__file__)
FIAT_ASSET_SYMBOLS = {"USD"}


def _optional_cell(row, column):
    if column not in row:
        return None

    value = row[column]
    if pd.isna(value):
        return None

    return value


def _excel_datetime(value):
    if hasattr(value, "replace") and getattr(value, "tzinfo", None):
        return value.replace(tzinfo=None)

    return value

# Updated imports to reflect the new modular structure
from parsers import *
from filters import *
from linking import *
from utils import *
from functools import lru_cache

class Transactions:
    def __init__(self, view=None):
        self.revision_num = None
        self.saves = self.load_saves()
        self.index = 0
        self.conversions = []
        self.asset_objects = []
        self.import_warnings = []
        
        if view is not None:
            self.transactions = self.load(view)
            self.view = view
        
        elif len(self.saves) > 0:
            view = self.saves[-1]['value']
            print(f"loading latest save {view}")
            
            self.transactions = self.load(view)
            self.view = view
        else:
            self.view = ""
            self.transactions = []
            self.revision_num = 0

    def __len__(self):
        return len(self.transactions)

    def __iter__(self):
        self.index = 0
        return self
    
    def __next__(self):
        try:
            result = self.transactions[self.index]
        except IndexError:
            raise StopIteration
        
        self.index += 1
        return result

    @property
    def links(self):
        links = set([
                link 
                for trans in self.transactions
                for link in trans.links
                ])
        
        return links

    def load_saves(self):
        saves = []

        match_object = "saved_"

        view_num = 1
        for root, dirs, files in os.walk(os.path.join(basedir, 'saves')):
            for f in files:  # Corrected the incomplete for loop
                save_as_filename = os.path.join(basedir, 'saves', f)
                if match_object in f and f.endswith('xlsx'):
                    description = ""
                    revision_num = None

                    # Check if the file is a valid zip file
                    if not zipfile.is_zipfile(save_as_filename):
                        print(f"File {save_as_filename} is not a zip file")
                        continue
                    workbook = load_workbook(filename=save_as_filename)
                    if 'Description' in workbook.sheetnames:
                        sheet = workbook['Description']
                        description = sheet.cell(column=1, row=1).value
                        revision_num = sheet.cell(column=2, row=1).value
                    else:
                        description = ""

                    workbook.close()
                    saves.append({
                        'label': save_as_filename,
                        'value': save_as_filename,
                        'description': description,
                        'revision_num': revision_num,
                        'modified_time': os.path.getmtime(save_as_filename),
                    })
                    view_num += 1

        saves.sort(key=lambda save: save['modified_time'])
        self.saves = saves
        
        return saves

    @property
    def assets(self):
        assets = set()

        for trans in self.transactions:
            if trans.symbol not in FIAT_ASSET_SYMBOLS:
                assets.add(trans.symbol)

        return assets

    def get_hodl(self, asset):
        asset = asset.upper()
        for asset_object in self.asset_objects:
            if asset_object.symbol == asset:
                return asset_object.hodl

        return None

    def set_hodl(self, asset, hodl):
        asset = asset.upper()
        hodl = float(hodl)

        for asset_object in self.asset_objects:
            if asset_object.symbol == asset:
                asset_object.hodl = hodl
                return asset_object

        asset_object = Asset(symbol=asset, hodl=hodl)
        self.asset_objects.append(asset_object)
        return asset_object

    def load(self, filename=None):
        # Check if the file is a valid zip file
        if not zipfile.is_zipfile(filename):
            raise zipfile.BadZipFile(f"File {filename} is not a zip file")

        workbook = load_workbook(filename=filename)
        if 'Description' in workbook.sheetnames:
            sheet = workbook['Description']
            description = sheet.cell(column=1, row=1).value
            revision_num = sheet.cell(column=2, row=1).value
            if revision_num is not None:
                self.revision_num = revision_num

        if 'Import Warnings' in workbook.sheetnames:
            warnings_sheet = workbook['Import Warnings']
            self.import_warnings = [
                row[0]
                for row in warnings_sheet.iter_rows(min_row=2, values_only=True)
                if row and row[0]
            ]
        else:
            self.import_warnings = []

        # Read Previously saved data into pandas df - Transactions    
        trans_df = pd.read_excel(filename, sheet_name='All Transactions', converters = {'my_str_column': list})
        trans_df.reset_index(inplace=True)

        # Read Previously saved data into pandas df - Conversions
        conversion_df = pd.read_excel(filename, sheet_name='Conversions', converters = {'my_str_column': list})
        conversion_df.reset_index(inplace=True)

        # Read Previously saved data into pandas df - Assets
        asset_df = pd.read_excel(filename, sheet_name='Assets', converters = {'my_str_column': list})

        # Read Previously saved data into pandas df - Links
        if 'Links' in workbook.sheetnames:
            links_df = pd.read_excel(filename, sheet_name='Links', converters = {'my_str_column': list})
        else:
            links_df = pd.DataFrame()

        # Split Buys and Sells into separate df's
        sell_df = trans_df[(trans_df['trans_type'] == 'sell')].copy()
        buy_df = trans_df[(trans_df['trans_type'] == 'buy')].copy()
        send_df = trans_df[(trans_df['trans_type'] == 'send')].copy()
        receive_df = trans_df[(trans_df['trans_type'] == 'receive')].copy()
        
        send_df.reset_index(inplace=True)
        sell_df.reset_index(inplace=True)
        buy_df.reset_index(inplace=True)
        receive_df.reset_index(inplace=True)
        
        sell_df.sort_values(by='time_stamp', inplace=True)
        buy_df.sort_values(by='time_stamp', inplace=True)
        send_df.sort_values(by='time_stamp', inplace=True)
        receive_df.sort_values(by='time_stamp', inplace=True)

        # Objects > 
        sells = []
        buys = []
        sends = []
        receives = []
        conversions = []
        asset_objects = []

        # Load Transactions into Objects
        # Load Sells
        for index, row in sell_df.iterrows():
            trans_obj = Sell(symbol=row['symbol'], quantity=row['quantity'], time_stamp=row['time_stamp'], usd_spot=row['usd_spot'], source=row['source'], uid=_optional_cell(row, 'uid'))
            # Check if 'fee' column exists before accessing it
            if 'fee' in row:
                trans_obj.fee = row['fee']
            sells.append(trans_obj)

        # Load Buys
        for index, row in buy_df.iterrows():
            trans_obj = Buy(symbol=row['symbol'], quantity=row['quantity'], time_stamp=row['time_stamp'], usd_spot=row['usd_spot'],  source=row['source'], uid=_optional_cell(row, 'uid'))
            # Check if 'fee' column exists before accessing it
            if 'fee' in row:
                trans_obj.fee = row['fee']
            buys.append(trans_obj)

        # Load Sends
        for index, row in send_df.iterrows():
            sends.append(Send(symbol=row['symbol'], quantity=row['quantity'], time_stamp=row['time_stamp'], usd_spot=row['usd_spot'],  source=row['source'], uid=_optional_cell(row, 'uid')))

        # Load Receives
        for index, row in receive_df.iterrows():
            receives.append(Receive(symbol=row['symbol'], quantity=row['quantity'], time_stamp=row['time_stamp'], usd_spot=row['usd_spot'], source=row['source'], uid=_optional_cell(row, 'uid')))

        # Load Conversions
        for index, row in conversion_df.iterrows():
            conversions.append(Conversion(
                input_trans_type=row['input_trans_type'],
                output_trans_type=row['output_trans_type'],
                input_symbol=row['symbol'],
                input_quantity=row['quantity'],
                input_time_stamp=row['time_stamp'],
                input_usd_spot=row['usd_spot'],
                input_usd_total=row['usd_total'],
                reason=row['reason'],
                source=row['source']
                )
            )

        # load Assets
        for index, row in asset_df.iterrows():
            asset_objects.append(Asset(symbol=row['symbol'], hodl=row['hodl']))

        self.asset_objects = asset_objects
        imported_transactions = buys + sells + sends + receives

        # Duplicates check
        transactions = set()
        for trans in imported_transactions:
            transactions.add(trans)

        # Multi-Link Indicators
        for trans in transactions:
            trans.update_linked_transactions()
            trans.set_multi_link()

        self.transactions = list(transactions)
        self.conversions = conversions

        # Re-import Re-Create Links
        print(f"Recreating links from {filename}...") # Add logging
        links_created_count = 0
        links_skipped_count = 0
        transactions_by_uid = {
            str(trans.uid): trans
            for trans in self.transactions
            if getattr(trans, 'uid', None)
        }
        for index, row in links_df.iterrows():
            buy_repr = row['buy'].strip("'") # Reads the __repr__ string like '2021-01-15 10:00:00 0.5'
            sell_repr = row['sell'].strip("'") # Reads the __repr__ string
            quantity_str = row['quantity'] # Reads quantity from Excel

            # Convert quantity string to float, handle potential errors
            try:
                # Ensure quantity_str is treated as string before float conversion
                quantity = float(str(quantity_str))
            except ValueError:
                print(f"  Warning: Could not convert link quantity '{quantity_str}' to float for link between {buy_repr} and {sell_repr}. Skipping link.")
                links_skipped_count += 1
                continue
            except TypeError:
                 print(f"  Warning: Invalid type for link quantity '{quantity_str}' ({type(quantity_str)}). Skipping link.")
                 links_skipped_count += 1
                 continue

            buy_obj = None
            sell_obj = None
            buy_uid = _optional_cell(row, 'buy_uid')
            sell_uid = _optional_cell(row, 'sell_uid')

            if buy_uid and sell_uid:
                buy_obj = transactions_by_uid.get(str(buy_uid))
                sell_obj = transactions_by_uid.get(str(sell_uid))

            if not buy_obj or not sell_obj:
                # Find transaction objects by name (repr string) for older saved files.
                # Using trans.name which is f"{time_stamp} {self.quantity}"
                for trans in self.transactions: # self.transactions contains objects created earlier
                    # Use a direct string comparison for the name/repr
                    if trans.name == sell_repr:
                        if trans.trans_type == 'sell':
                            sell_obj = trans
                    elif trans.name == buy_repr:
                        if trans.trans_type == 'buy':
                            buy_obj = trans
                    if buy_obj and sell_obj:
                        break # Found both

            if sell_obj and buy_obj:
                try:
                    # Ensure quantity is significantly positive before linking (use epsilon)
                    if quantity <= 1e-9:
                         print(f"  Warning: Skipping link with near-zero quantity ({quantity}) between {buy_repr} and {sell_repr}.")
                         links_skipped_count += 1
                         continue

                    # Call link_transaction - this calls Link.__init__ internally
                    sell_obj.link_transaction(buy_obj, link_quantity=quantity)
                    links_created_count += 1

                except ValueError as e:
                     # Catch the specific error from Link.__init__ and provide more context
                     print(f"\n--- ERROR Recreating Link from Saved File ---")
                     print(f"  File: {filename}")
                     print(f"  Link Index: {index}")
                     print(f"  Link Quantity from file: {quantity_str} -> {quantity}")
                     print(f"  Buy Trans Repr: {buy_repr}")
                     print(f"  Sell Trans Repr: {sell_repr}")
                     if buy_obj:
                         print(f"  Buy Obj Found: ID={buy_obj.id}, Current Unlinked: {buy_obj.unlinked_quantity:.8f}, Total Qty: {buy_obj.quantity:.8f}")
                     else:
                         print(f"  Buy Obj NOT Found for repr: {buy_repr}")
                     if sell_obj:
                         print(f"  Sell Obj Found: ID={sell_obj.id}, Current Unlinked: {sell_obj.unlinked_quantity:.8f}, Total Qty: {sell_obj.quantity:.8f}")
                     else:
                         print(f"  Sell Obj NOT Found for repr: {sell_repr}")
                     print(f"  Original Error: {e}")
                     print(f"---------------------------------------------\n")
                     links_skipped_count += 1
                     # Decide whether to raise or just continue
                     # raise e # Re-raise to stop execution as before
                     print("Continuing load process despite error...") # Or just log and continue

                except Exception as e:
                     print(f"  Unexpected error recreating link between {buy_repr} and {sell_repr}: {e}")
                     links_skipped_count += 1
                     # Handle other potential errors during linking

            else:
                print(f"  Warning: Could not find matching buy/sell objects for link between {buy_repr} and {sell_repr}. Skipping link.")
                links_skipped_count += 1

        print(f"Finished recreating links: {links_created_count} created, {links_skipped_count} skipped/errors.") # Add summary logging
        return list(transactions)    
    
    def save(self, description=None):
        save_as_filename = os.path.join(basedir, "saves", f"saved_{strftime('Y%Y-M%m-D%d_H%H-M%M-S%S')}.xlsx")
        
        # Make sure all transactions are properly updated before saving
        for trans in self.transactions:
            trans.update_linked_transactions()
            trans.set_multi_link()
            # Remove timezone info to prevent Excel issues
            if hasattr(trans.time_stamp, 'replace') and trans.time_stamp.tzinfo is not None:
                trans.time_stamp = trans.time_stamp.replace(tzinfo=None)
                
            # Update link timestamps too
            for link in trans.links:
                if hasattr(link.buy.time_stamp, 'replace') and link.buy.time_stamp.tzinfo is not None:
                    link.buy.time_stamp = link.buy.time_stamp.replace(tzinfo=None)
                if hasattr(link.sell.time_stamp, 'replace') and link.sell.time_stamp.tzinfo is not None:
                    link.sell.time_stamp = link.sell.time_stamp.replace(tzinfo=None)
        
        # Create DataFrames from transaction data
        trans_df = pd.DataFrame([vars(s) for s in self.transactions])
        conversion_df = pd.DataFrame([vars(s) for s in self.conversions])
        asset_df = pd.DataFrame([vars(s) for s in self.asset_objects])
        
        # Extract all links to ensure they're saved properly
        links_data = []
        for l in self.links:
            links_data.append({
                'id': l.symbol,
                'quantity': str(l.quantity),
                'buy_uid': getattr(l.buy, 'uid', None),
                'sell_uid': getattr(l.sell, 'uid', None),
                'buy': str(l.buy),
                'sell': str(l.sell),
                'symbol': l.symbol
            })
        links_df = pd.DataFrame(links_data)

        # Write all data to Excel file with pandas for consistent formatting
        with pd.ExcelWriter(save_as_filename, engine='xlsxwriter') as writer:
            trans_df.to_excel(writer, sheet_name="All Transactions")
            conversion_df.to_excel(writer, sheet_name="Conversions")
            asset_df.to_excel(writer, sheet_name="Assets")
            
            links_df.to_excel(writer, sheet_name="Links")
        
        # Open file to add description and update revision number
        workbook = load_workbook(filename=save_as_filename)
        
        # Add description sheet
        sheet = workbook.create_sheet('Description')
        sheet.cell(row=1, column=1, value=description)
        revision_num = self.revision_num        
        if revision_num is None:
            revision_num = 0
        sheet.cell(row=1, column=2, value=revision_num + 1)

        if getattr(self, 'import_warnings', None):
            warnings_sheet = workbook.create_sheet('Import Warnings')
            warnings_sheet.cell(row=1, column=1, value='Warning')
            for row_num, warning in enumerate(self.import_warnings, start=2):
                warnings_sheet.cell(row=row_num, column=1, value=warning)
        
        # Save and close
        workbook.save(save_as_filename)
        workbook.close()
        
        print(f"{'[' + description + ']' if description else ''} Saving to {save_as_filename}")

        # Update internal state
        self.saves = self.load_saves()
        self.view = save_as_filename
        
        return save_as_filename
 
    def export_to_excel(self, asset=None, date_range=None, by_year=True, output_dir=None):

        # Idea to programatically create Excel Links, Fancy ;-)
        # =HYPERLINK("[Export_Y2021-M03-D06_H19-M34.xlsx]Links!A20","Display Text")
        
        export_dir = output_dir or os.path.join(basedir, "exports")
        os.makedirs(export_dir, exist_ok=True)
        save_as_filename = os.path.join(export_dir, f"Export_{strftime('Y%Y-M%m-D%d_H%H-M%M')}.xlsx")
        
        # Template to use
        workbook = load_workbook(filename= os.path.join(basedir, 'Gainz_Export_Template-DO_NOT_MODIFY.xlsx'))
        c_sheet = workbook['Conversions']
        l_sheet = workbook['Gains']
        a_sheet = workbook['All Transactions']
        s_sheet = workbook['Stats']
        t8949_sheet = workbook['8949']
        sales_sheet = workbook['Sales']

        
        sales_rows = get_sales_report_rows(self)
        years = sorted({row["year"] for row in sales_rows})

        # Sales
        for year in years:
            print(f'exporting sales for {year}')

            ws = workbook.copy_worksheet(sales_sheet)
            ws.title = f'{year} Sales'

            row = 2
            for sale in sales_rows:
                if sale["year"] != year:
                    continue

                ws[f"A{row}"] = sale["description"]
                ws[f"B{row}"] = _excel_datetime(sale["date_acquired"])
                ws[f"C{row}"] = _excel_datetime(sale["date_sold"])
                ws[f"D{row}"] = sale["proceeds"]
                ws[f"D{row}"].number_format = '"$"#,##0.00_-'
                ws[f"E{row}"] = sale["cost_basis"]
                ws[f"E{row}"].number_format = '"$"#,##0.00_-'
                ws[f"F{row}"] = sale["gain_loss"]
                ws[f"F{row}"].number_format = '"$"#,##0.00_-'
                ws[f"G{row}"] = sale["source"]
                row += 1

        form_rows = get_form_8949_report_rows(self)
        form_years = sorted({row["year"] for row in form_rows})

        for term in ("short", "long"):
            term_label = term.capitalize()
            for year in form_years:
                ws = workbook.copy_worksheet(t8949_sheet)
                ws.title = f'{year} 8949 {term_label}'

                row = 2
                for form_row in form_rows:
                    if form_row["year"] != year or form_row["term"] != term:
                        continue

                    ws[f"A{row}"] = form_row["description"]
                    ws[f"B{row}"] = _excel_datetime(form_row["date_acquired"])
                    ws[f"C{row}"] = _excel_datetime(form_row["date_sold"])
                    ws[f"D{row}"] = form_row["proceeds"]
                    ws[f"D{row}"].number_format = '"$"#,##0.00_-'
                    ws[f"E{row}"] = form_row["cost_basis"]
                    ws[f"E{row}"].number_format = '"$"#,##0.00_-'
                    ws[f"H{row}"] = form_row["gain_loss"]
                    ws[f"H{row}"].number_format = '"$"#,##0.00_-'
                    ws[f"I{row}"] = form_row["source"]
                    row += 1

                if row == 2:
                    workbook.remove(ws)
                else:
                    row += 2
                    ws[f"C{row}"] = "Totals"
                    ws[f"D{row}"] = f"=SUM(D2:D{row -2})"
                    ws[f"D{row}"].number_format = '"$"#,##0.00_-'
                    ws[f"E{row}"] = f"=SUM(E2:E{row -2})"
                    ws[f"E{row}"].number_format = '"$"#,##0.00_-'
                    ws[f"H{row}"] = f"=SUM(H2:H{row -2})"
                    ws[f"H{row}"].number_format = '"$"#,##0.00_-'


        for asset in self.assets:
            
            # Conversions sheet
            sheetname = f'{asset} Conversions'
            conversions_sheet = workbook.copy_worksheet(c_sheet)
            conversions_sheet.title = sheetname

            column_names = []
            for cell in conversions_sheet[3]:
                column_names.append(cell.value)
            
            in_trans_type_index = column_names.index("In Transaction Type") + 1
            out_trans_type_index = column_names.index("Out Transaction Type") + 1
            symbol_index = column_names.index("Symbol") + 1
            time_stamp_index = column_names.index("Time Stamp") + 1
            quantity_index = column_names.index("Quantity") + 1
            usd_spot_index = column_names.index("USD Spot") + 1
            usd_total_index = column_names.index("USD Total") + 1
            reason_index = column_names.index("Reason") + 1

            row = 4
            for conversion in self.conversions:
                
                if conversion.symbol != asset:
                    continue

                conversions_sheet.cell(row=row, column=in_trans_type_index, value=conversion.input_trans_type)
                conversions_sheet.cell(row=row, column=out_trans_type_index, value=conversion.output_trans_type)
                conversions_sheet.cell(row=row, column=symbol_index, value=conversion.symbol)
                conversions_sheet.cell(row=row, column=time_stamp_index, value=_excel_datetime(conversion.time_stamp))
                conversions_sheet.cell(row=row, column=quantity_index, value=conversion.quantity)
                
                conversions_sheet.cell(row=row, column=usd_spot_index, value=conversion.usd_spot)
                conversions_sheet.cell(row=row, column=usd_spot_index).number_format = '"$"#,##0.00_-'
                
                conversions_sheet.cell(row=row, column=usd_total_index, value=conversion.usd_total)
                conversions_sheet.cell(row=row, column=usd_total_index).number_format = '"$"#,##0.00_-'
                
                conversions_sheet.cell(row=row, column=reason_index, value=conversion.reason)

                row += 1

            if row == 4:
                workbook.remove(conversions_sheet)


            # Gainz Sheet
            column_names = []
            for cell in l_sheet[1]:
                column_names.append(cell.value)

            sell_date_index = column_names.index("Sell Date") + 1
            sell_id_index = column_names.index("Sell ID") + 1
            sell_quantity_index = column_names.index("Sell Quantity") + 1
            sell_unlinked_index = column_names.index("Sell Unlinked") + 1
            sell_usd_total_index = column_names.index("Sell USD Total") + 1
            sell_usd_spot_index = column_names.index("Sell Spot USD") + 1
            sell_multi_link_index = column_names.index("Sell Multi-Link") + 1

            buy_link_usd_index = column_names.index("Link Buy in USD") + 1
            link_id_index = column_names.index("Link ID") + 1
            link_symbol_index = column_names.index("Link Asset") + 1
            link_quantity_index = column_names.index("Link Quantity") + 1
            link_profit_loss_index = column_names.index("Link Profit Loss") + 1
            sell_link_usd_index = column_names.index("Link Sell in USD") + 1
            date_acquired_index = column_names.index("Date Acquired") + 1

            buy_multi_link_index = column_names.index("Buy Multi-Link") + 1
            buy_date_index = column_names.index("Buy Date") + 1
            buy_id_index = column_names.index("Buy ID") + 1
            buy_quantity_index = column_names.index("Buy Quantity") + 1
            buy_unlinked_index = column_names.index("Buy Unlinked") + 1
            buy_usd_total_index = column_names.index("Buy USD Total") + 1
            buy_usd_spot_index = column_names.index("Buy Spot USD") + 1

            years = set()
            for link in self.links:
                if link.symbol != asset:
                    continue

                years.add(link.sell.time_stamp.year)

            for year in years:
                
                sheetname = f'{year} {asset} Gains'
                links_sheet = workbook.copy_worksheet(l_sheet)
                links_sheet.title = sheetname
                        
                row = 2
                profit_loss_total = 0.0
                for link in self.links:
                    
                    if link.symbol != asset:
                        continue
                        
                    if link.sell.time_stamp.year != year:
                        continue

                    if link.quantity <= 0.00000001:
                        continue

                    profit_loss_total += float(link.profit_loss)
                    
                    links_sheet.cell(row=row, column=link_symbol_index, value=link.sell.symbol)
                    links_sheet.cell(row=row, column=link_id_index, value=link.id)
                    links_sheet.cell(row=row, column=buy_id_index, value=link.buy.id)
                    links_sheet.cell(row=row, column=sell_id_index, value=link.sell.id)
                    links_sheet.cell(row=row, column=buy_date_index, value=_excel_datetime(link.buy.time_stamp))
                    links_sheet.cell(row=row, column=buy_quantity_index, value=link.buy.quantity)
                    links_sheet.cell(row=row, column=buy_unlinked_index, value=link.buy.unlinked_quantity)
                    
                    links_sheet.cell(row=row, column=buy_usd_spot_index, value=link.buy.usd_spot)
                    links_sheet.cell(row=row, column=buy_usd_spot_index).number_format = '"$"#,##0.00_-'

                    links_sheet.cell(row=row, column=buy_usd_total_index, value=link.buy.usd_total)
                    links_sheet.cell(row=row, column=buy_usd_total_index).number_format = '"$"#,##0.00_-'

                    links_sheet.cell(row=row, column=buy_link_usd_index, value=link.link_buy_price)
                    links_sheet.cell(row=row, column=buy_link_usd_index).number_format = '"$"#,##0.00_-'

                    links_sheet.cell(row=row, column=link_quantity_index, value=link.quantity)
                    
                    links_sheet.cell(row=row, column=link_profit_loss_index, value=link.profit_loss)
                    links_sheet.cell(row=row, column=link_profit_loss_index).number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'

                    links_sheet.cell(row=row, column=sell_link_usd_index, value=link.link_sell_price)
                    links_sheet.cell(row=row, column=sell_link_usd_index).number_format = '"$"#,##0.00_-'

                    links_sheet.cell(row=row, column=sell_date_index, value=_excel_datetime(link.sell.time_stamp))
                    links_sheet.cell(row=row, column=sell_quantity_index, value=link.sell.quantity)
                    links_sheet.cell(row=row, column=sell_unlinked_index, value=link.sell.unlinked_quantity)

                    links_sheet.cell(row=row, column=sell_usd_spot_index, value=link.sell.usd_spot)
                    links_sheet.cell(row=row, column=sell_usd_spot_index).number_format = '"$"#,##0.00_-'

                    links_sheet.cell(row=row, column=sell_usd_total_index, value=link.sell.usd_total)
                    links_sheet.cell(row=row, column=sell_usd_total_index).number_format = '"$"#,##0.00_-'
                    
                    links_sheet.cell(row=row, column=sell_multi_link_index, value=link.sell.multi_link)
                    links_sheet.cell(row=row, column=buy_multi_link_index, value=link.buy.multi_link)
                    
                    row += 1

                for trans in self.transactions:
                    
                    if trans.symbol != asset:
                        continue

                    if trans.trans_type != 'sell':
                        continue

                    if trans.time_stamp.year != year:
                        continue

                    if trans.unlinked_quantity <= 0.00000001:
                        continue

                    profit_loss_total += float(trans.unlinked_quantity * trans.usd_spot)

                    links_sheet.cell(row=row, column=link_symbol_index, value="N/A")
                    links_sheet.cell(row=row, column=link_id_index, value="N/A")
                    links_sheet.cell(row=row, column=buy_id_index, value="N/A")
                    links_sheet.cell(row=row, column=sell_id_index, value=trans.id)
                    links_sheet.cell(row=row, column=buy_date_index, value="N/A")
                    links_sheet.cell(row=row, column=buy_quantity_index, value="N/A")
                    links_sheet.cell(row=row, column=buy_unlinked_index, value="N/A")
                    links_sheet.cell(row=row, column=buy_usd_spot_index, value="N/A")
                    links_sheet.cell(row=row, column=buy_usd_total_index, value="N/A")
                    links_sheet.cell(row=row, column=buy_link_usd_index, value="N/A")
                    links_sheet.cell(row=row, column=link_quantity_index, value=trans.unlinked_quantity)

                    links_sheet.cell(row=row, column=link_profit_loss_index, value=(trans.unlinked_quantity * trans.usd_spot))
                    links_sheet.cell(row=row, column=link_profit_loss_index).number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'
                    
                    links_sheet.cell(row=row, column=sell_link_usd_index, value=trans.usd_total)
                    links_sheet.cell(row=row, column=sell_link_usd_index).number_format = '"$"#,##0.00_-'

                    links_sheet.cell(row=row, column=sell_date_index, value=_excel_datetime(trans.time_stamp))
                    links_sheet.cell(row=row, column=sell_quantity_index, value=trans.quantity)
                    links_sheet.cell(row=row, column=sell_unlinked_index, value=trans.unlinked_quantity)
                    links_sheet.cell(row=row, column=sell_usd_spot_index, value=trans.usd_spot)
                    links_sheet.cell(row=row, column=sell_usd_spot_index).number_format = '"$"#,##0.00_-'

                    links_sheet.cell(row=row, column=sell_usd_total_index, value=trans.usd_total)
                    links_sheet.cell(row=row, column=sell_usd_total_index).number_format = '"$"#,##0.00_-'

                    links_sheet.cell(row=row, column=sell_multi_link_index, value="N/A")
                    links_sheet.cell(row=row, column=buy_multi_link_index, value="N/A")

                    row += 1
                    

                row += 2

                links_sheet.cell(row=row, column=link_profit_loss_index, value="Profit/Loss Total: ${:,.2f}".format(profit_loss_total)) 

                if row == 4:
                    workbook.remove(links_sheet)

            # All Transactions sheet
            sheetname = f'{asset} Transactions'
            all_trans_sheet = workbook.copy_worksheet(a_sheet)
            all_trans_sheet.title = sheetname
            
            column_names = []
            for cell in all_trans_sheet[1]:
                column_names.append(cell.value)

            id_index = column_names.index("Transaction ID") + 1
            symbol_index = column_names.index("Symbol") + 1
            trans_type_index = column_names.index("Transaction Type") + 1
            time_stamp_index = column_names.index("Time Stamp") + 1 
            quantity_index = column_names.index("Quantity") + 1
            links_index = column_names.index("Links") + 1
            unlinked_index = column_names.index("Unlinked") + 1
            usd_spot_index = column_names.index("USD Spot") + 1
            usd_total_index = column_names.index("USD Total") + 1
            source_index = column_names.index("Source") + 1
            
            row = 2
            for trans in self.transactions:

                if trans.symbol != asset:
                    continue

                all_trans_sheet.cell(row=row, column=id_index, value=trans.id)
                all_trans_sheet.cell(row=row, column=symbol_index, value=trans.symbol)
                all_trans_sheet.cell(row=row, column=trans_type_index, value=trans.trans_type)
                all_trans_sheet.cell(row=row, column=time_stamp_index, value=_excel_datetime(trans.time_stamp))
                all_trans_sheet.cell(row=row, column=quantity_index, value=trans.quantity)
                all_trans_sheet.cell(row=row, column=unlinked_index, value=trans.unlinked_quantity)

                all_trans_sheet.cell(row=row, column=usd_spot_index, value=trans.usd_spot)
                all_trans_sheet.cell(row=row, column=usd_spot_index).number_format = '"$"#,##0.00_-'

                all_trans_sheet.cell(row=row, column=usd_total_index, value=trans.usd_total)
                all_trans_sheet.cell(row=row, column=usd_total_index).number_format = '"$"#,##0.00_-'

                all_trans_sheet.cell(row=row, column=source_index, value=trans.source)

                if len(trans.links) > 0:
                    all_trans_sheet.cell(row=row, column=links_index, value=str(trans.links))

                row += 1

            if row == 2:
                workbook.remove(all_trans_sheet)


            # Links Sheet
            sheetname = f'{asset} Stats'
            asset_stats_sheet = workbook.copy_worksheet(s_sheet)
            asset_stats_sheet.title = sheetname


            date_range = {
                'start_date': '',
                'end_date': ''
            }

            date_range = get_transactions_date_range(self, date_range)

            # get stats table data 
            stats_table_data = get_stats_table_data_range(self, date_range)

            # get stats for selected asset
            asset_stats = None
            for a in stats_table_data:
                if a['symbol'] == asset:
                    asset_stats = a
                    break
                
           
            # print(asset_stats)

            # Create detailed stats table data
            detailed_stats = [
                ["Quantity Purchased", asset_stats['total_purchased_quantity']],
                ["Number of Buys", asset_stats["num_buys"]],
                ["Number of Sells", asset_stats["num_sells"]],
                ["Number of Links", asset_stats["num_links"]],
                ["Average Buy Price", asset_stats["average_buy_price"]],
                ["Average Sell Price", asset_stats["average_sell_price"]],
                ["Quantity Sold", asset_stats['total_sold_quantity']],
                ["Quantity Sold Unlinked", asset_stats['total_sold_unlinked_quantity']],
                ["Quantity Purchased Unlinked", asset_stats['total_purchased_unlinked_quantity']],
                ["Quantity Purchased in USD", asset_stats['total_purchased_usd']],
                ["Quantity Sold in USD", asset_stats['total_sold_usd']],
                ["Profit / Loss in USD* (Valid when Quantity Sold Unlinked is 0)", asset_stats['profit_loss_total']],
            ]

            row = 2
            for i in detailed_stats:
                asset_stats_sheet.cell(row=row, column=1, value=i[0])
                asset_stats_sheet.cell(row=row, column=1).number_format = '"$"#,##0.00_-'
                asset_stats_sheet.cell(row=row, column=2, value=i[1])
                asset_stats_sheet.cell(row=row, column=2).number_format = '"$"#,##0.00_-'
                
                row += 1

        
        workbook.remove(a_sheet)
        workbook.remove(c_sheet)
        workbook.remove(l_sheet)
        workbook.remove(s_sheet)
        workbook.remove(t8949_sheet)
        workbook.remove(sales_sheet)
            
            
        workbook.save(save_as_filename)
        print(f"Saving to {save_as_filename}")

        return save_as_filename

    def _conversion_amount(self, asset, current_hodl=None, amount_to_convert=None):
        if amount_to_convert is not None:
            return max(float(amount_to_convert), 0.0)

        if current_hodl is None:
            return 0.0

        bought = sum(trans.quantity for trans in self.transactions if trans.symbol == asset and trans.trans_type == 'buy')
        sold = sum(trans.quantity for trans in self.transactions if trans.symbol == asset and trans.trans_type == 'sell')

        return max((bought - sold) - float(current_hodl), 0.0)

    def _reduce_transaction_quantity(self, trans, quantity):
        remaining = round_decimals_down(trans.quantity - quantity, decimals=9)
        if remaining <= 0.000000001:
            self.transactions.remove(trans)
            return

        trans.quantity = remaining
        trans.name = f"{trans.time_stamp} {trans.quantity}"

    def _record_conversion(self, trans, quantity, output_trans_type, reason):
        conversion = Conversion(
            input_trans_type=trans.trans_type,
            output_trans_type=output_trans_type,
            input_symbol=trans.symbol,
            input_quantity=quantity,
            input_time_stamp=trans.time_stamp,
            input_usd_spot=trans.usd_spot,
            input_usd_total=quantity * trans.usd_spot,
            reason=reason,
            source=f"{trans.source} Converted in Gainz App",
        )
        self.conversions.append(conversion)

    def convert_sends_to_sells(self, asset, current_hodl=None, amount_to_convert=None):
        amount_remaining = self._conversion_amount(asset, current_hodl=current_hodl, amount_to_convert=amount_to_convert)
        converted_quantity = 0.0
        converted_count = 0

        sends = [
            trans for trans in self.transactions
            if trans.symbol == asset and trans.trans_type == 'send'
        ]
        sends.sort(key=lambda x: x.time_stamp)

        for send in sends:
            if amount_remaining <= 0.000000001:
                break

            quantity = min(send.quantity, amount_remaining)
            sell = Sell(
                symbol=send.symbol,
                quantity=quantity,
                time_stamp=send.time_stamp,
                usd_spot=send.usd_spot,
                source="Gainz App Send to Sell",
            )

            self.transactions.append(sell)
            self._record_conversion(send, quantity, 'sell', 'Converted Send to Sell')
            self._reduce_transaction_quantity(send, quantity)

            amount_remaining -= quantity
            converted_quantity += quantity
            converted_count += 1

        return f"Converted {converted_quantity} {asset} from {converted_count} send transaction(s) to sell transaction(s)."

    def convert_buys_to_lost(self, asset, amount):
        amount_remaining = max(float(amount), 0.0)
        converted_quantity = 0.0
        converted_count = 0

        buys = [
            trans for trans in self.transactions
            if trans.symbol == asset and trans.trans_type == 'buy' and trans.unlinked_quantity > 0.000000001
        ]
        buys.sort(key=lambda x: x.time_stamp, reverse=True)

        for buy in buys:
            if amount_remaining <= 0.000000001:
                break

            quantity = min(buy.unlinked_quantity, amount_remaining)
            self._record_conversion(buy, quantity, 'lost', 'Converted Buy to Lost')
            self._reduce_transaction_quantity(buy, quantity)

            amount_remaining -= quantity
            converted_quantity += quantity
            converted_count += 1

        return f"Converted {converted_quantity} {asset} from {converted_count} buy transaction(s) to lost."

    def convert_receives_to_buys(self, asset, amount_to_convert):
        amount_remaining = max(float(amount_to_convert), 0.0)
        converted_quantity = 0.0
        converted_count = 0

        receives = [
            trans for trans in self.transactions
            if trans.symbol == asset and trans.trans_type == 'receive' and trans.unlinked_quantity > 0.000000001
        ]
        receives.sort(key=lambda x: x.time_stamp)

        for receive in receives:
            if amount_remaining <= 0.000000001:
                break

            quantity = min(receive.unlinked_quantity, amount_remaining)
            buy = Buy(
                symbol=receive.symbol,
                quantity=quantity,
                time_stamp=receive.time_stamp,
                usd_spot=receive.usd_spot,
                source="Gainz App Receive to Buy",
            )

            self.transactions.append(buy)
            self._record_conversion(receive, quantity, 'buy', 'Converted Receive to Buy')
            self._reduce_transaction_quantity(receive, quantity)

            amount_remaining -= quantity
            converted_quantity += quantity
            converted_count += 1

        return f"Converted {converted_quantity} {asset} from {converted_count} receive transaction(s) to buy transaction(s)."


    def delete(self, filename):
        os.rename(filename, f"{filename}.bak")

    def first_transaction_date(self, asset=None):
        """
        Returns a dictionary with the first transaction date for each asset or a specific asset.

        Args:
            asset (str, optional): The symbol of the asset to get the first transaction date for. 
                                   If None, returns dates for all assets. Defaults to None.

        Returns:
            dict: A dictionary with asset symbols as keys and their first transaction dates as values.
        """
        all_trans = {}
        
        for trans in self.transactions:
            # If asset is provided skip others
            if asset is not None:
                if trans.symbol != asset:
                    continue
            
            # Create key val for symbol
            if trans.symbol not in all_trans.keys():
                all_trans[trans.symbol] = []

            all_trans[trans.symbol].append(trans)

        # Sort By Time Stamp
        for key in all_trans.keys():
            all_trans[key].sort(key=lambda x: x.time_stamp.replace(tzinfo=None))        # Extract first transaction Date
        first_time_stamps = {}
        for key in all_trans.keys():
            first_time_stamps[key] = all_trans[key][0].time_stamp.replace(tzinfo=None)
            
        return first_time_stamps
        
    def last_transaction_date(self, asset=None):
        """
        Returns a dictionary with the last transaction date for each asset or a specific asset.

        Args:
            asset (str, optional): The symbol of the asset to get the last transaction date for. 
                                  If None, returns dates for all assets. Defaults to None.

        Returns:
            dict: A dictionary with asset symbols as keys and their last transaction dates as values.
        """
        all_trans = {}
        
        for trans in self.transactions:
            # If asset is provided skip others
            if asset is not None:
                if trans.symbol != asset:
                    continue
            
            # Create key val for symbol
            if trans.symbol not in all_trans.keys():
                all_trans[trans.symbol] = []

            all_trans[trans.symbol].append(trans)

        # Sort By Time Stamp
        for key in all_trans.keys():
            all_trans[key].sort(key=lambda x: x.time_stamp.replace(tzinfo=None))

        # Extract last transaction Date
        last_time_stamps = {}
        for key in all_trans.keys():
            if all_trans[key]:  # Check if the list is not empty
                last_time_stamps[key] = all_trans[key][-1].time_stamp.replace(tzinfo=None)
            
        return last_time_stamps
    
    def auto_link(self, asset=None, algo='fifo', date_range=None, min_unlinked=0.000001, year=None):
        """
        Automatically links buy and sell transactions based on specified algorithm.
        
        Args:
            asset (str, optional): The asset to link. If None, link all assets.
            algo (str, optional): The algorithm to use ('fifo' or 'filo'). Defaults to 'fifo'.
            date_range (dict, optional): Date range to filter transactions by.
            min_unlinked (float, optional): Minimum unlinked quantity to consider. Defaults to 0.000001.
            year (int, optional): Tax year to filter sells by. Ignored when date_range is provided.
            
        Returns:
            list: List of failures where sells couldn't be fully linked.
        """
        from dateutil import parser
        
        buys = {}
        sells = {}
        failures = []

        if year is not None and date_range is None:
            date_range = {
                'start_date': f"01/01/{year} 12:00 AM",
                'end_date': f"12/31/{year} 11:59 PM",
            }

        # Get buys and sells by asset
        for trans in self.transactions:
            if asset is not None and trans.symbol != asset:
                continue
                
            if trans.symbol not in buys:
                buys[trans.symbol] = []
                sells[trans.symbol] = []
                
            if trans.trans_type == 'buy':
                buys[trans.symbol].append(trans)
            elif trans.trans_type == 'sell':
                sells[trans.symbol].append(trans)
        
        # Filter by date range if provided
        if date_range is not None:
            try:
                tzinfos = {"EST": -5 * 3600, "EDT": -4 * 3600}
                start_date = parser.parse(date_range['start_date'], tzinfos=tzinfos)
                end_date = parser.parse(date_range['end_date'], tzinfos=tzinfos)
                
                # Filter Transactions to date range
                for key in sells.keys():
                    filtered_transactions = []
                    for trans in sells[key]:
                        if isinstance(trans.time_stamp, str):
                            trans_time_stamp = parser.parse(trans.time_stamp, tzinfos=tzinfos)
                        else:
                            trans_time_stamp = trans.time_stamp
                        
                        # Make all timestamps timezone-naive for comparison
                        if hasattr(trans_time_stamp, 'tzinfo') and trans_time_stamp.tzinfo:
                            trans_time_stamp = trans_time_stamp.replace(tzinfo=None)
                        if hasattr(start_date, 'tzinfo') and start_date.tzinfo:
                            start_date = start_date.replace(tzinfo=None)
                        if hasattr(end_date, 'tzinfo') and end_date.tzinfo:
                            end_date = end_date.replace(tzinfo=None)
                            
                        if trans_time_stamp >= start_date and trans_time_stamp <= end_date:              
                            filtered_transactions.append(trans)

                    sells[key] = filtered_transactions
                    
            except Exception as e:
                print(f"Error filtering by year: {e}")
        
        # sort for algo types
        if algo == 'fifo':
            # sort buys and sells by time_stamp
            for key in buys.keys():
                buys[key].sort(key=lambda x: x.time_stamp.replace(tzinfo=None) if hasattr(x.time_stamp, 'replace') else x.time_stamp)
            
            for key in sells.keys():
                sells[key].sort(key=lambda x: x.time_stamp.replace(tzinfo=None) if hasattr(x.time_stamp, 'replace') else x.time_stamp)

            from utils import round_decimals_down
            
            keys = list(sells.keys())
            keys.sort()
            for key in keys:
                quantity_linked = 0.0
                links = []

                # loop sells to find link candidate
                for sell in sells[key]:
                    # check if sell has remaining unlinked quantity
                    if sell.unlinked_quantity > min_unlinked:
                        
                        #loop buys to find link candidate
                        for buy in buys[key]:
                            link_quantity = None

                            # break if sell has no remaining unlinked quantity
                            if sell.unlinked_quantity <= min_unlinked:
                                break

                            # Skip if buy has no remaining unlinked quantity
                            if buy.unlinked_quantity <= min_unlinked:
                                continue
                                                    
                            # check if buy came before sell
                            buy_time = buy.time_stamp.replace(tzinfo=None) if hasattr(buy.time_stamp, 'replace') else buy.time_stamp
                            sell_time = sell.time_stamp.replace(tzinfo=None) if hasattr(sell.time_stamp, 'replace') else sell.time_stamp
                            
                            if buy_time >= sell_time:
                                continue

                            # Link 
                            # if sell unlinked is greater than buy unlinked, link quantity equals buy unlinked
                            if sell.unlinked_quantity >= buy.unlinked_quantity:
                                link_quantity = buy.unlinked_quantity
                            
                            # if sell unlinked is less than buy unlinked, link quantity equals sell unlinked
                            elif sell.unlinked_quantity <= buy.unlinked_quantity: 
                                link_quantity = sell.unlinked_quantity
                            
                            # Set max length of link 
                            link_quantity = round_decimals_down(link_quantity)

                            # Determine link profitability
                            buy_price = link_quantity * buy.usd_spot
                            sell_price = link_quantity * sell.usd_spot
                            profit = sell_price - buy_price

                            # if the link is less than $1.00 skip it
                            if abs(profit) < 1.0:
                                continue
                            
                            link = sell.link_transaction(buy, link_quantity)
                            links.append(link)
                            quantity_linked += link.quantity

                        if (sell.unlinked_quantity * sell.usd_spot) > min_unlinked:
                            failures.append({
                                'asset': sell.symbol, 
                                'unlinkable': sell.unlinked_quantity,
                                'quantity': sell.quantity,
                                'timestamp': sell.time_stamp,
                                'algo': algo
                            })
        
        elif algo == 'filo':
            for key in buys.keys():
                buys[key].sort(key=lambda x: x.time_stamp, reverse=True)
            
            for key in sells.keys():
                sells[key].sort(key=lambda x: x.time_stamp.replace(tzinfo=None) if hasattr(x.time_stamp, 'replace') else x.time_stamp)

            from utils import round_decimals_down
            
            keys = list(sells.keys())
            keys.sort()
            
            for key in keys:
                quantity_linked = 0.0
                links = []

                for sell in sells[key]:
                    # check if sell has remaining unlinked quantity
                    if sell.unlinked_quantity > min_unlinked:
                        for buy in buys[key]:
                            link_quantity = None

                            # break if sell has no remaining unlinked quantity
                            if sell.unlinked_quantity <= min_unlinked:
                                break

                            # Skip if buy has no remaining unlinked quantity
                            if buy.unlinked_quantity <= min_unlinked:
                                continue

                            # check if buy came before sell
                            buy_time = buy.time_stamp.replace(tzinfo=None) if hasattr(buy.time_stamp, 'replace') else buy.time_stamp
                            sell_time = sell.time_stamp.replace(tzinfo=None) if hasattr(sell.time_stamp, 'replace') else sell.time_stamp
                            
                            if buy_time >= sell_time:
                                continue

                            # Link 
                            # if sell unlinked is greater than buy unlinked, link quantity equals buy unlinked
                            if sell.unlinked_quantity >= buy.unlinked_quantity:
                                link_quantity = buy.unlinked_quantity

                            # if sell unlinked is less than buy unlinked, link quantity equals sell unlinked
                            elif sell.unlinked_quantity <= buy.unlinked_quantity: 
                                link_quantity = sell.unlinked_quantity

                            # Set max length of link 
                            link_quantity = round_decimals_down(link_quantity)

                            # Determine link profitability
                            buy_price = link_quantity * buy.usd_spot
                            sell_price = link_quantity * sell.usd_spot
                            profit = sell_price - buy_price

                            # if the link is less than 1 dollar skip it
                            if abs(profit) < 1.0:
                                continue

                            link = sell.link_transaction(buy, link_quantity)
                            links.append(link)
                            quantity_linked += link.quantity

                        if (sell.unlinked_quantity * sell.usd_spot) > min_unlinked:
                            failures.append({
                                'asset': sell.symbol, 
                                'unlinkable': sell.unlinked_quantity,
                                'quantity': sell.quantity,
                                'timestamp': sell.time_stamp,
                                'algo': algo
                            })

        elif algo in ('min_gain', 'min_gain_long'):
            for key in buys.keys():
                buys[key].sort(key=lambda x: x.usd_spot, reverse=True)
            
            for key in sells.keys():
                sells[key].sort(key=lambda x: x.time_stamp.replace(tzinfo=None) if hasattr(x.time_stamp, 'replace') else x.time_stamp)

            from utils import round_decimals_down
            
            keys = list(sells.keys())
            keys.sort()
            
            for key in keys:
                for sell in sells[key]:
                    if sell.unlinked_quantity > min_unlinked:
                        for buy in buys[key]:
                            link_quantity = None

                            if sell.unlinked_quantity <= min_unlinked:
                                break

                            if buy.unlinked_quantity <= min_unlinked:
                                continue

                            buy_time = buy.time_stamp.replace(tzinfo=None) if hasattr(buy.time_stamp, 'replace') else buy.time_stamp
                            sell_time = sell.time_stamp.replace(tzinfo=None) if hasattr(sell.time_stamp, 'replace') else sell.time_stamp
                            
                            if buy_time >= sell_time:
                                continue

                            if algo == 'min_gain_long' and (sell_time - buy_time).days <= 365:
                                continue

                            if sell.unlinked_quantity >= buy.unlinked_quantity:
                                link_quantity = buy.unlinked_quantity
                            elif sell.unlinked_quantity <= buy.unlinked_quantity: 
                                link_quantity = sell.unlinked_quantity

                            link_quantity = round_decimals_down(link_quantity)

                            buy_price = link_quantity * buy.usd_spot
                            sell_price = link_quantity * sell.usd_spot
                            profit = sell_price - buy_price

                            if abs(profit) < 1.0:
                                continue

                            sell.link_transaction(buy, link_quantity)

                        if (sell.unlinked_quantity * sell.usd_spot) > min_unlinked:
                            failures.append({
                                'asset': sell.symbol, 
                                'unlinkable': sell.unlinked_quantity,
                                'quantity': sell.quantity,
                                'timestamp': sell.time_stamp,
                                'algo': algo
                            })

        # Update transactions after linking
        for trans in self.transactions:
            trans.update_linked_transactions()
            trans.set_multi_link()
            
        return failures

if __name__ == "__main__":
    transactions = Transactions()
    asset = "BTC"
    buys = [trans for trans in transactions if trans.symbol == asset and trans.trans_type == "buy"]
    sells = [trans for trans in transactions if trans.symbol == asset and trans.trans_type == "sell"]
    sends = [trans for trans in transactions if trans.symbol == asset and trans.trans_type == "send"]
    receives = [trans for trans in transactions if trans.symbol == asset and trans.trans_type == "receive"]

    buys.sort(key=lambda x: x.time_stamp)
    sends.sort(key=lambda x: x.time_stamp)
    receives.sort(key=lambda x: x.time_stamp)

    sent_quantity = 0.0
    received_quantity = 0.0
    bought_quantity = 0.0
    sold_quantity = 0.0
    sold_unlinked = 0.0
    bought_unlinked = 0.0

    for r in receives:
        received_quantity += r.quantity

    for send in sends:
        sent_quantity += send.quantity

    for b in buys:
        bought_quantity += b.quantity
        bought_unlinked += b.unlinked_quantity

    for s in sells:
        sold_quantity += s.quantity
        sold_unlinked += s.unlinked_quantity

    print(f"\n bought {bought_quantity} \n sent {sent_quantity} \n received {received_quantity} \n sold {sold_quantity} \n sold unlinked {sold_unlinked} \n bought unlinked {bought_unlinked}")

    transactions.auto_link(asset=None, algo='fifo')

    buys = [trans for trans in transactions if trans.symbol == asset and trans.trans_type == "buy"]
    sells = [trans for trans in transactions if trans.symbol == asset and trans.trans_type == "sell"]
    sends = [trans for trans in transactions if trans.symbol == asset and trans.trans_type == "send"]
    receives = [trans for trans in transactions if trans.symbol == asset and trans.trans_type == "receive"]

    buys.sort(key=lambda x: x.time_stamp)
    sends.sort(key=lambda x: x.time_stamp)
    receives.sort(key=lambda x: x.time_stamp)

    sent_quantity = 0.0
    received_quantity = 0.0
    bought_quantity = 0.0
    sold_quantity = 0.0
    sold_unlinked = 0.0
    bought_unlinked = 0.0

    for r in receives:
        received_quantity += r.quantity

    for send in sends:
        sent_quantity += send.quantity

    for b in buys:
        bought_quantity += b.quantity
        bought_unlinked += b.unlinked_quantity

    for s in sells:
        sold_quantity += s.quantity
        sold_unlinked += s.unlinked_quantity

    print(f"\n bought {bought_quantity} \n sent {sent_quantity} \n received {received_quantity} \n sold {sold_quantity} \n sold unlinked {sold_unlinked} \n bought unlinked {bought_unlinked}")

    filename = transactions.save()
    transactions.load(filename=filename)

    buys = [trans for trans in transactions if trans.symbol == asset and trans.trans_type == "buy"]
    sells = [trans for trans in transactions if trans.symbol == asset and trans.trans_type == "sell"]
    sends = [trans for trans in transactions if trans.symbol == asset and trans.trans_type == "send"]
    receives = [trans for trans in transactions if trans.symbol == asset and trans.trans_type == "receive"]

    buys.sort(key=lambda x: x.time_stamp)
    sends.sort(key=lambda x: x.time_stamp)
    receives.sort(key=lambda x: x.time_stamp)

    sent_quantity = 0.0
    received_quantity = 0.0
    bought_quantity = 0.0
    sold_quantity = 0.0
    sold_unlinked = 0.0
    bought_unlinked = 0.0

    for r in receives:
        received_quantity += r.quantity

    for send in sends:
        sent_quantity += send.quantity

    for b in buys:
        bought_quantity += b.quantity
        bought_unlinked += b.unlinked_quantity

    for s in sells:
        sold_quantity += s.quantity
        sold_unlinked += s.unlinked_quantity    
        print(f"\n\n bought {bought_quantity} \n sent {sent_quantity} \n received {received_quantity} \n sold {sold_quantity} \n sold unlinked {sold_unlinked} \n bought unlinked {bought_unlinked}")

@lru_cache(maxsize=1024)
def calculate_gain(sell, buy):
    return sell.usd_spot * sell.quantity - buy.usd_spot * buy.quantity













