import csv
import os
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


YEAR_RE = re.compile(r"(?:19|20)\d{2}")
EXACT_YEAR_RE = re.compile(r"^(?:19|20)\d{2}(?:\.0+)?$")
CONTEXT_YEAR_RES = (
    re.compile(
        r"\b(?:tax|filing|calendar|reporting|return)\s*(?:year|period)?\s*[:#-]?\s*((?:19|20)\d{2})\b",
        re.IGNORECASE,
    ),
    re.compile(
        r"\b((?:19|20)\d{2})\s+(?:tax|filing|calendar|reporting)\s+(?:year|return|period)\b",
        re.IGNORECASE,
    ),
)
YEAR_COLUMN_LABELS = {
    "year",
    "tax year",
    "filing year",
    "calendar year",
    "reporting year",
}

TAX_EVIDENCE_TYPES = {
    "filed_return": "Filed return",
    "form_8949": "Form 8949",
    "schedule_d": "Schedule D",
    "payment_receipt": "Payment receipt",
    "account_transcript": "Tax account transcript",
    "crypto_workbook": "Crypto workbook",
    "broker_form": "Broker form",
    "transaction_csv": "Transaction CSV",
    "estimate": "Estimate",
    "zero_confirmation": "Not applicable / zero confirmed",
    "other": "Other tax evidence",
}

TAX_EVIDENCE_TYPE_CHOICES = [
    ("auto", "Auto-detect"),
    ("filed_return", TAX_EVIDENCE_TYPES["filed_return"]),
    ("form_8949", TAX_EVIDENCE_TYPES["form_8949"]),
    ("schedule_d", TAX_EVIDENCE_TYPES["schedule_d"]),
    ("payment_receipt", TAX_EVIDENCE_TYPES["payment_receipt"]),
    ("account_transcript", TAX_EVIDENCE_TYPES["account_transcript"]),
    ("crypto_workbook", TAX_EVIDENCE_TYPES["crypto_workbook"]),
    ("broker_form", TAX_EVIDENCE_TYPES["broker_form"]),
    ("transaction_csv", TAX_EVIDENCE_TYPES["transaction_csv"]),
    ("estimate", TAX_EVIDENCE_TYPES["estimate"]),
    ("zero_confirmation", TAX_EVIDENCE_TYPES["zero_confirmation"]),
    ("other", TAX_EVIDENCE_TYPES["other"]),
]


def _normalized_text(*values):
    text = " ".join(str(value or "") for value in values)
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def normalize_tax_evidence_type(value):
    value = str(value or "auto").strip().lower()
    if value in TAX_EVIDENCE_TYPES:
        return value
    return "auto"


def tax_evidence_type_label(value):
    return TAX_EVIDENCE_TYPES.get(normalize_tax_evidence_type(value), TAX_EVIDENCE_TYPES["other"])


def infer_tax_evidence_year(*values):
    for value in values:
        matches = YEAR_RE.findall(str(value or ""))
        for match in matches:
            year = int(match)
            if 2009 <= year <= 2100:
                return year

    return None


def infer_tax_evidence_years_from_file(path, max_rows=300, max_sheets=8):
    """Find tax years in supported local evidence without interpreting money values."""
    path = Path(path)
    years = set()

    def add_unqualified_years(value):
        for match in YEAR_RE.findall(str(value or "")):
            year = int(match)
            if 2009 <= year <= 2100:
                years.add(year)

    def add_exact_year(value):
        text = str(value or "").strip()
        if not EXACT_YEAR_RE.fullmatch(text):
            return
        year = int(float(text))
        if 2009 <= year <= 2100:
            years.add(year)

    def add_context_years(value):
        text = str(value or "")
        for pattern in CONTEXT_YEAR_RES:
            for match in pattern.findall(text):
                year = int(match)
                if 2009 <= year <= 2100:
                    years.add(year)

    def add_table_years(rows):
        year_columns = set()
        for row_index, row in enumerate(rows):
            if row_index >= max_rows:
                break
            values = list(row or [])
            for column_index, value in enumerate(values):
                normalized = _normalized_text(value)
                if normalized in YEAR_COLUMN_LABELS:
                    year_columns.add(column_index)
                add_context_years(value)
            for column_index in year_columns:
                if column_index < len(values):
                    add_exact_year(values[column_index])

    # A year in the file name is a useful label. A parent directory is accepted
    # only when its entire name is the year, so a scan/run date does not create a
    # false filing requirement.
    add_unqualified_years(path.name)
    add_exact_year(path.parent.name)
    try:
        if path.suffix.lower() == ".csv":
            with path.open(newline="", encoding="utf-8-sig", errors="replace") as file:
                add_table_years(csv.reader(file))
        elif path.suffix.lower() == ".txt":
            with path.open(encoding="utf-8", errors="replace") as file:
                for row_index, line in enumerate(file):
                    if row_index >= max_rows:
                        break
                    add_context_years(line)
        elif path.suffix.lower() == ".xlsx":
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                for sheet in workbook.worksheets[:max_sheets]:
                    add_context_years(sheet.title)
                    add_table_years(sheet.iter_rows(values_only=True))
            finally:
                workbook.close()
    except (OSError, ValueError, csv.Error, BadZipFile, InvalidFileException):
        pass

    return sorted(years)


def classify_tax_evidence(reference="", notes="", selected_type="auto"):
    selected_type = normalize_tax_evidence_type(selected_type)
    if selected_type != "auto":
        return selected_type

    text = _normalized_text(Path(str(reference or "")).name, reference, notes)

    suffix = Path(str(reference or "")).suffix.lower()

    if any(term in text for term in ("estimate", "estimated", "draft", "projection")):
        return "estimate"
    if any(term in text for term in ("account transcript", "record of account", "tax transcript")):
        return "account_transcript"
    if any(term in text for term in ("workbook", "worksheet", "crypto tax", "crypto taxes", "cointracker", "koinly", "taxbit", "spreadsheet")):
        return "crypto_workbook"
    if any(term in text for term in ("payment receipt", "payment confirmation", "eftps", "direct pay", "irs payment", "amount paid")):
        return "payment_receipt"
    if "8949" in text:
        return "form_8949"
    if "schedule d" in text or "sched d" in text or "scheduled" in text:
        return "schedule_d"
    if any(term in text for term in ("1099", "broker", "consolidated")):
        return "broker_form"
    if any(term in text for term in ("zero", "not applicable", "no crypto", "none reported", "n a")):
        return "zero_confirmation"
    if any(term in text for term in ("transaction", "transactions", "coinbase", "cash app", "cashapp", "exchange export")):
        return "transaction_csv"
    if any(term in text for term in ("1040", "return", "filed", "tax return")):
        return "filed_return"
    if suffix == ".csv":
        return "transaction_csv"
    if suffix in (".xlsx", ".xls"):
        return "crypto_workbook"

    return "other"


def evidence_display_name(record):
    label = str(record.get("evidence_label") or "").strip()
    if label:
        return label

    evidence_path = str(record.get("evidence_path") or "").strip()
    if evidence_path:
        return os.path.basename(evidence_path)

    return "Evidence item"


def evidence_records_by_year(transactions):
    by_year = defaultdict(list)
    unassigned = []
    for record in getattr(transactions, "tax_evidence_records", []) or []:
        year = record.get("year")
        if year in (None, ""):
            unassigned.append(record)
            continue
        by_year[int(year)].append(record)

    return by_year, unassigned


def _evidence_labels(records, evidence_types):
    labels = [
        evidence_display_name(record)
        for record in records
        if record.get("evidence_type") in evidence_types
    ]
    return ", ".join(labels) if labels else "Missing"


def _has_evidence(records, *evidence_types):
    return any(record.get("evidence_type") in evidence_types for record in records)


def _record_has_all_totals(record):
    return bool(
        record
        and record.get("reported_proceeds") is not None
        and record.get("reported_cost_basis") is not None
        and record.get("reported_gain_loss") is not None
    )


def _tax_record_needs_research(record):
    return bool(record and "research" in str(record.get("filing_status") or "").lower())


def _record_has_confirmed_totals(record):
    return _record_has_all_totals(record) and not _tax_record_needs_research(record)


def _found_summary(records, tax_record, calculated_rows):
    found = []
    if calculated_rows > 0:
        found.append("Gainz calculated Form 8949-style totals")
    if _record_has_confirmed_totals(tax_record):
        found.append("filed totals entered")
    elif _tax_record_needs_research(tax_record):
        found.append("filed totals marked needs research")
    if records:
        evidence_counts = defaultdict(int)
        for record in records:
            evidence_counts[record.get("evidence_type", "other")] += 1
        for evidence_type, count in sorted(evidence_counts.items(), key=lambda item: tax_evidence_type_label(item[0])):
            label = tax_evidence_type_label(evidence_type)
            found.append(f"{count} {label.lower()} item{'s' if count != 1 else ''}")

    return "; ".join(found) if found else "No tax evidence recorded yet"


def _needs_for_status(year, status, records, tax_record):
    if status == "Ready":
        return "Review and keep evidence with the audit packet."
    if status == "Needs research":
        return f"{year}: review the source evidence and enter filed totals only after they are confirmed from official filing records."
    if status == "Not applicable / zero confirmed":
        return "Keep the zero/not-applicable confirmation with the year record."
    if status == "Open tax year - filing not due":
        return f"{year}: continue reconciling current activity; filed-return totals are not required until this tax year is filed."
    if status == "Return found, crypto totals not found":
        return f"{year}: confirm no crypto was reported, or upload/record Form 8949, Schedule D, or filed crypto totals."
    if status == "Workbook totals found, filed return missing":
        return f"{year}: upload/record the filed return PDF and enter filed totals from the official return."
    if status == "Payment evidence missing":
        return f"{year}: record tax paid/refund evidence, or note why no separate crypto payment evidence applies."
    if status == "Estimate only":
        return f"{year}: replace the estimate with filed return evidence when the year is filed."
    if status == "Needs filed totals":
        return f"{year}: enter filed proceeds, cost basis, and gain/loss or upload evidence showing no crypto totals apply."

    return f"{year}: review the year and add the missing evidence or user confirmation."


def _legacy_evidence_records(tax_record):
    if not tax_record or not tax_record.get("evidence_reference"):
        return []

    reference = tax_record.get("evidence_reference", "")
    notes = tax_record.get("notes", "")
    text = _normalized_text(reference, notes)
    evidence_types = []

    if any(term in text for term in ("1040", "tax return", "filed return", "filed form 8949")):
        evidence_types.append("filed_return")
    if any(term in text for term in (
        "payment record",
        "payment receipt",
        "payment confirmation",
        "eftps",
        "direct pay",
        "irs payment",
        "amount paid",
    )):
        evidence_types.append("payment_receipt")
    if any(term in text for term in ("account transcript", "record of account", "tax transcript")):
        evidence_types.append("account_transcript")
    if "8949" in text:
        evidence_types.append("form_8949")
    if "schedule d" in text or "sched d" in text:
        evidence_types.append("schedule_d")
    if any(term in text for term in ("workbook", "worksheet", "crypto tax", "crypto taxes")):
        evidence_types.append("crypto_workbook")

    if not evidence_types:
        evidence_types.append(classify_tax_evidence(reference, notes))

    return [
        {
            "evidence_id": f"legacy-{tax_record.get('year')}-{evidence_type}",
            "year": tax_record.get("year"),
            "evidence_type": evidence_type,
            "evidence_label": reference,
            "evidence_path": "",
            "notes": "Legacy evidence reference from filed totals record.",
            "updated_at": tax_record.get("updated_at", ""),
        }
        for evidence_type in sorted(set(evidence_types))
    ]


def _year_inventory_status(year, calculated, tax_record, records, alignment_row):
    has_filed_return = _has_evidence(records, "filed_return")
    has_payment = _has_evidence(records, "payment_receipt")
    has_workbook = _has_evidence(records, "crypto_workbook")
    has_estimate = _has_evidence(records, "estimate")
    has_zero_confirmation = _has_evidence(records, "zero_confirmation")
    has_filing_detail = _has_evidence(records, "form_8949", "schedule_d", "broker_form")
    has_totals = _record_has_confirmed_totals(tax_record)
    has_payment_record = bool(tax_record and tax_record.get("tax_paid") is not None)

    filing_status = str((tax_record or {}).get("filing_status") or "").strip().lower()
    explicitly_filed = any(term in filing_status for term in ("filed", "amended", "accepted"))
    if int(year) >= datetime.now().year and not explicitly_filed:
        return "Open tax year - filing not due", "status-verified"

    if _tax_record_needs_research(tax_record):
        return "Needs research", "status-needs-review"
    if has_zero_confirmation and calculated["rows"] == 0 and not has_totals:
        return "Not applicable / zero confirmed", "status-verified"
    if has_estimate and not has_filed_return and not has_totals:
        return "Estimate only", "status-unlinked-sales"
    if has_workbook and not has_filed_return:
        return "Workbook totals found, filed return missing", "status-needs-review"
    if has_filed_return and not (has_totals or has_filing_detail or has_workbook):
        return "Return found, crypto totals not found", "status-needs-review"
    if not has_totals:
        return "Needs filed totals", "status-needs-declared-holdings"
    if alignment_row and alignment_row.get("status") not in ("Aligned", "Ready"):
        return alignment_row["status"], alignment_row["status_class"]
    if not (has_payment or has_payment_record):
        return "Payment evidence missing", "status-unlinked-sales"
    if not has_filed_return:
        return "Needs filed return evidence", "status-needs-review"

    return "Ready", "status-verified"


def get_tax_evidence_inventory_summary(transactions, alignment=None):
    from utils import currency, get_form_8949_totals_by_year

    totals_by_year = get_form_8949_totals_by_year(transactions)
    reliability_failed = any(
        str(getattr(transaction, "input_reliability_status", "") or "").upper() == "BLOCKING"
        for transaction in getattr(transactions, "transactions", []) or []
    )
    tax_records_by_year = {
        int(record["year"]): record
        for record in getattr(transactions, "tax_year_records", []) or []
        if record.get("year") is not None
    }
    evidence_by_year, unassigned = evidence_records_by_year(transactions)
    alignment_rows = {
        int(row["year"]): row
        for row in (alignment or {}).get("rows", [])
    }
    for year, tax_record in tax_records_by_year.items():
        if not evidence_by_year.get(year):
            evidence_by_year[year].extend(_legacy_evidence_records(tax_record))

    years = set(totals_by_year) | set(tax_records_by_year) | set(evidence_by_year)

    rows = []
    for year in sorted(years, reverse=True):
        calculated = totals_by_year.get(year, {
            "total": {"rows": 0, "proceeds": 0.0, "cost_basis": 0.0, "gain_loss": 0.0},
        })["total"]
        tax_record = tax_records_by_year.get(year)
        records = evidence_by_year.get(year, [])
        status, status_class = _year_inventory_status(
            year,
            calculated,
            tax_record,
            records,
            alignment_rows.get(year),
        )
        if reliability_failed:
            status = "Inputs not reliable"
            status_class = "status-needs-review"
            next_action = (
                "Correct source rows that failed quantity/value consistency checks before "
                "reviewing calculated totals against tax evidence."
            )
        else:
            next_action = _needs_for_status(year, status, records, tax_record)
        rows.append({
            "year": year,
            "calculated_rows": calculated["rows"],
            "calculated_totals": "Suppressed until source input integrity is restored" if reliability_failed else (
                f"{currency(calculated['gain_loss'])} gain/loss from "
                f"{currency(calculated['proceeds'])} proceeds and {currency(calculated['cost_basis'])} basis"
            ),
            "filed_return_evidence": _evidence_labels(records, {"filed_return", "account_transcript"}),
            "payment_evidence": _evidence_labels(records, {"payment_receipt"}),
            "crypto_total_evidence": _evidence_labels(records, {"form_8949", "schedule_d", "crypto_workbook", "broker_form", "estimate"}),
            "status": status,
            "status_class": status_class,
            "next_action": next_action,
            "what_gainz_found": _found_summary(records, tax_record, calculated["rows"]),
            "what_gainz_needs": next_action,
            "evidence_count": len(records),
        })

    status_counts = defaultdict(int)
    for row in rows:
        status_counts[row["status"]] += 1

    review_rows = [
        row
        for row in rows
        if row["status"] not in ("Ready", "Not applicable / zero confirmed", "Open tax year - filing not due")
    ]
    evidence_records = []
    for record in getattr(transactions, "tax_evidence_records", []) or []:
        display_record = dict(record)
        display_record["evidence_type_label"] = tax_evidence_type_label(record.get("evidence_type"))
        display_record["display_name"] = evidence_display_name(record)
        evidence_records.append(display_record)

    return {
        "rows": rows,
        "evidence_records": sorted(
            evidence_records,
            key=lambda record: (
                -(int(record["year"]) if record.get("year") else 0),
                record["evidence_type_label"],
                record["display_name"],
            ),
        ),
        "unassigned_records": unassigned,
        "status_counts": dict(status_counts),
        "review_rows": review_rows,
        "metrics": {
            "years": len(rows),
            "ready_years": status_counts.get("Ready", 0),
            "years_needing_review": len(review_rows),
            "evidence_items": len(getattr(transactions, "tax_evidence_records", []) or []),
            "unassigned_items": len(unassigned),
        },
        "next_action": (
            review_rows[0]["next_action"]
            if review_rows
            else "Tax evidence inventory is ready for review."
        ),
    }
