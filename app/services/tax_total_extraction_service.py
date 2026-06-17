import csv
import os
import re
from pathlib import Path

from openpyxl import load_workbook

from app.services.tax_evidence_service import (
    tax_evidence_type_label,
)
from utils import currency, parse_float_value


MAX_TEXT_ROWS = 250
MAX_WORKBOOK_ROWS = 250
MAX_WORKBOOK_SHEETS = 8

MONEY_RE = re.compile(r"\(?-?\$?\s*[0-9][0-9,]*(?:\.[0-9]{1,2})?\)?")
YEAR_RE = re.compile(r"(?:19|20)\d{2}")

FIELD_KEYWORDS = {
    "reported_proceeds": (
        "proceeds",
        "sales proceeds",
        "gross proceeds",
        "total proceeds",
        "amount realized",
    ),
    "reported_cost_basis": (
        "cost basis",
        "basis",
        "cost or other basis",
        "total cost",
        "adjusted basis",
    ),
    "reported_gain_loss": (
        "gain/loss",
        "gain or loss",
        "capital gain",
        "net gain",
        "net loss",
        "gain loss",
        "short term gain",
        "long term gain",
    ),
    "tax_paid": (
        "tax paid",
        "payment",
        "amount paid",
        "direct pay",
        "eftps",
        "balance paid",
    ),
}

HIGH_VALUE_TYPES = {"form_8949", "schedule_d", "crypto_workbook", "payment_receipt"}


def _empty_candidate(record):
    evidence_path = str(record.get("evidence_path") or "")
    evidence_label = record.get("evidence_label") or (Path(evidence_path).name if evidence_path else "Evidence item")
    return {
        "year": record.get("year"),
        "evidence_id": record.get("evidence_id", ""),
        "source_label": evidence_label,
        "source_path": evidence_path,
        "evidence_type": record.get("evidence_type", "other"),
        "evidence_type_label": tax_evidence_type_label(record.get("evidence_type")),
        "reported_proceeds": None,
        "reported_cost_basis": None,
        "reported_gain_loss": None,
        "tax_paid": None,
        "reported_proceeds_display": "",
        "reported_cost_basis_display": "",
        "reported_gain_loss_display": "",
        "tax_paid_display": "",
        "confidence": "Low",
        "confidence_class": "status-needs-review",
        "notes": "",
        "_conflict_counts": {},
        "matched_fields": [],
    }


def _normalize_text(value):
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _money_from_text(value):
    text = str(value or "")
    matches = MONEY_RE.findall(text)
    if not matches:
        return None

    return _parse_money(matches[-1])


def _parse_money(value):
    if value is None:
        return None

    text = str(value).strip()
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()").replace("$", "").replace(",", "").replace(" ", "")
    parsed = parse_float_value(text)
    if parsed is None:
        return None
    return -parsed if negative else parsed


def _numbers_from_row(row):
    numbers = []
    for cell in row:
        if isinstance(cell, (int, float)) and not isinstance(cell, bool):
            numbers.append(float(cell))
            continue
        parsed = _money_from_text(cell)
        if parsed is not None:
            numbers.append(parsed)
    return numbers


def _row_year_matches(row, year):
    if year in (None, ""):
        return True

    text = " ".join(str(cell or "") for cell in row)
    years = {int(match) for match in YEAR_RE.findall(text)}
    return not years or int(year) in years


def _merge_candidate_value(candidate, field, value, source_note):
    if value is None:
        return

    current = candidate.get(field)
    if current is None:
        candidate[field] = value
        candidate["matched_fields"].append(field)
        return

    if round(float(current), 2) == round(float(value), 2):
        return

    conflict_counts = candidate.setdefault("_conflict_counts", {})
    conflict_label = str(source_note or field.replace("_", " ")).strip()
    conflict_counts[conflict_label] = conflict_counts.get(conflict_label, 1) + 1


def _finalize_candidate_notes(candidate):
    notes = []
    for note in str(candidate.get("notes") or "").split(";"):
        note = note.strip()
        if note and note not in notes:
            notes.append(note)

    for label, count in sorted(candidate.get("_conflict_counts", {}).items()):
        notes.append(f"Multiple {label} values found ({count} candidates); review source.")

    candidate["notes"] = "; ".join(notes)
    candidate.pop("_conflict_counts", None)


def _scan_rows_for_totals(rows, candidate):
    year = candidate.get("year")
    header_indices = {}
    for row in rows:
        if not row:
            continue

        normalized_cells = [_normalize_text(cell) for cell in row]
        detected_indices = {}
        for index, cell_text in enumerate(normalized_cells):
            for field, keywords in FIELD_KEYWORDS.items():
                if any(keyword in cell_text for keyword in keywords):
                    detected_indices[field] = index

        if len(detected_indices) >= 2:
            header_indices = detected_indices
            continue

        if header_indices and _row_year_matches(row, year):
            for field, index in header_indices.items():
                if index < len(row):
                    _merge_candidate_value(candidate, field, _parse_money(row[index]), field.replace("_", " "))

        if not _row_year_matches(row, year):
            continue

        row_text = _normalize_text(" ".join(str(cell or "") for cell in row))
        row_numbers = _numbers_from_row(row)
        if not row_numbers:
            continue

        for field, keywords in FIELD_KEYWORDS.items():
            if any(keyword in row_text for keyword in keywords):
                matched_keyword = next((kw for kw in keywords if kw in row_text), field)
                _merge_candidate_value(candidate, field, row_numbers[-1], matched_keyword)


def _read_csv_rows(path):
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as file:
        reader = csv.reader(file)
        for index, row in enumerate(reader):
            if index >= MAX_TEXT_ROWS:
                break
            rows.append(row)
    return rows


def _read_xlsx_rows(path):
    rows = []
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_index, sheet in enumerate(workbook.worksheets):
            if sheet_index >= MAX_WORKBOOK_SHEETS:
                break
            for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_index >= MAX_WORKBOOK_ROWS:
                    break
                rows.append(list(row))
    finally:
        workbook.close()
    return rows


def _read_pdf_lines(path):
    reader_class = None
    try:
        from pypdf import PdfReader  # type: ignore

        reader_class = PdfReader
    except Exception:
        try:
            from PyPDF2 import PdfReader  # type: ignore

            reader_class = PdfReader
        except Exception:
            return [], "PDF text extraction is not available in this environment."

    lines = []
    try:
        reader = reader_class(path)
        for page in list(reader.pages)[:20]:
            text = page.extract_text() or ""
            lines.extend(text.splitlines())
    except Exception:
        return [], "Could not read PDF text; review this evidence manually."

    return [[line] for line in lines[:MAX_TEXT_ROWS]], ""


def _apply_confidence(candidate):
    matched = set(candidate["matched_fields"])
    totals_complete = {
        "reported_proceeds",
        "reported_cost_basis",
        "reported_gain_loss",
    }.issubset(matched)
    evidence_type = candidate.get("evidence_type")
    gain_consistent = False
    if totals_complete:
        expected = round(candidate["reported_proceeds"] - candidate["reported_cost_basis"], 2)
        reported = round(candidate["reported_gain_loss"], 2)
        gain_consistent = abs(expected - reported) <= max(2.0, abs(reported) * 0.02)

    if totals_complete and gain_consistent:
        candidate["confidence"] = "High"
        candidate["confidence_class"] = "status-verified"
    elif len(matched) >= 2 and evidence_type in HIGH_VALUE_TYPES:
        candidate["confidence"] = "Medium"
        candidate["confidence_class"] = "status-unlinked-sales"
    elif len(matched) >= 1:
        candidate["confidence"] = "Low"
        candidate["confidence_class"] = "status-needs-review"

    for field in ("reported_proceeds", "reported_cost_basis", "reported_gain_loss", "tax_paid"):
        value = candidate.get(field)
        candidate[f"{field}_display"] = currency(value) if value is not None else ""


def _candidate_from_record(record):
    candidate = _empty_candidate(record)
    if candidate.get("year") in (None, ""):
        return None

    path = str(record.get("evidence_path") or "")
    if not path or not os.path.exists(path):
        candidate["notes"] = "Evidence path is not available on disk."
        return None

    suffix = Path(path).suffix.lower()
    try:
        if suffix == ".csv":
            rows = _read_csv_rows(path)
        elif suffix == ".xlsx":
            rows = _read_xlsx_rows(path)
        elif suffix == ".pdf":
            rows, pdf_note = _read_pdf_lines(path)
            if pdf_note:
                candidate["notes"] = pdf_note
        else:
            candidate["notes"] = "This evidence type is listed but not parsed for totals yet."
            return None
    except Exception as exc:
        candidate["notes"] = f"Could not scan this file for totals: {exc}"
        return None

    _scan_rows_for_totals(rows, candidate)
    if not candidate["matched_fields"]:
        return None

    candidate["matched_fields"] = sorted(set(candidate["matched_fields"]))
    _finalize_candidate_notes(candidate)
    if not candidate["notes"]:
        candidate["notes"] = "Review source before confirming; extracted values are suggestions."
    _apply_confidence(candidate)
    return candidate


def _candidate_dedupe_key(candidate):
    return (
        candidate.get("year"),
        str(candidate.get("source_label") or "").strip().lower(),
        str(candidate.get("evidence_type") or "").strip().lower(),
        tuple(candidate.get("matched_fields") or []),
    )


def _split_notes(notes):
    parts = []
    for note in str(notes or "").split(";"):
        note = note.strip()
        if note and note not in parts:
            parts.append(note)
    return parts


def _merge_unique_values(candidates, field):
    values = []
    for candidate in candidates:
        value = candidate.get(field)
        if value is None:
            continue
        rounded = round(float(value), 2)
        if all(round(float(existing), 2) != rounded for existing in values):
            values.append(value)
    return values


def _merge_candidate_group(candidates):
    if len(candidates) == 1:
        return candidates[0]

    confidence_rank = {"High": 0, "Medium": 1, "Low": 2}
    ordered = sorted(
        candidates,
        key=lambda item: (
            confidence_rank.get(item.get("confidence"), 9),
            str(item.get("source_path") or ""),
            str(item.get("evidence_id") or ""),
        ),
    )
    merged = dict(ordered[0])
    merged["duplicate_count"] = len(candidates)

    notes = []
    for candidate in candidates:
        for note in _split_notes(candidate.get("notes")):
            if note not in notes:
                notes.append(note)

    for field in ("reported_proceeds", "reported_cost_basis", "reported_gain_loss", "tax_paid"):
        values = _merge_unique_values(candidates, field)
        if not values:
            merged[field] = None
            continue
        merged[field] = values[0]
        if len(values) > 1:
            label = field.replace("_", " ")
            notes.append(
                f"Multiple {label} values found across duplicate evidence rows "
                f"({len(values)} values); review source."
            )

    source_paths = []
    evidence_ids = []
    for candidate in candidates:
        source_path = str(candidate.get("source_path") or "").strip()
        if source_path and source_path not in source_paths:
            source_paths.append(source_path)
        evidence_id = str(candidate.get("evidence_id") or "").strip()
        if evidence_id and evidence_id not in evidence_ids:
            evidence_ids.append(evidence_id)

    if len(source_paths) == 1:
        merged["source_path"] = source_paths[0]
    elif source_paths:
        merged["source_path"] = "; ".join(source_paths)

    if len(evidence_ids) == 1:
        merged["evidence_id"] = evidence_ids[0]
    elif evidence_ids:
        merged["evidence_id"] = "; ".join(evidence_ids)

    notes.append(f"Merged {len(candidates)} duplicate suggested rows for this source/year/type.")
    merged["notes"] = "; ".join(note for note in notes if note)
    _apply_confidence(merged)
    return merged


def _dedupe_candidates(candidates):
    grouped = {}
    for candidate in candidates:
        grouped.setdefault(_candidate_dedupe_key(candidate), []).append(candidate)
    return [_merge_candidate_group(group) for group in grouped.values()]


def get_suggested_filed_totals(transactions):
    candidates = []
    records_by_key = {}
    for record in getattr(transactions, "tax_evidence_records", []) or []:
        key = record.get("evidence_id") or (
            record.get("year"),
            record.get("evidence_type"),
            record.get("evidence_label"),
            record.get("evidence_path"),
        )
        records_by_key[key] = record

    for record in records_by_key.values():
        candidate = _candidate_from_record(record)
        if candidate:
            candidates.append(candidate)

    confidence_rank = {"High": 0, "Medium": 1, "Low": 2}
    candidates = _dedupe_candidates(candidates)
    return sorted(
        candidates,
        key=lambda item: (
            -(int(item["year"]) if item.get("year") else 0),
            confidence_rank.get(item["confidence"], 9),
            item["source_label"],
        ),
    )
