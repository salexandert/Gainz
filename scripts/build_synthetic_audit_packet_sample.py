import csv
import datetime
import hashlib
import json
import shutil
import sys
import tempfile
import zipfile
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT))

from openpyxl import load_workbook

from app.services.audit_packet_service import AuditPacketService
from app.services.packet_plan_service import (
    CONSERVATIVE_MAX_GAIN_DISCLOSURE,
    reconciliation_work_order_rows,
)
from parsers import import_transactions
from transactions import Transactions
from utils import (
    get_audit_readiness_summary,
    get_form_8949_report_rows,
    get_form_8949_totals,
)


def empty_transactions():
    transactions = Transactions.__new__(Transactions)
    transactions.revision_num = 0
    transactions.saves = []
    transactions.index = 0
    transactions.conversions = []
    transactions.asset_objects = []
    transactions.import_warnings = []
    transactions.import_warning_reviews = []
    transactions.basis_review_notes = []
    transactions.tax_year_records = []
    transactions.tax_evidence_records = []
    transactions.work_order_reviews = []
    transactions.view = ""
    transactions.transactions = []
    transactions.saved_descriptions = []

    def fake_save(description=None):
        transactions.saved_descriptions.append(description)

    transactions.save = fake_save
    return transactions


def set_expected_holdings(transactions):
    for asset in sorted(transactions.assets):
        expected = 0.0
        for transaction in transactions.transactions:
            if transaction.symbol != asset:
                continue
            if transaction.trans_type in ("buy", "receive"):
                expected += transaction.quantity
            elif transaction.trans_type in ("sell", "send"):
                expected -= transaction.quantity
        transactions.set_holdings(asset, expected)


def normalize_demo_source_paths(transactions):
    for transaction in transactions.transactions:
        source = Path(str(transaction.source))
        try:
            transaction.source = str(source.resolve().relative_to(REPO_ROOT))
        except (OSError, ValueError):
            transaction.source = source.name or str(transaction.source)


def set_synthetic_filed_totals(transactions):
    by_year = {}
    for row in get_form_8949_report_rows(transactions):
        year = int(row["year"])
        totals = by_year.setdefault(
            year,
            {"proceeds": 0.0, "cost_basis": 0.0, "gain_loss": 0.0},
        )
        totals["proceeds"] += float(row["proceeds"])
        totals["cost_basis"] += float(row["cost_basis"])
        totals["gain_loss"] += float(row["gain_loss"])

    for year, totals in sorted(by_year.items()):
        transactions.set_tax_year_record(
            year,
            reported_proceeds=totals["proceeds"],
            reported_cost_basis=totals["cost_basis"],
            reported_gain_loss=totals["gain_loss"],
            tax_paid=max(totals["gain_loss"], 0) * 0.25,
            filing_status="Synthetic demo only",
            evidence_reference=f"Synthetic {year} filed return summary",
            notes="Synthetic public sample record. Not a real tax filing.",
        )
        transactions.set_tax_evidence_record(
            year=year,
            evidence_type="filed_return",
            evidence_label=f"Synthetic {year} filed return summary (reference-only demo)",
            notes="Synthetic reference for the public sample packet. No real evidence file is copied.",
        )


def apply_synthetic_partial_basis_resolution(transactions):
    transactions.auto_link(asset="BCH", algo="fifo")
    transactions.set_holdings("BCH", 0.0)

    readiness = get_audit_readiness_summary(transactions)
    item = next(
        row
        for row in reconciliation_work_order_rows(readiness, transactions)
        if row.get("blocker_type") == "Missing acquisition basis"
        and row.get("asset") == "BCH"
    )
    sell = next(
        transaction
        for transaction in transactions.transactions
        if transaction.uid == item["target_transaction_uid"]
    )
    quantity = float(item["quantity"])
    proceeds = sell.prorated_tax_usd(quantity)
    before = get_form_8949_totals(transactions)["total"]
    evidence_reference = "Synthetic professional workpaper BCH-2024-01"

    adjustment_buy, _link = transactions.apply_cpa_basis_resolution(
        target_sell_uid=sell.uid,
        quantity=quantity,
        acquisition_date=sell.time_stamp,
        basis_value=0.0,
        proceeds_value=proceeds,
        basis_method="Conservative unknown basis recorded as $0",
        evidence_reference=evidence_reference,
        work_order_item_id=item["item_id"],
        acquisition_date_method="cpa_conservative_short_term",
    )
    after = get_form_8949_totals(transactions)["total"]
    receipt = {
        "changes_calculations": True,
        "decision": "conservative_max_gain",
        "decision_label": "Apply conservative $0-basis short-term treatment",
        "asset": "BCH",
        "quantity": quantity,
        "source_file": Path(str(sell.source)).name,
        "source_gross": sell.prorated_gross_usd(quantity),
        "source_fee": sell.prorated_fee_usd(quantity),
        "source_net": proceeds,
        "added_proceeds": proceeds,
        "added_basis": 0.0,
        "added_gain_loss": proceeds,
        "before_proceeds": before["proceeds"],
        "before_basis": before["cost_basis"],
        "before_gain_loss": before["gain_loss"],
        "after_proceeds": after["proceeds"],
        "after_basis": after["cost_basis"],
        "after_gain_loss": after["gain_loss"],
        "actual_after_proceeds": after["proceeds"],
        "actual_after_basis": after["cost_basis"],
        "actual_after_gain_loss": after["gain_loss"],
        "term": "Short-term assumption",
        "assumption_disclosure": CONSERVATIVE_MAX_GAIN_DISCLOSURE,
        "evidence_reference": evidence_reference,
        "reviewer_name": "Synthetic Example Tax Professional",
        "direction_date": "2026-07-18",
        "direction_entered_by": "Synthetic sample builder",
        "reviewer_credential": "Synthetic example only",
        "reviewer_jurisdiction": "Demo",
    }
    transactions.set_work_order_review(
        item["item_id"],
        decision="conservative_max_gain",
        note=(
            "Synthetic demonstration only. Documented FIFO basis was preserved for 0.2 BCH; "
            "the exact unsupported 0.3 BCH remainder uses the recorded conservative treatment."
        ),
        cpa_question="Synthetic example: confirm the documented treatment before any real filing.",
        reviewer_name="Synthetic Example Tax Professional",
        reviewer_role="cpa_ea_tax_professional",
        direction_date="2026-07-18",
        direction_entered_by="Synthetic sample builder",
        reviewer_credential="Synthetic example only",
        reviewer_jurisdiction="Demo",
        event_classification="conservative_unknown_disposition",
        proceeds_method="allocated_source_value",
        proceeds_value=f"{proceeds:.2f}",
        basis_method="unknown_zero_for_review",
        basis_value="0.00",
        evidence_reference=evidence_reference,
        resolution_status="cpa_reviewed_position",
        professional_attestation="Yes",
        blocker_type=item.get("blocker_type"),
        asset=item.get("asset"),
        year=item.get("year"),
        date=item.get("date"),
        quantity=item.get("quantity"),
        transaction_quantity=item.get("transaction_quantity"),
        source_file=item.get("source_file"),
        suspected_issue=item.get("suspected_issue"),
        target_transaction_uid=item.get("target_transaction_uid"),
        acquisition_date_method="cpa_conservative_short_term",
        acquisition_date="",
        assumption_disclosure=CONSERVATIVE_MAX_GAIN_DISCLOSURE,
        calculation_applied="Yes",
        adjustment_transaction_uid=adjustment_buy.uid,
        calculation_receipt_json=json.dumps(receipt, sort_keys=True),
        resolution_applied_at="2026-07-18 12:00:00",
    )

    return receipt


def _replacement_pairs(packet_path):
    home = Path.home()
    pairs = [
        (str(packet_path), "gainz-synthetic-audit-packet-sample"),
        (str(packet_path).replace("\\", "/"), "gainz-synthetic-audit-packet-sample"),
        (str(packet_path.parent), "gainz-synthetic-audit-packet-build"),
        (str(packet_path.parent).replace("\\", "/"), "gainz-synthetic-audit-packet-build"),
        (str(REPO_ROOT), "gainz-demo-repo"),
        (str(REPO_ROOT).replace("\\", "/"), "gainz-demo-repo"),
        (str(home), "<user-home>"),
        (str(home).replace("\\", "/"), "<user-home>"),
    ]
    return pairs + [(old.replace("\\", "\\\\"), new) for old, new in pairs]


def _sanitize_text_file(path, replacements):
    text = path.read_text(encoding="utf-8")
    sanitized = text
    for old, new in replacements:
        sanitized = sanitized.replace(old, new)
    if sanitized != text:
        path.write_text(sanitized, encoding="utf-8")


def _sanitize_workbook(path, replacements):
    workbook = load_workbook(path)
    changed = False
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                value = cell.value
                for old, new in replacements:
                    value = value.replace(old, new)
                if value != cell.value:
                    cell.value = value
                    changed = True
    if changed:
        workbook.save(path)
    workbook.close()
    return changed


def sanitize_public_sample_packet(packet_path, audit_service):
    replacements = _replacement_pairs(packet_path)
    for path in packet_path.rglob("*"):
        if not path.is_file():
            continue
        if path.suffix.lower() in {".csv", ".json", ".md", ".txt"}:
            _sanitize_text_file(path, replacements)
        elif path.suffix.lower() == ".xlsx":
            _sanitize_workbook(path, replacements)

    manifest_path = packet_path / "03_manifests" / "evidence_manifest.csv"
    with open(manifest_path, newline="", encoding="utf-8") as file:
        rows = list(csv.DictReader(file))

    for row in rows:
        if row["category"] != "generated_report":
            continue
        report_path = packet_path / row["packet_relative_path"]
        report_hash = audit_service._sha256(report_path)
        row["source_path"] = row["packet_relative_path"]
        row["source_sha256"] = report_hash
        row["packet_sha256"] = report_hash
        row["size_bytes"] = str(report_path.stat().st_size)

    with open(manifest_path, "w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)

    audit_service._write_inventory(packet_path)


def build_packet_zip():
    downloads_dir = REPO_ROOT / "docs" / "assets" / "downloads"
    downloads_dir.mkdir(parents=True, exist_ok=True)
    output_zip = downloads_dir / "gainz-synthetic-audit-packet-sample.zip"

    transactions = empty_transactions()
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir)
        source_path = REPO_ROOT / "demo_data" / "coinbase_partial_basis_fee_sample.csv"
        imported_count, skipped_count = import_transactions(str(source_path), transactions)
        if imported_count != 2 or skipped_count != 0 or transactions.import_warnings:
            raise RuntimeError(
                "Synthetic partial-basis source did not import cleanly: "
                f"imported={imported_count}, skipped={skipped_count}, warnings={transactions.import_warnings}"
            )
        receipt = apply_synthetic_partial_basis_resolution(transactions)
        if round(float(receipt["actual_after_gain_loss"]), 2) != 469.50:
            raise RuntimeError("Synthetic sample must produce the fee-inclusive $469.50 gain.")
        normalize_demo_source_paths(transactions)
        set_expected_holdings(transactions)
        set_synthetic_filed_totals(transactions)

        packet_root = temp_path / "packets"
        export_root = temp_path / "exports"
        audit_service = AuditPacketService(packet_root, export_root)
        packet_path = Path(audit_service.create_packet(transactions))
        sanitize_public_sample_packet(packet_path, audit_service)

        staged_zip = temp_path / output_zip.name
        with zipfile.ZipFile(staged_zip, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for path in sorted(packet_path.rglob("*")):
                if path.is_file():
                    archive.write(path, path.relative_to(packet_path))

        shutil.copy2(staged_zip, output_zip)

    sample_hash = hashlib.sha256(output_zip.read_bytes()).hexdigest()
    version = (REPO_ROOT / "VERSION").read_text(encoding="utf-8").strip()
    metadata = {
        "version": version,
        "generated_date": datetime.date.today().isoformat(),
        "sha256": sample_hash,
        "scenario": (
            "Synthetic partial-basis BCH sale with a fee-inclusive Coinbase import, documented FIFO basis first, "
            "and an exact professional-directed conservative treatment for the unsupported remainder."
        ),
        "totals": {
            "proceeds": 495.00,
            "cost_basis": 25.50,
            "gain_loss": 469.50,
            "short_term_gain_loss": 297.00,
            "long_term_gain_loss": 172.50,
            "source_fees": 5.50,
        },
    }
    metadata_path = downloads_dir / "gainz-synthetic-audit-packet-sample.json"
    with metadata_path.open("w", encoding="utf-8", newline="\n") as metadata_file:
        metadata_file.write(json.dumps(metadata, indent=2) + "\n")

    return output_zip


if __name__ == "__main__":
    print(build_packet_zip())
