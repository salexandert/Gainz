import datetime
import os
import tempfile
import unittest
from pathlib import Path

import transactions as transactions_module
from app.auto_link.routes import _preview_auto_link_failures
from app.model.routes import _model_sale_payload
from app.services.auto_link_service import AutoLinkService
from app.services.tax_evidence_service import (
    classify_tax_evidence,
    get_tax_evidence_inventory_summary,
)
from app.services.tax_total_extraction_service import get_suggested_filed_totals
from app.services.packet_plan_service import reconciliation_work_order_rows
from app.stats.routes import _auto_fix_safe_issues
from openpyxl import Workbook
from transaction import Buy, Receive, Sell, Send
from transactions import Transactions
from utils import (
    get_audit_readiness_summary,
    get_current_holdings_lot_table_data,
    get_holdings_difference_breakdown,
    get_multi_asset_holdings_reconciliation_table_data,
    get_stats_table_data_range,
    get_tax_filing_alignment_summary,
    get_transactions_date_range,
    get_unrealized_chart_data,
)


def empty_transactions():
    transactions = Transactions.__new__(Transactions)
    transactions.revision_num = 0
    transactions.saves = []
    transactions.index = 0
    transactions.conversions = []
    transactions.asset_objects = []
    transactions.import_warnings = []
    transactions.import_warning_reviews = []
    transactions.basis_review_notes = []
    transactions.tax_year_records = []
    transactions.tax_evidence_records = []
    transactions.work_order_reviews = []
    transactions.view = ""
    transactions.transactions = []
    transactions.saved_descriptions = []

    def fake_save(description=None):
        transactions.saved_descriptions.append(description)

    transactions.save = fake_save
    return transactions


class TransactionsEngineTests(unittest.TestCase):
    def test_default_load_uses_most_recent_save_not_highest_revision(self):
        original_basedir = transactions_module.basedir
        original_load = Transactions.load

        def write_save(path, revision_num):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Description"
            sheet.cell(row=1, column=1, value=path.stem)
            sheet.cell(row=1, column=2, value=revision_num)
            workbook.save(path)
            workbook.close()

        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            saves_dir = temp_path / "saves"
            saves_dir.mkdir()

            older = saves_dir / "saved_older_high_revision.xlsx"
            newer = saves_dir / "saved_newer_low_revision.xlsx"
            write_save(older, 1000)
            write_save(newer, 1)
            os.utime(older, (100, 100))
            os.utime(newer, (200, 200))

            loaded_paths = []

            def fake_load(self, filename):
                loaded_paths.append(Path(filename).name)
                return []

            try:
                transactions_module.basedir = str(temp_path)
                Transactions.load = fake_load

                transactions = Transactions()
            finally:
                transactions_module.basedir = original_basedir
                Transactions.load = original_load

        self.assertEqual("saved_newer_low_revision.xlsx", Path(transactions.view).name)
        self.assertEqual(["saved_newer_low_revision.xlsx"], loaded_paths)

    def test_save_creates_persistent_saves_directory(self):
        original_basedir = transactions_module.basedir

        with tempfile.TemporaryDirectory() as temp_dir:
            data_dir = Path(temp_dir) / "Gainz-Windows"
            transactions = Transactions.__new__(Transactions)
            transactions.revision_num = 0
            transactions.saves = []
            transactions.index = 0
            transactions.conversions = []
            transactions.asset_objects = []
            transactions.import_warnings = []
            transactions.import_warning_reviews = []
            transactions.basis_review_notes = []
            transactions.tax_year_records = []
            transactions.tax_evidence_records = []
            transactions.work_order_reviews = []
            transactions.view = ""
            transactions.transactions = []

            try:
                transactions_module.basedir = str(data_dir)
                save_path = Transactions.save(transactions, "Packaged save path smoke test")
            finally:
                transactions_module.basedir = original_basedir

            save_path = Path(save_path)
            self.assertEqual(data_dir / "saves", save_path.parent)
            self.assertTrue(save_path.is_file())

    def test_assets_excludes_fiat_symbols_but_keeps_transactions(self):
        transactions = empty_transactions()
        transactions.transactions = [
            Receive("USD", 100, datetime.datetime(2024, 1, 1), "cash", 1),
            Buy("BTC", 1, datetime.datetime(2024, 1, 2), 100, "exchange"),
        ]

        self.assertEqual({"BTC"}, transactions.assets)
        self.assertEqual(2, len(transactions.transactions))

    def test_stats_date_range_handles_empty_transactions(self):
        transactions = empty_transactions()

        date_range = get_transactions_date_range(
            transactions,
            {"start_date": "", "end_date": ""},
        )
        stats = get_stats_table_data_range(transactions, date_range)

        self.assertIsNone(date_range["start_date"])
        self.assertIsNone(date_range["end_date"])
        self.assertEqual([], stats)

    def test_stats_all_time_range_includes_transactions(self):
        transactions = empty_transactions()
        transactions.transactions = [
            Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "exchange"),
        ]

        date_range = get_transactions_date_range(
            transactions,
            {"start_date": "", "end_date": ""},
        )
        stats = get_stats_table_data_range(transactions, date_range)

        self.assertEqual(1, len(stats))
        self.assertEqual("BTC", stats[0]["symbol"])
        self.assertEqual("1", stats[0]["total_purchased_quantity"])

    def test_current_holdings_lots_show_remaining_basis_by_acquisition_date(self):
        transactions = empty_transactions()
        first_buy = Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "exchange")
        second_buy = Buy("BTC", 2, datetime.datetime(2024, 2, 1), 200, "exchange")
        sell = Sell("BTC", 0.4, datetime.datetime(2024, 3, 1), 300, "exchange")
        sell.link_transaction(first_buy, 0.4)
        transactions.transactions = [first_buy, second_buy, sell]

        lots = get_current_holdings_lot_table_data(transactions, "BTC")

        self.assertEqual(2, len(lots))
        self.assertEqual("2024-01-01 00:00:00", lots[0][2])
        self.assertEqual("0.6", lots[0][3])
        self.assertEqual("$60.00", lots[0][6])
        self.assertEqual("2024-02-01 00:00:00", lots[1][2])
        self.assertEqual("2", lots[1][3])
        self.assertEqual("$400.00", lots[1][6])

    def test_current_holdings_lots_include_receives_and_ignore_fiat(self):
        transactions = empty_transactions()
        transactions.transactions = [
            Receive("ETH", 3, datetime.datetime(2024, 1, 1), "wallet", 50),
            Receive("USD", 100, datetime.datetime(2024, 1, 2), "cash", 1),
        ]

        lots = get_current_holdings_lot_table_data(transactions)

        self.assertEqual(1, len(lots))
        self.assertEqual("ETH", lots[0][0])
        self.assertEqual("Receive", lots[0][1])
        self.assertEqual("$150.00", lots[0][6])

    def test_declared_holdings_allocates_current_holdings_to_newest_available_lots(self):
        transactions = empty_transactions()
        transactions.set_holdings("BTC", 0.5)
        transactions.transactions = [
            Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "old"),
            Buy("BTC", 2, datetime.datetime(2024, 2, 1), 200, "new"),
        ]

        lots = get_current_holdings_lot_table_data(transactions, "BTC")

        self.assertEqual(1, len(lots))
        self.assertEqual("2024-02-01 00:00:00", lots[0][2])
        self.assertEqual("0.5", lots[0][3])
        self.assertEqual("$100.00", lots[0][6])

    def test_unrealized_chart_data_uses_current_lots_and_usd_spot(self):
        transactions = empty_transactions()
        first_buy = Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "exchange")
        second_buy = Buy("BTC", 2, datetime.datetime(2024, 2, 1), 200, "exchange")
        sell = Sell("BTC", 0.4, datetime.datetime(2024, 3, 1), 300, "exchange")
        sell.link_transaction(first_buy, 0.4)
        transactions.transactions = [first_buy, second_buy, sell]

        chart = get_unrealized_chart_data(transactions, "BTC", 300)

        self.assertEqual(300, chart["current_usd_spot"])
        self.assertEqual(2, len(chart["points"]))
        self.assertEqual("2024-01-01 00:00:00", chart["points"][0]["x"])
        self.assertEqual(120, chart["points"][0]["y"])
        self.assertEqual("$60.00", chart["points"][0]["cost_basis"])
        self.assertEqual("$180.00", chart["points"][0]["current_value"])
        self.assertEqual(200, chart["points"][1]["y"])

    def test_multi_asset_holdings_reconciliation_includes_all_crypto_assets(self):
        transactions = empty_transactions()
        transactions.transactions = [
            Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "exchange"),
            Sell("BTC", 0.25, datetime.datetime(2024, 2, 1), 200, "exchange"),
            Receive("ETH", 2, datetime.datetime(2024, 1, 1), "wallet", 50),
            Receive("USD", 100, datetime.datetime(2024, 1, 2), "cash", 1),
        ]
        transactions.transactions[1].link_transaction(transactions.transactions[0], 0.25)
        transactions.set_holdings("BTC", 0.75)

        rows = get_multi_asset_holdings_reconciliation_table_data(transactions)

        self.assertEqual(["BTC", "ETH"], [row[0] for row in rows])
        self.assertEqual("0.75", rows[0][1])
        self.assertEqual("0.75", rows[0][2])
        self.assertEqual("0", rows[0][5])
        self.assertEqual("Verified", rows[0][6])
        self.assertEqual("N/A", rows[1][1])
        self.assertEqual("Needs declared holdings", rows[1][6])

    def test_holdings_difference_breakdown_explains_running_timeline(self):
        transactions = empty_transactions()
        transactions.set_holdings("BTC", 0.5)
        transactions.transactions = [
            Buy("BTC", 2, datetime.datetime(2024, 1, 1), 100, "C:\\private\\coinbase.csv"),
            Sell("BTC", 0.25, datetime.datetime(2024, 2, 1), 200, "C:\\private\\coinbase.csv"),
            Send("BTC", 0.75, datetime.datetime(2024, 3, 1), 250, "C:\\private\\cash_app.csv"),
            Receive("BTC", 0.1, datetime.datetime(2025, 1, 1), "wallet.csv", 300),
        ]

        breakdown = get_holdings_difference_breakdown(transactions, "BTC")

        self.assertEqual("1.75", breakdown["summary"]["expected_holdings"])
        self.assertEqual("1.25", breakdown["summary"]["difference"])
        self.assertEqual("1.1", breakdown["summary"]["imported_net"])
        self.assertIn("Buys 2 - sells 0.25 = 1.75 BTC", breakdown["summary"]["expected_formula"])
        self.assertIn("Expected 1.75 - declared 0.5 = 1.25 BTC", breakdown["summary"]["difference_formula"])
        self.assertEqual(2, len(breakdown["classification_rows"]))
        self.assertEqual("Needs classification", breakdown["classification_rows"][0][5])
        self.assertIn("No nearby same-quantity receive", breakdown["classification_rows"][0][7])
        self.assertEqual("Needs source/basis", breakdown["classification_rows"][1][5])
        self.assertEqual([
            "2024",
            3,
            "2",
            "0.25",
            "0.75",
            "0",
            "1.75",
            "1.75",
            "1",
            "1",
        ], breakdown["yearly_rows"][0])
        self.assertEqual([
            "2025",
            1,
            "0",
            "0",
            "0",
            "0.1",
            "0",
            "1.75",
            "0.1",
            "1.1",
        ], breakdown["yearly_rows"][1])
        self.assertEqual(4, len(breakdown["transaction_rows"]))
        self.assertEqual("coinbase.csv", breakdown["transaction_rows"][0][9])
        self.assertEqual("0", breakdown["transaction_rows"][2][5])
        self.assertEqual("-0.75", breakdown["transaction_rows"][2][7])
        self.assertEqual("1", breakdown["transaction_rows"][2][8])

    def test_holdings_difference_breakdown_flags_possible_owner_transfer(self):
        transactions = empty_transactions()
        transactions.transactions = [
            Send("ETH", 1.5, datetime.datetime(2024, 3, 1), 2500, "exchange_a.csv"),
            Receive("ETH", 1.5, datetime.datetime(2024, 3, 3), "exchange_b.csv", 2500),
        ]

        breakdown = get_holdings_difference_breakdown(transactions, "ETH")

        self.assertEqual(2, len(breakdown["classification_rows"]))
        self.assertEqual("Possible owner transfer", breakdown["classification_rows"][0][5])
        self.assertIn("Nearby receive found 1.5", breakdown["classification_rows"][0][7])
        self.assertEqual("Possible owner transfer", breakdown["classification_rows"][1][5])
        self.assertIn("Nearby send found 1.5", breakdown["classification_rows"][1][7])

    def test_auto_link_preview_reports_failures_without_creating_links(self):
        buy = Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "test")
        covered_sell = Sell("BTC", 0.5, datetime.datetime(2024, 2, 1), 300, "test")
        early_sell = Sell("BTC", 0.5, datetime.datetime(2023, 12, 1), 300, "test")

        covered_failures = _preview_auto_link_failures("BTC", [buy], [covered_sell], "fifo")
        early_failures = _preview_auto_link_failures("BTC", [buy], [early_sell], "fifo")

        self.assertEqual([], covered_failures)
        self.assertEqual(0, len(buy.links))
        self.assertEqual(0, len(covered_sell.links))
        self.assertEqual(1, len(early_failures))
        self.assertEqual(0.5, early_failures[0]["unlinkable"])

    def test_set_holdings_creates_and_updates_asset_record(self):
        transactions = empty_transactions()

        transactions.set_holdings("btc", 0.5)
        transactions.set_holdings("BTC", 0.75)

        self.assertEqual(1, len(transactions.asset_objects))
        self.assertEqual("BTC", transactions.asset_objects[0].symbol)
        self.assertEqual(0.75, transactions.get_holdings("BTC"))

    def test_tax_filing_alignment_matches_recorded_filed_totals(self):
        transactions = empty_transactions()
        buy = Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "exchange")
        sell = Sell("BTC", 1, datetime.datetime(2024, 6, 1), 300, "exchange")
        sell.link_transaction(buy, 1)
        transactions.transactions = [buy, sell]
        transactions.set_tax_year_record(
            2024,
            reported_proceeds=300,
            reported_cost_basis=100,
            reported_gain_loss=200,
            tax_paid=50,
            filing_status="Filed",
            evidence_reference="2024 return and payment confirmation",
        )

        alignment = get_tax_filing_alignment_summary(transactions)

        self.assertEqual("Aligned", alignment["overall_status"])
        self.assertEqual(1, alignment["metrics"]["aligned_years"])
        self.assertEqual("Aligned", alignment["rows"][0]["status"])
        self.assertEqual("$0.00", alignment["rows"][0]["difference_gain_loss_display"])
        self.assertEqual("$50.00", alignment["rows"][0]["tax_paid_display"])

    def test_tax_filing_alignment_flags_unlinked_sales_before_matching(self):
        transactions = empty_transactions()
        transactions.transactions = [
            Sell("BTC", 1, datetime.datetime(2024, 6, 1), 300, "exchange")
        ]
        transactions.set_tax_year_record(
            2024,
            reported_proceeds=300,
            reported_cost_basis=100,
            reported_gain_loss=200,
            tax_paid=50,
        )

        alignment = get_tax_filing_alignment_summary(transactions)

        self.assertEqual("Needs review", alignment["overall_status"])
        self.assertEqual("Needs basis review", alignment["rows"][0]["status"])
        self.assertEqual(1, alignment["rows"][0]["unlinked_sell_count"])

    def test_tax_evidence_inventory_classifies_filenames_and_specific_year_statuses(self):
        transactions = empty_transactions()
        buy = Buy("BTC", 1, datetime.datetime(2023, 1, 1), 100, "exchange")
        sell = Sell("BTC", 1, datetime.datetime(2023, 6, 1), 300, "exchange")
        sell.link_transaction(buy, 1)
        transactions.transactions = [buy, sell]
        transactions.set_tax_year_record(
            2023,
            reported_proceeds=300,
            reported_cost_basis=100,
            reported_gain_loss=200,
            tax_paid=50,
            filing_status="Filed",
        )
        transactions.set_tax_evidence_record(
            year=2023,
            evidence_type=classify_tax_evidence("2023_filed_return.pdf"),
            evidence_label="2023_filed_return.pdf",
        )
        transactions.set_tax_evidence_record(
            year=2023,
            evidence_type=classify_tax_evidence("2023_payment_confirmation.pdf"),
            evidence_label="2023_payment_confirmation.pdf",
        )
        transactions.set_tax_evidence_record(
            year=2022,
            evidence_type=classify_tax_evidence("2022_crypto_tax_workbook.xlsx"),
            evidence_label="2022_crypto_tax_workbook.xlsx",
        )
        transactions.set_tax_evidence_record(
            year=2020,
            evidence_type=classify_tax_evidence("2020_filed_return.pdf"),
            evidence_label="2020_filed_return.pdf",
            notes="Virtual currency yes, crypto totals not found.",
        )
        transactions.set_tax_evidence_record(
            year=2025,
            evidence_type=classify_tax_evidence("2025_estimate_only.xlsx"),
            evidence_label="2025_estimate_only.xlsx",
        )

        inventory = get_tax_evidence_inventory_summary(transactions)
        statuses = {row["year"]: row["status"] for row in inventory["rows"]}

        self.assertEqual("Ready", statuses[2023])
        self.assertEqual("Workbook totals found, filed return missing", statuses[2022])
        self.assertEqual("Return found, crypto totals not found", statuses[2020])
        self.assertEqual("Estimate only", statuses[2025])
        self.assertEqual(3, inventory["metrics"]["years_needing_review"])
        needs_by_year = {row["year"]: row["what_gainz_needs"] for row in inventory["rows"]}
        self.assertIn("2022: upload/record the filed return PDF", needs_by_year[2022])

    def test_suggested_filed_totals_extracts_reviewable_values_from_csv(self):
        transactions = empty_transactions()
        with tempfile.TemporaryDirectory() as temp_dir:
            evidence_path = Path(temp_dir) / "2024_crypto_tax_workbook.csv"
            evidence_path.write_text(
                "Year,Reported Proceeds,Reported Cost Basis,Reported Gain Loss,Tax Paid\n"
                "2024,300.00,100.00,200.00,25.00\n",
                encoding="utf-8",
            )
            transactions.set_tax_evidence_record(
                year=2024,
                evidence_type="crypto_workbook",
                evidence_label=evidence_path.name,
                evidence_path=str(evidence_path),
            )

            suggestions = get_suggested_filed_totals(transactions)

        self.assertEqual(1, len(suggestions))
        suggestion = suggestions[0]
        self.assertEqual(2024, suggestion["year"])
        self.assertEqual(300.0, suggestion["reported_proceeds"])
        self.assertEqual(100.0, suggestion["reported_cost_basis"])
        self.assertEqual(200.0, suggestion["reported_gain_loss"])
        self.assertEqual(25.0, suggestion["tax_paid"])
        self.assertEqual("High", suggestion["confidence"])

    def test_suggested_filed_totals_deduplicates_repeated_conflict_notes(self):
        transactions = empty_transactions()
        with tempfile.TemporaryDirectory() as temp_dir:
            evidence_path = Path(temp_dir) / "2024_crypto_tax_workbook.csv"
            evidence_path.write_text(
                "Year,Reported Proceeds,Reported Cost Basis,Reported Gain Loss\n"
                "2024,300.00,100.00,200.00\n"
                "2024,301.00,100.00,201.00\n"
                "2024,302.00,100.00,202.00\n",
                encoding="utf-8",
            )
            transactions.set_tax_evidence_record(
                year=2024,
                evidence_type="crypto_workbook",
                evidence_label=evidence_path.name,
                evidence_path=str(evidence_path),
            )

            suggestions = get_suggested_filed_totals(transactions)

        self.assertEqual(1, len(suggestions))
        notes = suggestions[0]["notes"]
        self.assertEqual(1, notes.count("Multiple reported proceeds values found"))
        self.assertEqual(1, notes.count("Multiple reported gain loss values found"))
        self.assertIn("3 candidates", notes)

    def test_suggested_filed_totals_deduplicates_rows_by_source_year_type_and_fields(self):
        transactions = empty_transactions()
        with tempfile.TemporaryDirectory() as temp_dir:
            evidence_path = Path(temp_dir) / "2024_crypto_tax_workbook.csv"
            evidence_path.write_text(
                "Year,Reported Proceeds,Reported Cost Basis,Reported Gain Loss\n"
                "2024,300.00,100.00,200.00\n",
                encoding="utf-8",
            )
            for evidence_id in ("duplicate-a", "duplicate-b"):
                transactions.set_tax_evidence_record(
                    year=2024,
                    evidence_type="crypto_workbook",
                    evidence_label=evidence_path.name,
                    evidence_path=str(evidence_path),
                    evidence_id=evidence_id,
                )

            suggestions = get_suggested_filed_totals(transactions)

        self.assertEqual(1, len(suggestions))
        suggestion = suggestions[0]
        self.assertEqual(2, suggestion["duplicate_count"])
        self.assertEqual(2, suggestion["combined_suggestions_count"])
        self.assertEqual("$300.00", suggestion["reported_proceeds_display"])
        self.assertIn("Similar suggestions were combined for this source.", suggestion["notes"])
        self.assertNotIn("Merged 2 duplicate suggested rows", suggestion["notes"])

    def test_audit_readiness_distinguishes_reviewed_blocking_import_warnings(self):
        transactions = empty_transactions()
        warning = "Imported row 299 from cash_app_report.csv with $0 USD spot price."
        transactions.import_warnings = [warning]
        transactions.set_import_warning_review(
            warning,
            decision="needs_manual_usd_value",
            note="Needs source USD value.",
        )

        readiness = get_audit_readiness_summary(transactions)

        self.assertIn(
            "1 import warning reviewed but still blocking: Needs manual USD value.",
            readiness["warnings"],
        )
        self.assertNotIn("Choose review decisions for 1 import warning.", readiness["warnings"])
        import_warning_group = next(
            group for group in readiness["blocker_groups"]
            if group["key"] == "import_warnings"
        )
        self.assertEqual("Reviewed blocker", import_warning_group["status"])
        self.assertIn("reviewed but still blocking", import_warning_group["detail"])
        checklist = {item["label"]: item for item in readiness["checklist"]}
        self.assertTrue(checklist["Import warning decisions recorded"]["complete"])
        self.assertFalse(checklist["Import warning blockers resolved"]["complete"])
        self.assertIn(
            "Needs manual USD value",
            checklist["Import warning blockers resolved"]["detail"],
        )

    def test_reconciliation_work_order_prioritizes_actionable_blockers(self):
        readiness = {
            "is_ready": False,
            "missing_records": {
                "current_holdings": [
                    {"asset": "ETH", "message": "Current holdings missing."},
                ],
                "holdings_explanations": [
                    {"asset": "BTC", "message": "Holdings explanation needed."},
                ],
                "basis": [
                    {
                        "asset": "BCH",
                        "date": "2019-01-07",
                        "source": "coinbase.csv",
                        "message": "This sale needs earlier acquisition basis.",
                        "status": "Needs research",
                    },
                ],
                "source_overlaps": [
                    {
                        "source_a": "Coinbase_All.csv",
                        "source_b": "Coinbase_2025.csv",
                        "message": "Possible overlapping source exports.",
                        "next_action": "Review duplicate coverage.",
                        "status": "Review",
                    },
                ],
                "filed_totals": [
                    {
                        "year": 2022,
                        "status": "Workbook totals found, filed return missing",
                        "what_gainz_found": "Workbook found.",
                        "what_gainz_needs": "Upload filed return.",
                    },
                ],
            },
            "unresolved_import_warning_rows": [
                {
                    "source": "cash_app_report.csv",
                    "issue": "$0 USD spot price",
                    "decision": "needs_manual_usd_value",
                    "decision_label": "Needs manual USD value",
                    "next_action": "Map USD value.",
                },
            ],
        }

        rows = reconciliation_work_order_rows(readiness)

        self.assertEqual("Missing acquisition basis", rows[0]["blocker_type"])
        self.assertEqual("Reviewed import warning blocker", rows[1]["blocker_type"])
        self.assertTrue(rows[0]["item_id"])
        self.assertEqual(10, rows[0]["priority"])
        self.assertEqual(20, rows[1]["priority"])
        self.assertTrue(rows[0]["priority_label"].startswith("P1"))
        self.assertTrue(rows[-1]["priority_label"].startswith("P5"))

    def test_reconciliation_work_order_applies_saved_review_state(self):
        transactions = empty_transactions()
        readiness = {
            "is_ready": False,
            "missing_records": {
                "current_holdings": [],
                "holdings_explanations": [],
                "basis": [
                    {
                        "asset": "BCH",
                        "date": "2019-01-07",
                        "source": "coinbase.csv",
                        "message": "This sale needs earlier acquisition basis.",
                        "status": "Missing acquisition basis",
                    },
                ],
                "source_overlaps": [],
                "filed_totals": [],
            },
            "unresolved_import_warning_rows": [],
        }

        initial_rows = reconciliation_work_order_rows(readiness, transactions)
        transactions.set_work_order_review(
            initial_rows[0]["item_id"],
            decision="needs_research",
            note="User will investigate source records later.",
        )

        reviewed_rows = reconciliation_work_order_rows(readiness, transactions)

        self.assertEqual("needs_research", reviewed_rows[0]["review_decision"])
        self.assertEqual("Needs research", reviewed_rows[0]["review_decision_label"])
        self.assertEqual("Needs research", reviewed_rows[0]["status"])
        self.assertEqual(
            "User will investigate source records later.",
            reviewed_rows[0]["review_note"],
        )

    def test_partial_tax_year_research_record_stays_needs_research(self):
        transactions = empty_transactions()
        buy = Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "exchange")
        sell = Sell("BTC", 1, datetime.datetime(2024, 6, 1), 300, "exchange")
        sell.link_transaction(buy, 1)
        transactions.transactions = [buy, sell]
        transactions.set_tax_year_record(
            2024,
            filing_status="Needs research",
            evidence_reference="2024 return PDF",
            notes="Review source evidence before recording filed totals.",
        )

        alignment = get_tax_filing_alignment_summary(transactions)

        self.assertEqual("Needs research", alignment["rows"][0]["status"])
        self.assertIn("enter filed proceeds", alignment["rows"][0]["next_action"])

    def test_tax_year_records_round_trip_through_save_file(self):
        original_basedir = transactions_module.basedir
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            (temp_path / "saves").mkdir()
            transactions = empty_transactions()
            transactions.save = Transactions.save.__get__(transactions, Transactions)
            transactions.transactions = [
                Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "exchange")
            ]
            transactions.set_tax_year_record(
                2024,
                reported_proceeds=300,
                reported_cost_basis=100,
                reported_gain_loss=200,
                tax_paid=50,
                evidence_reference="2024 return",
            )

            try:
                transactions_module.basedir = str(temp_path)
                save_path = transactions.save(description="Tax record test")
                loaded = Transactions(save_path)
            finally:
                transactions_module.basedir = original_basedir

        record = loaded.get_tax_year_record(2024)
        self.assertIsNotNone(record)
        self.assertEqual(300.0, record["reported_proceeds"])
        self.assertEqual(50.0, record["tax_paid"])
        self.assertEqual("2024 return", record["evidence_reference"])

    def test_tax_evidence_records_round_trip_through_save_file(self):
        original_basedir = transactions_module.basedir
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            (temp_path / "saves").mkdir()
            transactions = empty_transactions()
            transactions.save = Transactions.save.__get__(transactions, Transactions)
            transactions.transactions = [
                Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "exchange")
            ]
            transactions.set_tax_evidence_record(
                year=2024,
                evidence_type="payment_receipt",
                evidence_label="2024 payment receipt",
                evidence_path=str(temp_path / "payment.pdf"),
                notes="Proof of payment kept locally.",
            )

            try:
                transactions_module.basedir = str(temp_path)
                save_path = transactions.save(description="Tax evidence record test")
                loaded = Transactions(save_path)
            finally:
                transactions_module.basedir = original_basedir

        self.assertEqual(1, len(loaded.tax_evidence_records))
        record = loaded.tax_evidence_records[0]
        self.assertEqual(2024, record["year"])
        self.assertEqual("payment_receipt", record["evidence_type"])
        self.assertEqual("2024 payment receipt", record["evidence_label"])

    def test_work_order_reviews_round_trip_through_save_file(self):
        original_basedir = transactions_module.basedir
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            (temp_path / "saves").mkdir()
            transactions = empty_transactions()
            transactions.save = Transactions.save.__get__(transactions, Transactions)
            transactions.transactions = [
                Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "exchange")
            ]
            transactions.set_work_order_review(
                "abc123",
                decision="sent_to_cpa",
                note="Asked CPA to review missing basis.",
            )

            try:
                transactions_module.basedir = str(temp_path)
                save_path = transactions.save(description="Work order review test")
                loaded = Transactions(save_path)
            finally:
                transactions_module.basedir = original_basedir

        record = loaded.get_work_order_review("abc123")
        self.assertIsNotNone(record)
        self.assertEqual("sent_to_cpa", record["decision"])
        self.assertEqual("Asked CPA to review missing basis.", record["note"])

    def test_review_decisions_round_trip_through_save_file(self):
        original_basedir = transactions_module.basedir
        warning = "Imported row 10 from exchange.csv with $0 USD spot price."
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            (temp_path / "saves").mkdir()
            transactions = empty_transactions()
            transactions.save = Transactions.save.__get__(transactions, Transactions)
            transactions.transactions = [
                Sell("LTC", 1, datetime.datetime(2024, 1, 1), 100, "exchange.csv")
            ]
            transactions.import_warnings = [warning]
            transactions.set_import_warning_review(
                warning,
                decision="needs_manual_usd_value",
                note="Find source USD value later.",
            )
            transactions.set_basis_review_note(
                "LTC",
                status="needs_research",
                note="User will investigate source records later.",
            )

            try:
                transactions_module.basedir = str(temp_path)
                save_path = transactions.save(description="Review record test")
                loaded = Transactions(save_path)
            finally:
                transactions_module.basedir = original_basedir

        self.assertEqual(
            "needs_manual_usd_value",
            loaded.get_import_warning_review(warning)["decision"],
        )
        self.assertEqual(
            "User will investigate source records later.",
            loaded.get_basis_review_note("LTC")["note"],
        )

    def test_audit_readiness_flags_missing_links_and_holdings(self):
        transactions = empty_transactions()
        transactions.transactions = [
            Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "exchange"),
            Sell("BTC", 1, datetime.datetime(2024, 2, 1), 200, "exchange"),
        ]

        readiness = get_audit_readiness_summary(transactions)

        self.assertFalse(readiness["is_ready"])
        self.assertEqual("Not ready", readiness["status"])
        self.assertEqual(1, readiness["metrics"]["assets_with_unlinked_sales"])
        self.assertEqual(1, readiness["metrics"]["assets_needing_holdings"])
        self.assertEqual(0, readiness["metrics"]["form_8949_rows"])
        self.assertTrue(any("Missing acquisition basis" in blocker for blocker in readiness["blockers"]))
        self.assertEqual(1, readiness["metrics"]["missing_basis_rows"])
        self.assertIn("BTC sale on 2024-02-01", readiness["missing_records"]["basis"][0]["message"])

    def test_audit_readiness_names_tax_evidence_inventory_blockers(self):
        transactions = empty_transactions()
        transactions.set_tax_evidence_record(
            year=2024,
            evidence_type="crypto_workbook",
            evidence_label="2024 crypto workbook",
        )

        readiness = get_audit_readiness_summary(transactions)

        self.assertFalse(readiness["is_ready"])
        self.assertEqual(1, readiness["metrics"]["tax_evidence_years_needing_review"])
        self.assertIn("Review tax evidence inventory for: 2024", readiness["blockers"])
        self.assertEqual(
            "Workbook totals found, filed return missing",
            readiness["missing_records"]["filed_totals"][0]["status"],
        )


    def test_audit_readiness_is_ready_when_links_holdings_and_filed_totals_match(self):
        transactions = empty_transactions()
        buy = Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "exchange")
        sell = Sell("BTC", 0.25, datetime.datetime(2024, 6, 1), 300, "exchange")
        sell.link_transaction(buy, 0.25)
        transactions.transactions = [buy, sell]
        transactions.set_holdings("BTC", 0.75)
        transactions.set_tax_year_record(
            2024,
            reported_proceeds=75,
            reported_cost_basis=25,
            reported_gain_loss=50,
            tax_paid=5,
            evidence_reference="2024 filed return",
        )

        readiness = get_audit_readiness_summary(transactions)

        self.assertTrue(readiness["is_ready"])
        self.assertEqual("Ready for review", readiness["status"])
        self.assertEqual(1, readiness["metrics"]["form_8949_rows"])
        self.assertEqual("$75.00", readiness["metrics"]["form_8949_proceeds"])
        self.assertEqual("$25.00", readiness["metrics"]["form_8949_cost_basis"])
        self.assertEqual("$50.00", readiness["metrics"]["form_8949_gain_loss"])

    def test_fifo_auto_link_uses_earliest_buy(self):
        transactions = empty_transactions()
        first_buy = Buy("BTC", 1, datetime.datetime(2023, 1, 1), 100, "test")
        second_buy = Buy("BTC", 1, datetime.datetime(2023, 2, 1), 200, "test")
        sell = Sell("BTC", 1, datetime.datetime(2023, 3, 1), 300, "test")
        transactions.transactions = [second_buy, sell, first_buy]

        failures = transactions.auto_link(asset="BTC", algo="fifo")

        self.assertEqual([], failures)
        self.assertEqual(first_buy.uid, sell.links[0].buy.uid)
        self.assertAlmostEqual(200, sell.links[0].profit_loss)

    def test_stats_safe_auto_fix_links_unlinked_sales_only(self):
        transactions = empty_transactions()
        buy = Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "test")
        sell = Sell("BTC", 1, datetime.datetime(2024, 2, 1), 300, "test")
        mismatch_buy = Buy("SOL", 2, datetime.datetime(2024, 1, 1), 10, "test")
        transactions.transactions = [buy, sell, mismatch_buy]
        transactions.set_holdings("BTC", 0)
        transactions.set_holdings("SOL", 0)

        payload = _auto_fix_safe_issues(transactions)
        statuses = {
            row[0]: row[6]
            for row in payload["holdings_reconciliation_table_data"]
        }

        self.assertEqual(1, payload["links_created"])
        self.assertEqual(["BTC"], payload["fixed_assets"])
        self.assertEqual("Verified", statuses["BTC"])
        self.assertEqual("Needs Review", statuses["SOL"])
        self.assertIn("SOL", payload["review_required_assets"])
        self.assertIn("Added 1 FIFO basis link", payload["message"])
        self.assertEqual(
            ["Added FIFO basis links from Stats review"],
            transactions.saved_descriptions,
        )

    def test_auto_link_service_links_all_unlinked_sales_for_selected_year(self):
        transactions = empty_transactions()
        btc_buy = Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "test")
        btc_sell = Sell("BTC", 1, datetime.datetime(2024, 2, 1), 300, "test")
        eth_buy = Buy("ETH", 1, datetime.datetime(2025, 1, 1), 100, "test")
        eth_sell = Sell("ETH", 1, datetime.datetime(2025, 2, 1), 300, "test")
        transactions.transactions = [btc_buy, btc_sell, eth_buy, eth_sell]

        result = AutoLinkService().auto_link_unlinked_sales(
            transactions,
            algo="fifo",
            year=2024,
            save_description="test guided link",
        )

        self.assertEqual(1, result["links_created"])
        self.assertEqual(["BTC"], result["fixed_assets"])
        self.assertEqual([], result["failures"])
        self.assertEqual(0, btc_sell.unlinked_quantity)
        self.assertEqual(1, eth_sell.unlinked_quantity)
        self.assertEqual(["test guided link"], transactions.saved_descriptions)

    def test_auto_link_service_normalizes_year_values(self):
        service = AutoLinkService()

        self.assertIsNone(service.selected_year(None))
        self.assertIsNone(service.selected_year(""))
        self.assertIsNone(service.selected_year(" All Time "))
        self.assertEqual(2024, service.selected_year(" 2024 "))

    def test_model_sell_defaults_to_fifo(self):
        transactions = empty_transactions()
        old_buy = Buy("BTC", 1, datetime.datetime(2020, 1, 1), 100, "old")
        new_buy = Buy("BTC", 1, datetime.datetime(2021, 1, 1), 250, "new")
        transactions.transactions = [new_buy, old_buy]

        payload = _model_sale_payload(
            transactions,
            "BTC",
            sale_usd_spot=300,
            total_in_usd=300,
        )
        fifo = payload["batches_by_key"]["fifo"]

        self.assertEqual("fifo", payload["default_batch_key"])
        self.assertEqual("old", fifo["rows"][0][0])
        self.assertEqual("$100.00", fifo["summary"]["cost_basis_display"])
        self.assertEqual("$200.00", fifo["summary"]["gain_loss_display"])

    def test_model_sell_uses_declared_current_holdings_lots(self):
        transactions = empty_transactions()
        old_buy = Buy("BTC", 1, datetime.datetime(2020, 1, 1), 100, "old")
        new_buy = Buy("BTC", 1, datetime.datetime(2021, 1, 1), 250, "new")
        transactions.transactions = [old_buy, new_buy]
        transactions.set_holdings("BTC", 0.5)

        payload = _model_sale_payload(
            transactions,
            "BTC",
            sale_usd_spot=300,
            quantity=0.5,
        )
        fifo = payload["batches_by_key"]["fifo"]

        self.assertEqual("new", fifo["rows"][0][0])
        self.assertEqual("0.5", fifo["summary"]["quantity_display"])
        self.assertEqual("$125.00", fifo["summary"]["cost_basis_display"])
        self.assertEqual("$25.00", fifo["summary"]["gain_loss_display"])

    def test_min_gain_auto_link_uses_highest_cost_basis(self):
        transactions = empty_transactions()
        low_basis_buy = Buy("ETH", 1, datetime.datetime(2023, 1, 1), 100, "test")
        high_basis_buy = Buy("ETH", 1, datetime.datetime(2023, 2, 1), 250, "test")
        sell = Sell("ETH", 1, datetime.datetime(2023, 3, 1), 300, "test")
        transactions.transactions = [low_basis_buy, sell, high_basis_buy]

        failures = transactions.auto_link(asset="ETH", algo="min_gain")

        self.assertEqual([], failures)
        self.assertEqual(high_basis_buy.uid, sell.links[0].buy.uid)
        self.assertAlmostEqual(50, sell.links[0].profit_loss)

    def test_convert_sends_to_sells_creates_sell_and_conversion_record(self):
        transactions = empty_transactions()
        send = Send("SOL", 3, datetime.datetime(2024, 1, 1), 50, "wallet")
        transactions.transactions = [send]

        message = transactions.convert_sends_to_sells("SOL", amount_to_convert=2)

        sells = [trans for trans in transactions if trans.trans_type == "sell"]
        sends = [trans for trans in transactions if trans.trans_type == "send"]
        self.assertIn("Recorded 2.0 SOL", message)
        self.assertIn("taxable disposal", message)
        self.assertEqual(1, len(sells))
        self.assertEqual(1, len(sends))
        self.assertAlmostEqual(2, sells[0].quantity)
        self.assertAlmostEqual(1, sends[0].quantity)
        self.assertEqual(1, len(transactions.conversions))

    def test_convert_receives_to_buys_creates_buy_and_reduces_receive(self):
        transactions = empty_transactions()
        receive = Receive("ADA", 10, datetime.datetime(2024, 1, 1), "wallet", 0.50)
        transactions.transactions = [receive]

        transactions.convert_receives_to_buys("ADA", amount_to_convert=4)

        buys = [trans for trans in transactions if trans.trans_type == "buy"]
        receives = [trans for trans in transactions if trans.trans_type == "receive"]
        self.assertEqual(1, len(buys))
        self.assertEqual(1, len(receives))
        self.assertAlmostEqual(4, buys[0].quantity)
        self.assertAlmostEqual(6, receives[0].quantity)
        self.assertEqual(1, len(transactions.conversions))

    def test_convert_buys_to_lost_uses_only_unlinked_quantity(self):
        transactions = empty_transactions()
        buy = Buy("MATIC", 5, datetime.datetime(2024, 1, 1), 1, "test")
        sell = Sell("MATIC", 2, datetime.datetime(2024, 2, 1), 2, "test")
        sell.link_transaction(buy, 2)
        transactions.transactions = [buy, sell]

        transactions.convert_buys_to_lost("MATIC", amount=2)

        self.assertAlmostEqual(3, buy.quantity)
        self.assertAlmostEqual(1, buy.unlinked_quantity)
        self.assertEqual(1, len(transactions.conversions))


if __name__ == "__main__":
    unittest.main()
