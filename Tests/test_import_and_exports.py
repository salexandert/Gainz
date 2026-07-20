import datetime
import csv
import io
import json
import tempfile
import unittest
import warnings
from pathlib import Path
from unittest.mock import patch

from dateutil.parser import UnknownTimezoneWarning
from openpyxl import Workbook, load_workbook

from app import create_app
from app.extensions import db
from app.import_transactions import routes as import_routes
from app.import_transactions.routes import (
    _data_source_summary,
    _manual_batch_rows,
    _manual_row_is_blank,
    _manual_transaction_from_values,
    _public_import_result,
    _remove_data_source,
)
from app.services.audit_packet_service import AuditPacketService
from app.services.import_service import ImportService
from app.services.import_warning_service import import_warning_review_rows
from app.services.tax_evidence_service import (
    classify_tax_evidence,
    get_tax_evidence_inventory_summary,
    infer_tax_evidence_years_from_file,
)
from app.services.tax_total_extraction_service import get_suggested_filed_totals
from app.services.packet_plan_service import (
    cpa_resolution_workpaper_rows,
    get_packet_preview,
    reconciliation_work_order_rows,
)
from app_version import APP_VERSION
from transaction import Buy, Sell, Send
from transactions import Transactions
from utils import (
    get_audit_readiness_summary,
    get_form_8949_report_rows,
    get_form_8949_table_data,
    get_form_8949_totals,
    get_import_economics_rows,
    get_tax_filing_alignment_summary,
)
from configs.config import config_dict
from parsers import analyze_csv_import, import_transactions, parse_quantity_value
from werkzeug.datastructures import MultiDict


def empty_transactions():
    transactions = Transactions.__new__(Transactions)
    transactions.revision_num = 0
    transactions.saves = []
    transactions.index = 0
    transactions.conversions = []
    transactions.asset_objects = []
    transactions.import_warnings = []
    transactions.import_warning_reviews = []
    transactions.basis_review_notes = []
    transactions.tax_year_records = []
    transactions.tax_evidence_records = []
    transactions.work_order_reviews = []
    transactions.import_receipts = []
    transactions.view = ""
    transactions.transactions = []
    transactions.saved_descriptions = []

    def fake_save(description=None):
        transactions.saved_descriptions.append(description)

    transactions.save = fake_save
    return transactions


class FileUpload:
    def __init__(self, source):
        self.source = Path(source)
        self.filename = self.source.name

    def save(self, destination):
        Path(destination).write_bytes(self.source.read_bytes())


def import_template_context(**overrides):
    context = {
        "manual_trans": import_routes.ManualTransaction(),
        "current_holdings": import_routes.CurrentHoldings(),
        "save_summary": {
            "revision": 0,
            "save_count": 0,
            "current_file": "Unsaved session",
            "current_name": "Unsaved session",
            "current_description": "",
            "current_modified": "N/A",
            "recent_saves": [],
        },
        "data_summary": {
            "transaction_count": 0,
            "asset_count": 0,
            "source_count": 0,
            "link_count": 0,
            "source_overlap_count": 0,
            "source_overlaps": [],
            "sources": [],
            "import_warnings": [],
            "import_warning_rows": [],
            "unresolved_import_warning_rows": [],
            "unresolved_import_warning_count": 0,
            "import_economics_count": 0,
            "import_economics_warning_count": 0,
            "import_economics_rows": [],
            "input_reliability_failure_count": 0,
            "import_receipt_count": 0,
            "import_receipts": [],
            "type_counts": {
                "buy": 0,
                "sell": 0,
                "send": 0,
                "receive": 0,
            },
        },
        "guided_import": False,
    }
    context.update(overrides)
    return context


class ImportAndExportTests(unittest.TestCase):
    def test_import_service_imports_cash_app_sample(self):
        transactions = empty_transactions()
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path("demo_data/cash_app_sample.csv")
            result = ImportService(temp_dir).import_upload(FileUpload(source), transactions)

        self.assertEqual(3, result["imported_count"])
        self.assertEqual(0, result["skipped_count"])
        self.assertEqual(["buy", "buy", "sell"], [t.trans_type for t in transactions.transactions])
        self.assertEqual({"BTC"}, transactions.assets)

    def test_import_page_get_does_not_build_heavy_unused_tables(self):
        transactions = empty_transactions()
        transactions.saves = [{
            "value": "C:/Development/Gainz/saves/saved_test.xlsx",
            "description": "Current test save",
            "revision_num": 7,
            "modified_time": 1710000000,
        }]
        transactions.view = transactions.saves[0]["value"]

        app = create_app(config_dict["Debug"], selenium=True)
        app.config.update(
            WTF_CSRF_ENABLED=False,
            transactions=transactions,
            UPLOAD_FOLDER="uploads",
        )

        with app.test_request_context("/import_transactions/"):
            with patch.object(import_routes, "render_template", return_value="ok") as render_mock:
                with patch.object(import_routes, "get_stats_table_data", side_effect=AssertionError("unused")):
                    with patch.object(import_routes, "get_all_trans_table_data", side_effect=AssertionError("unused")):
                        with patch.object(transactions, "load_saves", side_effect=AssertionError("unused")):
                            response = import_routes.import_wizard.__wrapped__()

        self.assertEqual("ok", response)
        context = render_mock.call_args.kwargs
        self.assertNotIn("stats_table_data", context)
        self.assertNotIn("transactions", context)
        self.assertEqual(7, context["save_summary"]["revision"])
        self.assertFalse(context["guided_import"])

    def test_import_page_renders_manual_batch_table(self):
        app = create_app(config_dict["Debug"], selenium=True)
        app.config.update(WTF_CSRF_ENABLED=False)

        with app.test_request_context("/import_transactions/#manual-transactions"):
            rendered_page = app.jinja_env.get_template(
                "import_transactions.html"
            ).render(**import_template_context())

        self.assertIn('id="manual_transactions_table"', rendered_page)
        self.assertIn('name="manual_type[]"', rendered_page)
        self.assertIn('name="manual_batch_submit"', rendered_page)
        self.assertIn("Blank rows are ignored", rendered_page)
        self.assertIn("<summary>Add manual rows</summary>", rendered_page)

    def test_guided_import_page_prioritizes_upload_and_collapses_operational_details(self):
        app = create_app(config_dict["Debug"], selenium=True)
        app.config.update(WTF_CSRF_ENABLED=False)

        with app.test_request_context("/import_transactions/?guided=1"):
            rendered_page = app.jinja_env.get_template(
                "import_transactions.html"
            ).render(**import_template_context(guided_import=True))

        self.assertIn("Step 1: Import Data", rendered_page)
        self.assertIn("Current task", rendered_page)
        self.assertIn("Upload source data", rendered_page)
        self.assertIn("Try demo data or upload one exchange CSV.", rendered_page)
        self.assertIn("Try Missing-Basis Demo", rendered_page)
        self.assertIn("/import_transactions/demo_missing_basis", rendered_page)
        self.assertIn("Step 1.1: Start With Source Data", rendered_page)
        self.assertIn("Optional: Add Manual Transactions", rendered_page)
        self.assertIn("Choose CSV file", rendered_page)
        self.assertIn('id="upload_exchange_csv"', rendered_page)
        self.assertIn('aria-label="Upload exchange CSV"', rendered_page)
        self.assertIn('action="/import_transactions/?guided=1"', rendered_page)
        self.assertIn("<summary>Column review options</summary>", rendered_page)
        self.assertIn("<summary>Show current import status</summary>", rendered_page)
        self.assertIn("<summary>Add manual rows</summary>", rendered_page)
        self.assertIn("<summary>Review data sources and revisions</summary>", rendered_page)
        self.assertLess(
            rendered_page.index("Step 1.1: Start With Source Data"),
            rendered_page.index("Show current import status"),
        )

    def test_guided_import_page_hides_advanced_sidebar_tools(self):
        app = create_app(config_dict["Debug"], selenium=True)

        with app.test_request_context("/import_transactions/?guided=1"):
            guided_sidebar = app.jinja_env.get_template(
                "site_template/sidebar.html"
            ).render()

        with app.test_request_context("/import_transactions/"):
            direct_sidebar = app.jinja_env.get_template(
                "site_template/sidebar.html"
            ).render()

        self.assertNotIn("Advanced tools", guided_sidebar)
        self.assertNotIn("Auto Link", guided_sidebar)
        self.assertIn("Advanced tools", direct_sidebar)

        with app.test_request_context("/holdings_accounting/?guided=1&mode=declare"):
            guided_holdings_sidebar = app.jinja_env.get_template(
                "site_template/sidebar.html"
            ).render()

        with app.test_request_context("/export/?guided=1"):
            guided_export_sidebar = app.jinja_env.get_template(
                "site_template/sidebar.html"
            ).render()

        self.assertNotIn("Advanced tools", guided_holdings_sidebar)
        self.assertNotIn("Advanced tools", guided_export_sidebar)

    def test_holdings_page_has_separate_guided_declare_and_reconcile_modes(self):
        app = create_app(config_dict["Debug"], selenium=True)
        app.config.update(WTF_CSRF_ENABLED=False)
        stats_rows = [
            {
                "symbol": "BTC",
                "total_purchased_quantity": "1",
                "total_sold_quantity": "0",
                "total_sold_unlinked_quantity": "0",
                "total_purchased_unlinked_quantity": "0",
                "total_purchased_usd": "$100.00",
                "total_sold_usd": "$0.00",
                "total_profit_loss": "$0.00",
                "holdings": "N/A",
            }
        ]
        holdings_summary = {
            "asset_count": 1,
            "assets_needing_holdings": 1,
            "assets_matched": 0,
            "assets_with_mismatch": 0,
        }

        with app.test_request_context("/holdings_accounting/?guided=1&mode=declare"):
            declare_page = app.jinja_env.get_template(
                "holdings_accounting.html"
            ).render(
                stats_table_data=stats_rows,
                holdings_summary=holdings_summary,
                guided_mode=True,
                holdings_mode="declare",
            )
        with app.test_request_context("/holdings_accounting/?guided=1&mode=declare"):
            ready_declare_page = app.jinja_env.get_template(
                "holdings_accounting.html"
            ).render(
                stats_table_data=stats_rows,
                holdings_summary={
                    **holdings_summary,
                    "assets_needing_holdings": 0,
                },
                guided_mode=True,
                holdings_mode="declare",
            )

        with app.test_request_context("/holdings_accounting/?guided=1&mode=reconcile"):
            reconcile_page = app.jinja_env.get_template(
                "holdings_accounting.html"
            ).render(
                stats_table_data=stats_rows,
                holdings_summary=holdings_summary,
                guided_mode=True,
                holdings_mode="reconcile",
            )
        with app.test_request_context("/holdings_accounting/?guided=1&mode=reconcile"):
            complete_reconcile_page = app.jinja_env.get_template(
                "holdings_accounting.html"
            ).render(
                stats_table_data=stats_rows,
                holdings_summary={
                    **holdings_summary,
                    "assets_needing_holdings": 0,
                    "assets_with_mismatch": 0,
                    "assets_matched": 1,
                },
                guided_mode=True,
                holdings_mode="reconcile",
            )

        self.assertIn("Step 2 of 4", declare_page)
        self.assertIn("Declare Holdings", declare_page)
        self.assertIn("Declare what you currently hold", declare_page)
        self.assertIn("Declared holdings give Gainz the real-world target", declare_page)
        self.assertNotIn("Continue to Reconcile Gaps", declare_page)
        self.assertIn("Continue to Reconcile Gaps", ready_declare_page)
        self.assertIn("Current holdings are recorded. Gainz can now compare them to imported activity.", ready_declare_page)
        self.assertIn('id="holdings_current_task"', ready_declare_page)
        self.assertIn('id="holdings_gap_queue" class="row" style="display: none;"', ready_declare_page)
        self.assertIn('id="holdings_workbench_section" class="row" style="display: none;"', ready_declare_page)
        self.assertLess(
            ready_declare_page.index('id="holdings_current_task"'),
            ready_declare_page.index("Current holdings are recorded"),
        )
        self.assertIn("holdings-guided-card-grid", declare_page)
        self.assertIn("holdings-data-source-only", declare_page)
        self.assertIn("Step 2.1: Bulk Current Holdings", declare_page)
        self.assertIn("Step 2.2: Choose An Asset To Declare", declare_page)
        self.assertIn("Step 2.3: Enter Current Holdings", declare_page)
        self.assertIn("Confirm zero holdings", declare_page)
        self.assertIn("Add Another Holding", declare_page)
        self.assertIn("Save These Holdings And Set The Rest To 0", declare_page)
        self.assertIn("Save Declared Holdings", declare_page)
        self.assertNotIn("Documented Activity Classification", declare_page)
        self.assertNotIn("Step 4: Review Readiness", declare_page)
        self.assertNotIn("Step 3 Asset Workbench", declare_page)

        self.assertIn("Step 3 of 4", reconcile_page)
        self.assertIn("Reconcile Gaps", reconcile_page)
        self.assertNotIn("Work one asset gap at a time.", reconcile_page)
        self.assertIn("Start with Step 3.1 below", reconcile_page)
        self.assertIn("Step 3.1: Choose One Gap To Review", reconcile_page)
        self.assertIn("Step 3.2: Understand This Gap", reconcile_page)
        self.assertIn("Step 3.3: Record A Gap Decision", reconcile_page)
        self.assertIn("Step 3.4: Confirm The Next Action", reconcile_page)
        self.assertIn("Review one holdings gap", reconcile_page)
        self.assertIn("A gap means imported activity does not yet explain declared holdings", reconcile_page)
        self.assertIn('id="holdings_current_help"', reconcile_page)
        self.assertIn("What missing basis means", reconcile_page)
        self.assertNotIn("Open Guided Review Queue", reconcile_page)
        self.assertIn("Start Step 3.1", reconcile_page)
        self.assertIn("Current gap", reconcile_page)
        self.assertIn("Advanced gap details", reconcile_page)
        self.assertIn("Open Declare Holdings", reconcile_page)
        self.assertIn("Review conservative full-proceeds gain using $0 basis", reconcile_page)
        self.assertIn("What the conservative option does", reconcile_page)
        self.assertIn("does not prove the missing", reconcile_page)
        self.assertNotIn("Save Declared Holdings", reconcile_page)
        self.assertNotIn("Save 0 Holdings", reconcile_page)
        self.assertIn("Reconciliation gaps complete", complete_reconcile_page)
        self.assertIn("Continue to Reports &amp; Export", complete_reconcile_page)
        self.assertNotIn("Start Step 3.1", complete_reconcile_page)

        custom_js = Path("app/base/static/assets/js/custom.js").read_text(encoding="utf-8")
        self.assertIn("Step 3.2: Understand ", custom_js)
        self.assertIn("What a holdings gap means", custom_js)
        self.assertIn("Leave Holdings Gap As Needs Research", custom_js)
        self.assertIn("Reconciliation gaps complete", custom_js)
        self.assertIn("decision=conservative_max_gain", custom_js)

    def test_import_page_renders_import_warning_workflow(self):
        app = create_app(config_dict["Debug"], selenium=True)
        app.config.update(WTF_CSRF_ENABLED=False)
        warnings = [
            "Skipped row 12 from coinbase.csv: unrecognized transaction type 'Mystery Reward'"
        ]
        warning_rows = import_warning_review_rows(warnings)

        with app.test_request_context("/import_transactions/"):
            rendered_page = app.jinja_env.get_template(
                "import_transactions.html"
            ).render(
                **import_template_context(
                    data_summary={
                        **import_template_context()["data_summary"],
                        "transaction_count": 1,
                        "import_warnings": warnings,
                        "import_warning_rows": warning_rows,
                        "unresolved_import_warning_rows": warning_rows,
                        "unresolved_import_warning_count": 1,
                    }
                )
            )

        self.assertIn('id="import_warning_workflow"', rendered_page)
        self.assertIn("Import warnings need review", rendered_page)
        self.assertIn("Gainz found an import row that needs review", rendered_page)
        self.assertIn("Show all import warning rows", rendered_page)
        self.assertIn("coinbase.csv", rendered_page)
        self.assertIn("What should happen with this row?", rendered_page)
        self.assertIn("Advanced import repair", rendered_page)
        self.assertIn("Show source path", rendered_page)
        self.assertIn("Open Advanced Import / Column Mapping", rendered_page)
        self.assertIn("Remove this source and re-import", rendered_page)
        self.assertIn("I do not know yet", rendered_page)
        self.assertIn("Leave unresolved for draft only", rendered_page)
        self.assertIn("Import data is loaded, but review is still needed", rendered_page)
        self.assertIn("Review 1 unresolved import warning", rendered_page)

    def test_guided_import_current_decision_prioritizes_warning(self):
        app = create_app(config_dict["Debug"], selenium=True)
        app.config.update(WTF_CSRF_ENABLED=False)
        warnings = [
            "Imported row 261 from cash_app_report_btc2019.csv with $0 USD spot price."
        ]
        warning_rows = import_warning_review_rows(warnings)

        with app.test_request_context("/import_transactions/?guided=1"):
            rendered_page = app.jinja_env.get_template(
                "import_transactions.html"
            ).render(
                **import_template_context(
                    guided_import=True,
                    data_summary={
                        **import_template_context()["data_summary"],
                        "transaction_count": 2,
                        "import_warnings": warnings,
                        "import_warning_rows": warning_rows,
                        "unresolved_import_warning_rows": warning_rows,
                        "unresolved_import_warning_count": 1,
                    }
                )
        )

        self.assertIn("Decide what row 261 represents", rendered_page)
        self.assertIn("Step 1.2: Import warnings need review", rendered_page)
        self.assertIn("Step 1.3: Confirm Imported Values", rendered_page)
        self.assertIn("Step 1.4: Import data is loaded, but review is still needed", rendered_page)
        self.assertIn("This needs a corrected value", rendered_page)
        self.assertNotIn("Try demo data or upload one exchange CSV.", rendered_page)

    def test_source_overlap_review_starts_with_plain_language_decision(self):
        app = create_app(config_dict["Debug"], selenium=True)
        app.config.update(WTF_CSRF_ENABLED=False)
        overlap = {
            "source_a": "exchange_full_history.csv",
            "source_b": "exchange_year_export.csv",
            "name_a": "exchange_full_history.csv",
            "name_b": "exchange_year_export.csv",
            "date_range_a": "2017-2025",
            "date_range_b": "2025",
            "matching_rows": 42,
            "overlap_percent": "95%",
            "status": "Needs review",
            "message": "These sources appear to overlap.",
            "next_action": "Confirm whether one export duplicates the other before removing anything.",
        }

        with app.test_request_context("/import_transactions/?guided=1"):
            rendered_page = app.jinja_env.get_template(
                "import_transactions.html"
            ).render(
                **import_template_context(
                    guided_import=True,
                    data_summary={
                        **import_template_context()["data_summary"],
                        "transaction_count": 3,
                        "source_overlap_count": 1,
                        "source_overlaps": [overlap],
                    }
                )
            )

        self.assertIn("Review possible duplicate source files", rendered_page)
        self.assertIn("Decide whether these files duplicate the same activity", rendered_page)
        self.assertIn("Review row coverage before removing anything", rendered_page)
        self.assertIn("Show source overlap evidence table", rendered_page)

    def test_import_page_hides_continue_until_warnings_are_resolved(self):
        app = create_app(config_dict["Debug"], selenium=True)
        app.config.update(WTF_CSRF_ENABLED=False)
        warnings = ["Imported row 261 from cash_app_report_btc2019.csv with $0 USD spot price."]
        warning_rows = import_warning_review_rows(warnings)

        with app.test_request_context("/import_transactions/?guided=1"):
            blocked_page = app.jinja_env.get_template("import_transactions.html").render(
                **import_template_context(
                    guided_import=True,
                    data_summary={
                        **import_template_context()["data_summary"],
                        "transaction_count": 2,
                        "import_warnings": warnings,
                        "import_warning_rows": warning_rows,
                        "unresolved_import_warning_rows": warning_rows,
                        "unresolved_import_warning_count": 1,
                    },
                )
            )

        with app.test_request_context("/import_transactions/?guided=1"):
            ready_page = app.jinja_env.get_template("import_transactions.html").render(
                **import_template_context(
                    guided_import=True,
                    data_summary={
                        **import_template_context()["data_summary"],
                        "transaction_count": 2,
                        "import_warnings": warnings,
                        "import_warning_rows": warning_rows,
                        "unresolved_import_warning_rows": [],
                        "unresolved_import_warning_count": 0,
                    },
                )
            )

        self.assertIn("Import data is loaded, but review is still needed", blocked_page)
        self.assertIn("Step 1.3: Confirm Imported Values", blocked_page)
        self.assertIn("Step 1.4: Import data is loaded, but review is still needed", blocked_page)
        self.assertIn('class="btn btn-primary btn-round btn-sm import-continue-action"', blocked_page)
        self.assertIn('style="display: none;"', blocked_page)
        self.assertIn("Import data is ready for the next step", ready_page)
        self.assertIn("Step 1.4: Import data is ready for the next step", ready_page)
        self.assertIn("Continue to Declare Holdings", ready_page)

    def test_import_page_renders_column_review_workflow(self):
        app = create_app(config_dict["Debug"], selenium=True)
        app.config.update(WTF_CSRF_ENABLED=False)

        with app.test_request_context("/import_transactions/"):
            rendered_page = app.jinja_env.get_template(
                "import_transactions.html"
            ).render(**import_template_context())

        self.assertIn('id="import_column_mapper"', rendered_page)
        self.assertIn("Column review needed", rendered_page)
        self.assertIn("Choose Header And Columns", rendered_page)

    def test_public_import_result_strips_server_paths(self):
        result = _public_import_result({
            "file_path": r"C:\private\uploads\transactions.csv",
            "imported_count": 1,
            "warnings": [
                "Skipped row 12 from transactions.csv: unrecognized transaction type 'Mystery Reward'"
            ],
            "files": [
                {
                    "file_path": r"C:\private\demo_data\cash_app_sample.csv",
                    "imported_count": 3,
                }
            ],
        })

        self.assertNotIn("file_path", result)
        self.assertEqual("transactions.csv", result["filename"])
        self.assertNotIn("file_path", result["files"][0])
        self.assertEqual("cash_app_sample.csv", result["files"][0]["filename"])
        self.assertEqual("transactions.csv", result["warning_rows"][0]["source"])
        self.assertEqual("12", result["warning_rows"][0]["row"])

    def test_usd_spot_import_warning_points_to_source_row_and_advanced_import(self):
        warning_rows = import_warning_review_rows([
            "Imported row 261 from cash_app_report_btc2019.csv with $0 USD spot price."
        ])

        self.assertEqual(1, len(warning_rows))
        self.assertEqual("cash_app_report_btc2019.csv", warning_rows[0]["source"])
        self.assertEqual("261", warning_rows[0]["row"])
        self.assertIn("Open the source file and check row 261", warning_rows[0]["next_action"])
        self.assertIn("USD spot/total USD value column", warning_rows[0]["next_action"])
        self.assertIn("re-import using Advanced Import", warning_rows[0]["next_action"])

    def test_usd_spot_withdrawal_warning_uses_owner_transfer_decision_card(self):
        transactions = empty_transactions()

        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "cash_app_report_btc2019.csv"
            source.write_text(
                "\n".join([
                    "Transaction ID,Date,Transaction Type,Currency,Amount,Fee,Net Amount,Asset Type,Asset Price,Asset Amount,Status,Notes,Name,Account",
                    "m11cax,2020-02-16 21:59:29 EST,Bitcoin Withdrawal,USD,$0,$0,$0,BTC,,0.03760644,COMPLETED,Withdrawing BTC 0.03760644,,Your Cash",
                    "h88o3a,2020-02-16 21:58:49 EST,Bitcoin Buy,USD,($370),($6.71),($376.71),BTC,$9838.74,0.03760644,COMPLETED,purchase of BTC 0.03760644,,Your Cash",
                ]),
                encoding="utf-8",
            )
            transactions.transactions = [
                Send("BTC", 0.03760644, datetime.datetime(2020, 2, 16, 21, 59, 29), 0, str(source))
            ]

            warning_rows = import_warning_review_rows([
                "Imported row 2 from cash_app_report_btc2019.csv with $0 USD spot price."
            ], transactions=transactions)

        warning = warning_rows[0]
        self.assertEqual("zero_usd_transfer", warning["mode"])
        self.assertEqual("Gainz found a $0 BTC withdrawal", warning["card_title"])
        self.assertEqual("What happened to this BTC?", warning["question"])
        self.assertEqual("Bitcoin Withdrawal", warning["raw_row_type"])
        self.assertEqual("$0", warning["raw_usd_amount"])
        self.assertIn("Withdrawing BTC 0.03760644", warning["notes"])
        self.assertIn("row 3: Bitcoin Buy", warning["nearby_summary"])
        self.assertIn("0.03760644 BTC", warning["nearby_summary"])
        self.assertEqual(
            {
                "decision": "true_zero_value_transfer",
                "label": "This went to my own wallet/account",
                "style": "primary",
            },
            warning["decision_options"][0],
        )

    def test_public_mapping_prompt_does_not_create_warning_rows(self):
        result = _public_import_result({
            "file_path": r"C:\private\uploads\cash_app_report.csv",
            "imported_count": 0,
            "skipped_count": 0,
            "warnings": [
                "Column review needed. Gainz could not confidently identify the required import columns."
            ],
            "mapping_required": True,
        })

        self.assertEqual("cash_app_report.csv", result["filename"])
        self.assertEqual([], result["warning_rows"])

    def test_import_route_errors_do_not_expose_exception_details(self):
        app = create_app(config_dict["Debug"], selenium=True)
        app.config.update(
            WTF_CSRF_ENABLED=False,
            transactions=empty_transactions(),
            UPLOAD_FOLDER="uploads",
        )

        with app.test_request_context("/import_transactions/demo", method="POST"):
            with patch.object(
                import_routes.ImportService,
                "import_demo_data",
                side_effect=RuntimeError(r"C:\private\secret.csv"),
            ):
                response, status_code = import_routes.import_demo_data.__wrapped__()

        data = response.get_json()
        self.assertEqual(400, status_code)
        self.assertIn("could not complete", data["error"])
        self.assertNotIn("secret.csv", data["error"])
        self.assertNotIn("C:\\", data["error"])

    def test_manual_batch_rows_skip_blanks_and_build_transactions(self):
        form_data = MultiDict([
            ("manual_type[]", "buy"),
            ("manual_timestamp[]", "2024-01-01T09:30"),
            ("manual_symbol[]", "btc"),
            ("manual_quantity[]", "0.25"),
            ("manual_usd_spot[]", "42000"),
            ("manual_type[]", "sell"),
            ("manual_timestamp[]", "2024-02-02T10:45"),
            ("manual_symbol[]", "ETH"),
            ("manual_quantity[]", "1.5"),
            ("manual_usd_spot[]", "2500"),
            ("manual_type[]", "buy"),
            ("manual_timestamp[]", ""),
            ("manual_symbol[]", ""),
            ("manual_quantity[]", ""),
            ("manual_usd_spot[]", ""),
        ])

        rows = [
            row
            for row in _manual_batch_rows(form_data)
            if not _manual_row_is_blank(row)
        ]
        transactions = [
            _manual_transaction_from_values(row, index + 1)
            for index, row in enumerate(rows)
        ]

        self.assertEqual(2, len(transactions))
        self.assertIsInstance(transactions[0], Buy)
        self.assertIsInstance(transactions[1], Sell)
        self.assertEqual("BTC", transactions[0].symbol)
        self.assertEqual(0.25, transactions[0].quantity)
        self.assertEqual(2500, transactions[1].usd_spot)
        self.assertEqual("Gainz App Manual Add", transactions[1].source)

    def test_manual_batch_rejects_partial_rows(self):
        with self.assertRaisesRegex(ValueError, "Row 3: complete USD Spot"):
            _manual_transaction_from_values(
                {
                    "type": "buy",
                    "timestamp": "2024-01-01T09:30",
                    "symbol": "BTC",
                    "quantity": "0.5",
                    "usd_spot": "",
                },
                row_number=3,
            )

    def test_import_page_posts_manual_batch_as_one_revision(self):
        transactions = empty_transactions()
        saved_descriptions = []
        transactions.save = lambda description=None: saved_descriptions.append(description)
        app = create_app(config_dict["Debug"], selenium=True)
        app.config.update(
            WTF_CSRF_ENABLED=False,
            transactions=transactions,
            UPLOAD_FOLDER="uploads",
        )

        form_data = {
            "manual_batch_submit": "1",
            "manual_type[]": ["buy", "sell", "buy"],
            "manual_timestamp[]": ["2024-01-01T09:30", "2024-02-02T10:45", ""],
            "manual_symbol[]": ["BTC", "ETH", ""],
            "manual_quantity[]": ["0.25", "1.5", ""],
            "manual_usd_spot[]": ["42000", "2500", ""],
        }

        with app.test_request_context(
            "/import_transactions/",
            method="POST",
            data=form_data,
        ):
            response = import_routes.import_wizard.__wrapped__()

        self.assertEqual(302, response.status_code)
        self.assertIn("manual_added=2", response.location)
        self.assertEqual(["Manually Added 2 Transactions"], saved_descriptions)
        self.assertEqual(["buy", "sell"], [t.trans_type for t in transactions.transactions])
        self.assertEqual(["BTC", "ETH"], [t.symbol for t in transactions.transactions])

    def test_remove_data_source_removes_transactions_and_cleans_links(self):
        transactions = empty_transactions()
        removed_buy = Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "remove.csv")
        remaining_buy = Buy("BTC", 2, datetime.datetime(2024, 1, 2), 100, "keep.csv")
        remaining_sell = Sell("BTC", 0.5, datetime.datetime(2024, 2, 1), 300, "keep.csv")
        remaining_sell.link_transaction(removed_buy, 0.5)
        transactions.transactions = [removed_buy, remaining_buy, remaining_sell]
        transactions.import_warnings = [
            "Skipped row 4 from remove.csv: unrecognized transaction type 'Mystery'",
            "Skipped row 6 from remove.csv: unrecognized transaction type 'Other'",
            "Skipped row 5 from keep.csv: unrecognized transaction type 'Mystery'",
        ]
        transactions.set_import_warning_review(
            "Skipped row 4 from remove.csv: unrecognized transaction type 'Mystery'",
            decision="note",
            note="Reviewed before source cleanup.",
        )

        result = _remove_data_source(transactions, "remove.csv")

        self.assertEqual(1, result["removed_count"])
        self.assertEqual(["keep.csv", "keep.csv"], [transaction.source for transaction in transactions.transactions])
        self.assertEqual([], remaining_sell.links)
        self.assertEqual([], remaining_sell.linked_transactions)
        self.assertEqual([
            "Skipped row 5 from keep.csv: unrecognized transaction type 'Mystery'"
        ], transactions.import_warnings)
        self.assertEqual(
            "Reviewed before source cleanup.",
            transactions.get_import_warning_review(
                "Skipped row 4 from remove.csv: unrecognized transaction type 'Mystery'"
            )["note"],
        )
        self.assertEqual(
            "cleared_by_source_update",
            transactions.get_import_warning_review(
                "Skipped row 6 from remove.csv: unrecognized transaction type 'Other'"
            )["decision"],
        )

    def test_import_service_imports_coinbase_sample(self):
        transactions = empty_transactions()
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path("demo_data/coinbase_sample.csv")
            result = ImportService(temp_dir).import_upload(FileUpload(source), transactions)

        self.assertEqual(3, result["imported_count"])
        self.assertEqual(0, result["skipped_count"])
        self.assertEqual(["buy", "buy", "sell"], [t.trans_type for t in transactions.transactions])
        self.assertEqual({"ETH"}, transactions.assets)

    def test_import_service_imports_coinbase_convert_as_sell_and_buy(self):
        transactions = empty_transactions()
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path("demo_data/coinbase_convert_sample.csv")
            result = ImportService(temp_dir).import_upload(FileUpload(source), transactions)

        self.assertEqual(2, result["imported_count"])
        self.assertEqual(0, result["skipped_count"])
        self.assertEqual([], result["warnings"])
        self.assertEqual([], getattr(transactions, "import_warnings", []))

        self.assertEqual(["sell", "buy"], [t.trans_type for t in transactions.transactions])
        self.assertEqual(["ETH", "BTC"], [t.symbol for t in transactions.transactions])
        self.assertEqual([0.5, 0.02], [t.quantity for t in transactions.transactions])
        self.assertEqual(4000, transactions.transactions[0].usd_spot)
        self.assertEqual(100000, transactions.transactions[1].usd_spot)

    def test_import_service_imports_gdax_fills(self):
        transactions = empty_transactions()
        csv_text = "\n".join([
            "trade id,product,side,created at,size,size unit,price,fee,total,price/fee/total unit",
            "1,BTC-USD,BUY,2024-01-05T10:15:00Z,0.1,BTC,40000,1,4000,USD",
            "2,BTC-USD,SELL,2024-02-05T10:15:00Z,0.05,BTC,50000,1,2500,USD",
        ])

        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "gdax_fills.csv"
            upload_dir = Path(temp_dir) / "uploads"
            source.write_text(csv_text, encoding="utf-8")
            result = ImportService(upload_dir).import_upload(FileUpload(source), transactions)

        self.assertEqual(2, result["imported_count"])
        self.assertEqual(0, result["skipped_count"])
        self.assertEqual(["buy", "sell"], [t.trans_type for t in transactions.transactions])
        self.assertEqual({"BTC"}, transactions.assets)
        self.assertEqual([0.1, 0.05], sorted([t.quantity for t in transactions.transactions], reverse=True))

    def test_coinbase_partial_basis_fee_golden_totals(self):
        transactions = empty_transactions()

        with tempfile.TemporaryDirectory() as temp_dir:
            result = ImportService(temp_dir).import_upload(
                FileUpload(Path("Tests/fixtures/coinbase_partial_basis_fees.csv")),
                transactions,
            )

        self.assertEqual(2, result["imported_count"])
        self.assertEqual([], result["warnings"])
        buy, sell = sorted(transactions.transactions, key=lambda item: item.time_stamp)
        self.assertEqual(25.0, buy.gross_usd_total)
        self.assertEqual(0.5, buy.fee)
        self.assertEqual(25.5, buy.net_usd_total)
        self.assertEqual(500.0, sell.gross_usd_total)
        self.assertEqual(5.0, sell.fee)
        self.assertEqual(495.0, sell.net_usd_total)

        sell.link_transaction(buy, 0.2)
        unresolved_quantity = sell.unlinked_quantity
        transactions.apply_cpa_basis_resolution(
            target_sell_uid=sell.uid,
            quantity=unresolved_quantity,
            acquisition_date=sell.time_stamp,
            basis_value=0.0,
            proceeds_value=sell.prorated_tax_usd(unresolved_quantity),
            basis_method="unknown_zero_for_review",
            evidence_reference="Synthetic golden fixture",
            work_order_item_id="golden-bch-partial-basis",
            acquisition_date_method="cpa_conservative_short_term",
        )

        totals = get_form_8949_totals(transactions)
        self.assertAlmostEqual(495.0, totals["total"]["proceeds"])
        self.assertAlmostEqual(25.5, totals["total"]["cost_basis"])
        self.assertAlmostEqual(469.5, totals["total"]["gain_loss"])
        preview = get_packet_preview(
            transactions,
            get_audit_readiness_summary(transactions),
            "C:/synthetic-output",
        )
        self.assertEqual(1, preview["material_assumption_count"])
        self.assertIn("$5.50", preview["material_assumptions"][0]["detail"])

        import_summary = _data_source_summary(transactions)
        self.assertEqual(2, import_summary["import_economics_count"])
        sell_row = next(
            row for row in import_summary["import_economics_rows"]
            if row["transaction_type"] == "sell"
        )
        self.assertEqual("Net proceeds", sell_row["total_label"])
        self.assertAlmostEqual(500.0, sell_row["gross_usd"])
        self.assertAlmostEqual(5.0, sell_row["fee_usd"])
        self.assertAlmostEqual(495.0, sell_row["net_tax_usd"])

    def test_cash_app_fee_golden_round_trip(self):
        transactions = empty_transactions()

        with tempfile.TemporaryDirectory() as temp_dir:
            result = ImportService(temp_dir).import_upload(
                FileUpload(Path("Tests/fixtures/cash_app_fee_round_trip.csv")),
                transactions,
            )

        self.assertEqual(2, result["imported_count"])
        self.assertEqual([], result["warnings"])
        self.assertEqual([], transactions.auto_link(asset=None, algo="fifo"))
        totals = get_form_8949_totals(transactions)
        self.assertAlmostEqual(1485.0, totals["total"]["proceeds"])
        self.assertAlmostEqual(1010.0, totals["total"]["cost_basis"])
        self.assertAlmostEqual(475.0, totals["total"]["gain_loss"])

    def test_gdax_fee_golden_round_trip(self):
        transactions = empty_transactions()

        with tempfile.TemporaryDirectory() as temp_dir:
            result = ImportService(temp_dir).import_upload(
                FileUpload(Path("Tests/fixtures/gdax_fee_round_trip.csv")),
                transactions,
            )

        self.assertEqual(2, result["imported_count"])
        self.assertEqual([], result["warnings"])
        self.assertEqual([], transactions.auto_link(asset=None, algo="fifo"))
        totals = get_form_8949_totals(transactions)
        self.assertAlmostEqual(1485.0, totals["total"]["proceeds"])
        self.assertAlmostEqual(1010.0, totals["total"]["cost_basis"])
        self.assertAlmostEqual(475.0, totals["total"]["gain_loss"])

    def test_coinbase_convert_fee_golden_round_trip(self):
        transactions = empty_transactions()

        with tempfile.TemporaryDirectory() as temp_dir:
            result = ImportService(temp_dir).import_upload(
                FileUpload(Path("Tests/fixtures/coinbase_convert_fee_round_trip.csv")),
                transactions,
            )

        self.assertEqual(3, result["imported_count"])
        self.assertEqual(1, len(result["warnings"]))
        self.assertIn("acquired-asset basis", result["warnings"][0])
        self.assertEqual([], transactions.auto_link(asset=None, algo="fifo"))
        totals = get_form_8949_totals(transactions)
        self.assertAlmostEqual(1485.0, totals["total"]["proceeds"])
        self.assertAlmostEqual(1010.0, totals["total"]["cost_basis"])
        self.assertAlmostEqual(475.0, totals["total"]["gain_loss"])
        readiness = get_audit_readiness_summary(transactions)
        self.assertEqual(1, readiness["metrics"]["import_economics_warnings"])
        self.assertTrue(any(
            group["key"] == "import_economics"
            for group in readiness["blocker_groups"]
        ))
        self.assertFalse(readiness["is_ready"])

    def test_mapped_csv_fee_golden_round_trip(self):
        transactions = empty_transactions()
        csv_text = "\n".join([
            "When,Kind,Thing,Units,Unit USD,Gross USD,Fee USD,Net USD,Reference",
            "2023-01-05 10:15:00 UTC,Acquire,BTC,1,1000,1000,10,1010,mapped-buy-001",
            "2024-02-05 10:15:00 UTC,Dispose,BTC,1,1500,1500,15,1485,mapped-sell-001",
        ])

        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "mapped_fee_round_trip.csv"
            source.write_text(csv_text, encoding="utf-8")
            result = ImportService(Path(temp_dir) / "uploads").import_mapped_file(
                source,
                transactions,
                header_row=1,
                column_mapping={
                    "date": "When",
                    "transaction_type": "Kind",
                    "asset_type": "Thing",
                    "asset_amount": "Units",
                    "asset_price": "Unit USD",
                    "fiat_amount": "Gross USD",
                    "fee": "Fee USD",
                    "net_amount": "Net USD",
                    "transaction_id": "Reference",
                },
            )

        self.assertEqual(2, result["imported_count"])
        self.assertEqual([], result["warnings"])
        self.assertEqual([], transactions.auto_link(asset=None, algo="fifo"))
        totals = get_form_8949_totals(transactions)
        self.assertAlmostEqual(1485.0, totals["total"]["proceeds"])
        self.assertAlmostEqual(1010.0, totals["total"]["cost_basis"])
        self.assertAlmostEqual(475.0, totals["total"]["gain_loss"])

    def test_kraken_template_fee_golden_round_trip(self):
        transactions = empty_transactions()

        with tempfile.TemporaryDirectory() as temp_dir:
            result = ImportService(Path(temp_dir) / "uploads").import_mapped_file(
                Path("Tests/fixtures/kraken_mapped_fee_round_trip.csv"),
                transactions,
                header_row=1,
                column_mapping={
                    "date": "time",
                    "transaction_type": "type",
                    "asset_type": "pair",
                    "asset_amount": "vol",
                    "asset_price": "price",
                    "fiat_amount": "cost",
                    "fee": "fee",
                    "transaction_id": "txid",
                },
            )

        self.assertEqual(2, result["imported_count"])
        self.assertEqual([], result["warnings"])
        self.assertEqual([], transactions.auto_link(asset=None, algo="fifo"))
        totals = get_form_8949_totals(transactions)
        self.assertAlmostEqual(1485.0, totals["total"]["proceeds"])
        self.assertAlmostEqual(1010.0, totals["total"]["cost_basis"])
        self.assertAlmostEqual(475.0, totals["total"]["gain_loss"])

    def test_mapped_crypto_fee_is_preserved_and_blocks_readiness(self):
        transactions = empty_transactions()
        csv_text = "\n".join([
            "When,Kind,Thing,Units,Unit USD,Gross USD,Fee,Fee Currency,Reference",
            "2024-02-05 10:15:00 UTC,Dispose,ETH,1,1500,1500,0.01,ETH,crypto-fee-001",
        ])

        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "crypto_fee.csv"
            source.write_text(csv_text, encoding="utf-8")
            result = ImportService(Path(temp_dir) / "uploads").import_mapped_file(
                source,
                transactions,
                header_row=1,
                column_mapping={
                    "date": "When",
                    "transaction_type": "Kind",
                    "asset_type": "Thing",
                    "asset_amount": "Units",
                    "asset_price": "Unit USD",
                    "fiat_amount": "Gross USD",
                    "fee": "Fee",
                    "fee_currency": "Fee Currency",
                    "transaction_id": "Reference",
                },
            )

        self.assertEqual(1, result["imported_count"])
        self.assertTrue(any("not converted to USD" in warning for warning in result["warnings"]))
        transaction = transactions.transactions[0]
        self.assertIsNone(transaction.fee)
        self.assertEqual(0.01, transaction.source_fee_amount)
        self.assertEqual("ETH", transaction.fee_currency)
        self.assertIn("not converted to USD", transaction.economics_warning)
        readiness = get_audit_readiness_summary(transactions)
        self.assertFalse(readiness["is_ready"])
        self.assertEqual(1, readiness["metrics"]["import_economics_warnings"])

    def test_transaction_economics_survive_save_reload(self):
        transactions = empty_transactions()
        del transactions.save
        buy = Buy("BTC", 1, datetime.datetime(2023, 1, 5), 1000, "fee-source.csv")
        buy.set_economics(
            fee=10,
            gross_usd_total=1000,
            net_usd_total=1010,
            source_row=2,
            source_transaction_id="saved-buy-001",
            economics_source="gross value, fee, net amount",
        )
        transactions.transactions = [buy]

        with tempfile.TemporaryDirectory() as temp_dir, patch("transactions.basedir", temp_dir):
            save_path = transactions.save(description="Fee persistence golden")
            loaded = Transactions(view=save_path)

        loaded_buy = loaded.transactions[0]
        self.assertEqual(10.0, loaded_buy.fee)
        self.assertEqual(1000.0, loaded_buy.gross_usd_total)
        self.assertEqual(1010.0, loaded_buy.net_usd_total)
        self.assertEqual(2, loaded_buy.source_row)
        self.assertEqual("saved-buy-001", loaded_buy.source_transaction_id)

    def test_import_service_imports_cash_app_with_updated_headers(self):
        transactions = empty_transactions()
        csv_text = "\n".join([
            "Reference ID,Transaction Date,Activity,Fiat Currency,USD Amount,Network Fee,Crypto,Spot Price USD,Crypto Quantity",
            "cash-new-001,2024-01-05 10:15:00 PST,Bitcoin Purchase,USD,-$1000.00,$0.00,Bitcoin,\"$40,000.00\",0.025 BTC",
            "cash-new-002,2024-08-20 14:00:00 PDT,Bitcoin Sale,USD,$1200.00,-$12.00,Bitcoin,\"$60,000.00\",0.020 BTC",
        ])

        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "cash_app_updated.csv"
            upload_dir = Path(temp_dir) / "uploads"
            source.write_text(csv_text, encoding="utf-8")
            result = ImportService(upload_dir).import_upload(FileUpload(source), transactions)

        self.assertEqual(2, result["imported_count"])
        self.assertEqual(0, result["skipped_count"])
        self.assertEqual([], result["warnings"])
        self.assertEqual(["buy", "sell"], [t.trans_type for t in transactions.transactions])
        self.assertEqual(["BTC", "BTC"], [t.symbol for t in transactions.transactions])

    def test_import_service_imports_coinbase_with_updated_headers(self):
        transactions = empty_transactions()
        csv_text = "\n".join([
            "Trade ID,Date UTC,Activity Type,Asset Symbol,Amount Transacted,Unit Price,Total Amount,Memo",
            "cb-new-001,2023-01-10 12:00:00 UTC,Purchased,ETH,2.0,$1200.00,$2400.00,Bought 2 ETH",
            "cb-new-002,2024-03-20 12:00:00 UTC,Sold,ETH,-1.5,$3000.00,$4500.00,Sold 1.5 ETH",
        ])

        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "coinbase_updated.csv"
            upload_dir = Path(temp_dir) / "uploads"
            source.write_text(csv_text, encoding="utf-8")
            result = ImportService(upload_dir).import_upload(FileUpload(source), transactions)

        self.assertEqual(2, result["imported_count"])
        self.assertEqual(0, result["skipped_count"])
        self.assertEqual([], result["warnings"])
        self.assertEqual(["buy", "sell"], [t.trans_type for t in transactions.transactions])
        self.assertEqual([2.0, 1.5], [t.quantity for t in transactions.transactions])

    def test_import_service_imports_coinbase_earn_as_receive(self):
        transactions = empty_transactions()
        csv_text = "\n".join([
            "Timestamp,Transaction Type,Asset,Quantity Transacted,Total Inclusive of Fees and Spread,Notes",
            "2024-02-01 09:00:00 UTC,Coinbase Earn,SOL,3.5,$35.00,Earned SOL",
        ])

        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "coinbase_earn.csv"
            upload_dir = Path(temp_dir) / "uploads"
            source.write_text(csv_text, encoding="utf-8")
            result = ImportService(upload_dir).import_upload(FileUpload(source), transactions)

        self.assertEqual(1, result["imported_count"])
        self.assertEqual(0, result["skipped_count"])
        self.assertEqual([], result["warnings"])
        self.assertEqual("receive", transactions.transactions[0].trans_type)
        self.assertEqual("SOL", transactions.transactions[0].symbol)
        self.assertEqual(10.0, transactions.transactions[0].usd_spot)

    def test_import_service_imports_generic_alias_csv(self):
        transactions = empty_transactions()
        csv_text = "\n".join([
            "Created At,Operation,Token Symbol,Token Quantity,Transaction Value",
            "2024-02-01 09:00:00 UTC,Receive,SOL,3.5,$350.00",
            "2024-02-15 09:00:00 UTC,Send,SOL,1.0,$125.00",
        ])

        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "wallet_export.csv"
            upload_dir = Path(temp_dir) / "uploads"
            source.write_text(csv_text, encoding="utf-8")
            result = ImportService(upload_dir).import_upload(FileUpload(source), transactions)

        self.assertEqual(2, result["imported_count"])
        self.assertEqual(0, result["skipped_count"])
        self.assertEqual([], result["warnings"])
        self.assertEqual(["receive", "send"], [t.trans_type for t in transactions.transactions])
        self.assertEqual([100.0, 125.0], [t.usd_spot for t in transactions.transactions])

    def test_import_service_handles_eastern_timezone_abbreviations_without_warning(self):
        transactions = empty_transactions()
        csv_text = "\n".join([
            "Created At,Operation,Token Symbol,Token Quantity,Transaction Value",
            "2019-02-01 09:00:00 EST,Receive,BTC,0.1,$350.00",
            "2019-08-15 14:30:00 EDT,Send,BTC,0.05,$250.00",
        ])

        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "wallet_export_eastern_time.csv"
            upload_dir = Path(temp_dir) / "uploads"
            source.write_text(csv_text, encoding="utf-8")
            with warnings.catch_warnings(record=True) as caught:
                warnings.simplefilter("always")
                result = ImportService(upload_dir).import_upload(FileUpload(source), transactions)

        unknown_timezone_warnings = [
            warning for warning in caught
            if issubclass(warning.category, UnknownTimezoneWarning)
        ]
        self.assertEqual([], unknown_timezone_warnings)
        self.assertEqual(2, result["imported_count"])
        self.assertEqual(["receive", "send"], [t.trans_type for t in transactions.transactions])

    def test_import_service_finds_header_row_after_preamble(self):
        transactions = empty_transactions()
        csv_text = "\n".join([
            "Wallet Export",
            "Generated for demo testing",
            "Created At,Operation,Token Symbol,Token Quantity,Transaction Value",
            "2024-02-01 09:00:00 UTC,Receive,SOL,3.5,$350.00",
        ])

        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "wallet_export_with_preamble.csv"
            upload_dir = Path(temp_dir) / "uploads"
            source.write_text(csv_text, encoding="utf-8")
            result = ImportService(upload_dir).import_upload(FileUpload(source), transactions)

        self.assertEqual(1, result["imported_count"])
        self.assertEqual(3, result["header_row_used"])
        self.assertEqual("SOL", transactions.transactions[0].symbol)

    def test_import_service_can_pause_for_column_review(self):
        transactions = empty_transactions()
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path("demo_data/cash_app_sample.csv")
            result = ImportService(temp_dir).import_upload(
                FileUpload(source),
                transactions,
                review_columns=True,
            )

        self.assertTrue(result["mapping_required"])
        self.assertEqual(0, result["imported_count"])
        self.assertEqual([], transactions.transactions)
        self.assertEqual(1, result["mapping"]["header_row"])
        self.assertEqual(2, result["mapping"]["data_start_row"])
        self.assertTrue(result["mapping"]["sample_rows"])

    def test_import_service_requests_mapping_when_pricing_column_missing(self):
        transactions = empty_transactions()
        csv_text = "\n".join([
            "Created At,Operation,Token Symbol,Token Quantity",
            "2024-02-01 09:00:00 UTC,Receive,SOL,3.5",
        ])

        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "wallet_export_without_usd.csv"
            upload_dir = Path(temp_dir) / "uploads"
            source.write_text(csv_text, encoding="utf-8")
            result = ImportService(upload_dir).import_upload(FileUpload(source), transactions)

        self.assertTrue(result["mapping_required"])
        self.assertIn("USD", result["warnings"][0])
        self.assertEqual(0, result["imported_count"])
        self.assertEqual([], transactions.transactions)

    def test_import_service_imports_mapped_csv_with_later_data_start_row(self):
        transactions = empty_transactions()
        csv_text = "\n".join([
            "Wallet Export",
            "Generated for demo testing",
            "Created At,Operation,Token Symbol,Token Quantity,Transaction Value",
            "Notes: internal transfers included below,,,,",
            "2024-02-01 09:00:00 UTC,Receive,SOL,3.5,$350.00",
        ])

        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "wallet_export_with_note_row.csv"
            upload_dir = Path(temp_dir) / "uploads"
            source.write_text(csv_text, encoding="utf-8")
            service = ImportService(upload_dir)
            upload_result = service.import_upload(FileUpload(source), transactions, review_columns=True)
            mapped_result = service.import_mapped_file(
                upload_result["file_path"],
                transactions,
                header_row=3,
                data_start_row=5,
                column_mapping={
                    "date": "Created At",
                    "transaction_type": "Operation",
                    "asset_type": "Token Symbol",
                    "asset_amount": "Token Quantity",
                    "fiat_amount": "Transaction Value",
                },
            )

        self.assertEqual(1, mapped_result["imported_count"])
        self.assertEqual([], mapped_result["warnings"])
        self.assertEqual("SOL", transactions.transactions[0].symbol)
        self.assertEqual("receive", transactions.transactions[0].trans_type)
        self.assertEqual(100.0, transactions.transactions[0].usd_spot)

    def test_import_warnings_do_not_expose_parser_exception_details(self):
        transactions = empty_transactions()
        csv_text = "\n".join([
            "Created At,Operation,Token Symbol,Token Quantity,Transaction Value",
            "not a date,Receive,SOL,3.5,$350.00",
            "2024-02-15 09:00:00 UTC,Send,SOL,1.0,$125.00",
        ])

        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "wallet_export_bad_row.csv"
            upload_dir = Path(temp_dir) / "uploads"
            source.write_text(csv_text, encoding="utf-8")
            result = ImportService(upload_dir).import_upload(FileUpload(source), transactions)

        self.assertEqual(1, result["imported_count"])
        self.assertEqual(1, len(result["warnings"]))
        warning = result["warnings"][0]
        self.assertIn("could not parse this row", warning)
        self.assertIn("wallet_export_bad_row.csv", warning)
        self.assertNotIn("Unknown string format", warning)
        self.assertNotIn("not a date", warning)
        self.assertNotIn(str(source.parent), warning)

    def test_import_service_returns_mapping_prompt_for_unknown_columns(self):
        transactions = empty_transactions()
        csv_text = "\n".join([
            "When,Kind,Thing,Units,Value",
            "2024-02-01 09:00:00 UTC,Acquire,SOL,3.5,$350.00",
        ])

        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "unknown_columns.csv"
            upload_dir = Path(temp_dir) / "uploads"
            source.write_text(csv_text, encoding="utf-8")
            service = ImportService(upload_dir)
            result = service.import_upload(FileUpload(source), transactions)

            self.assertTrue(result["mapping_required"])
            mapped_result = service.import_mapped_file(
                result["file_path"],
                transactions,
                header_row=1,
                column_mapping={
                    "date": "When",
                    "transaction_type": "Kind",
                    "asset_type": "Thing",
                    "asset_amount": "Units",
                    "fiat_amount": "Value",
                },
            )

        self.assertEqual(1, mapped_result["imported_count"])
        self.assertEqual([], mapped_result["warnings"])
        self.assertEqual("buy", transactions.transactions[0].trans_type)
        self.assertEqual(100.0, transactions.transactions[0].usd_spot)

    def test_demo_data_golden_form_8949_totals_after_fifo_linking(self):
        transactions = empty_transactions()

        with tempfile.TemporaryDirectory() as temp_dir:
            result = ImportService(temp_dir).import_demo_data(transactions, repo_root=Path.cwd())

        self.assertEqual(8, result["imported_count"])
        self.assertEqual([], result["warnings"])
        self.assertEqual([], transactions.auto_link(asset=None, algo="fifo"))

        totals = get_form_8949_totals(transactions)
        self.assertEqual(1, totals["short"]["rows"])
        self.assertEqual(2000.0, totals["short"]["proceeds"])
        self.assertEqual(605.0, totals["short"]["cost_basis"])
        self.assertEqual(1395.0, totals["short"]["gain_loss"])
        self.assertEqual(2, totals["long"]["rows"])
        self.assertEqual(5648.0, totals["long"]["proceeds"])
        self.assertEqual(3815.0, totals["long"]["cost_basis"])
        self.assertEqual(1833.0, totals["long"]["gain_loss"])
        self.assertEqual(3, totals["total"]["rows"])
        self.assertEqual(7648.0, totals["total"]["proceeds"])
        self.assertEqual(4420.0, totals["total"]["cost_basis"])
        self.assertEqual(3228.0, totals["total"]["gain_loss"])

    def test_demo_import_route_runs_default_fifo_automatically(self):
        transactions = empty_transactions()

        with tempfile.TemporaryDirectory() as temp_dir:
            app = create_app(config_dict["Debug"], selenium=True)
            app.config.update(
                WTF_CSRF_ENABLED=False,
                transactions=transactions,
                UPLOAD_FOLDER=temp_dir,
            )

            with app.test_request_context("/import_transactions/demo", method="POST"):
                response = import_routes.import_demo_data.__wrapped__()

        payload = response.get_json()
        totals = get_form_8949_totals(transactions)

        self.assertEqual(8, payload["imported_count"])
        self.assertEqual(3, payload["auto_link"]["links_created"])
        self.assertEqual("fifo", payload["auto_link"]["algo"])
        self.assertEqual([], payload["auto_link"]["failures"])
        self.assertEqual(3, totals["total"]["rows"])
        self.assertIn(
            "Automatically added FIFO basis links after demo data import",
            transactions.saved_descriptions,
        )

    def test_missing_basis_demo_exposes_partial_bch_basis_workflow(self):
        transactions = empty_transactions()

        with tempfile.TemporaryDirectory() as temp_dir:
            result = ImportService(temp_dir).import_missing_basis_demo(
                transactions,
                repo_root=Path.cwd(),
            )

        self.assertEqual(2, result["imported_count"])
        self.assertEqual("missing_basis", result["demo_kind"])
        self.assertEqual("BCH", result["demo_profile"]["asset"])
        self.assertEqual("0", result["demo_profile"]["declared_holdings"])
        transactions.auto_link(asset=None, algo="fifo")
        sell = next(transaction for transaction in transactions.transactions if transaction.trans_type == "sell")
        self.assertAlmostEqual(0.3, sell.unlinked_quantity)
        self.assertAlmostEqual(495.0, sell.tax_usd_total)
        self.assertAlmostEqual(5.0, sell.fee)

    def test_missing_basis_demo_route_runs_fifo_and_returns_judge_guidance(self):
        transactions = empty_transactions()

        with tempfile.TemporaryDirectory() as temp_dir:
            app = create_app(config_dict["Debug"], selenium=True)
            app.config.update(
                WTF_CSRF_ENABLED=False,
                transactions=transactions,
                UPLOAD_FOLDER=temp_dir,
            )

            with app.test_request_context("/import_transactions/demo_missing_basis", method="POST"):
                response = import_routes.import_missing_basis_demo.__wrapped__()

        payload = response.get_json()
        sell = next(transaction for transaction in transactions.transactions if transaction.trans_type == "sell")
        self.assertEqual(2, payload["imported_count"])
        self.assertEqual(1, payload["auto_link"]["links_created"])
        self.assertEqual("fifo", payload["auto_link"]["algo"])
        self.assertAlmostEqual(0.3, sell.unlinked_quantity)
        self.assertIn("declare BCH holdings as 0", payload["demo_profile"]["guidance"])
        self.assertIn(
            "Automatically added FIFO basis links after missing-basis demo import",
            transactions.saved_descriptions,
        )

    def test_export_includes_8949_short_totals(self):
        transactions = empty_transactions()
        buy = Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "demo")
        sell = Sell("BTC", 1, datetime.datetime(2024, 6, 1), 300, "demo")
        sell.link_transaction(buy, 1)
        transactions.transactions = [buy, sell]

        with tempfile.TemporaryDirectory() as temp_dir:
            export_path = transactions.export_to_excel(output_dir=temp_dir)
            workbook = load_workbook(export_path, data_only=False)
            sheet = workbook["2024 8949 Short"]

            self.assertEqual("Crypto BTC", sheet["A2"].value)
            self.assertEqual(300, sheet["D2"].value)
            self.assertEqual(100, sheet["E2"].value)
            self.assertEqual(200, sheet["H2"].value)

    def test_export_handles_timezone_aware_imported_datetimes(self):
        transactions = empty_transactions()
        buy = Buy(
            "ETH",
            1,
            datetime.datetime(2024, 1, 1, tzinfo=datetime.timezone.utc),
            100,
            "coinbase",
        )
        sell = Sell(
            "ETH",
            1,
            datetime.datetime(2024, 6, 1, tzinfo=datetime.timezone.utc),
            300,
            "coinbase",
        )
        sell.link_transaction(buy, 1)
        transactions.transactions = [buy, sell]

        with tempfile.TemporaryDirectory() as temp_dir:
            export_path = transactions.export_to_excel(output_dir=temp_dir)
            workbook = load_workbook(export_path, data_only=False)
            sheet = workbook["2024 8949 Short"]

            self.assertIsNone(sheet["B2"].value.tzinfo)
            self.assertIsNone(sheet["C2"].value.tzinfo)

    def test_holdings_classifies_selected_documented_send_and_runs_fifo(self):
        transactions = empty_transactions()
        buy = Buy("BTC", 2, datetime.datetime(2024, 1, 1), 100, "demo")
        send = Send("BTC", 1, datetime.datetime(2024, 6, 1), 200, "wallet")
        transactions.transactions = [buy, send]
        transactions.set_holdings("BTC", 1)

        with tempfile.TemporaryDirectory() as temp_dir:
            class RouteTestConfig(config_dict["Debug"]):
                TESTING = True
                WTF_CSRF_ENABLED = False
                INSTANCE_PATH = temp_dir
                SQLALCHEMY_DATABASE_URI = f"sqlite:///{temp_dir}/test.db"

            app = create_app(RouteTestConfig, selenium=True)
            app.config["transactions"] = transactions

            with app.test_client() as client:
                response = client.post(
                    "/holdings_accounting/sends_to_sells",
                    json={
                        "asset": ["BTC"],
                        "send_uid": send.uid,
                        "event_classification": "cash_sale",
                        "proceeds_value": "200",
                        "evidence_reference": "Exchange statement row 42",
                        "auto_link": True,
                    },
                )

            self.assertEqual(200, response.status_code)
            payload = response.get_json()
            sells = [trans for trans in transactions if trans.trans_type == "sell"]
            sends = [trans for trans in transactions if trans.trans_type == "send"]

            self.assertIn("documented sale for cash", payload["message"])
            self.assertEqual(1, payload["links_created"])
            self.assertEqual([], payload["auto_link_failures"])
            self.assertEqual(1, len(sells))
            self.assertEqual(0, len(sends))
            self.assertEqual(1, len(sells[0].links))
            self.assertEqual("0", payload["difference_breakdown"]["summary"]["difference"])

            with app.app_context():
                db.drop_all()
                db.session.remove()
                db.engine.dispose()

    def test_holdings_rejects_quantity_based_send_inference(self):
        transactions = empty_transactions()
        send = Send("BTC", 1, datetime.datetime(2024, 6, 1), 200, "wallet")
        transactions.transactions = [send]

        with tempfile.TemporaryDirectory() as temp_dir:
            class RouteTestConfig(config_dict["Debug"]):
                TESTING = True
                WTF_CSRF_ENABLED = False
                INSTANCE_PATH = temp_dir
                SQLALCHEMY_DATABASE_URI = f"sqlite:///{temp_dir}/test.db"

            app = create_app(RouteTestConfig, selenium=True)
            app.config["transactions"] = transactions

            with app.test_client() as client:
                response = client.post(
                    "/holdings_accounting/sends_to_sells",
                    json={
                        "asset": ["BTC"],
                        "quantity": "1",
                        "auto_link": True,
                    },
                )

            self.assertEqual(400, response.status_code)
            self.assertIn("Select the exact send row", response.get_json()["message"])
            self.assertEqual([send], transactions.transactions)
            self.assertEqual([], transactions.conversions)

            with app.app_context():
                db.drop_all()
                db.session.remove()
                db.engine.dispose()

    def test_holdings_can_apply_professional_conservative_fallback_to_exact_send(self):
        transactions = empty_transactions()
        send = Send("BTC", 0.5, datetime.datetime(2024, 6, 1), 1000, "wallet.csv")
        transactions.transactions = [send]
        transactions.set_holdings("BTC", 0)

        with tempfile.TemporaryDirectory() as temp_dir:
            class RouteTestConfig(config_dict["Debug"]):
                TESTING = True
                WTF_CSRF_ENABLED = False
                INSTANCE_PATH = temp_dir
                SQLALCHEMY_DATABASE_URI = f"sqlite:///{temp_dir}/test.db"

            app = create_app(RouteTestConfig, selenium=True)
            app.config["transactions"] = transactions

            with app.test_client() as client:
                response = client.post(
                    "/holdings_accounting/sends_to_sells",
                    json={
                        "asset": ["BTC"],
                        "send_uid": send.uid,
                        "event_classification": "conservative_unknown_disposition",
                        "proceeds_method": "disposition_date_fmv",
                        "proceeds_value": "500",
                        "evidence_reference": "Wallet row 17; price workpaper BTC-2024-06-01",
                        "reviewer_name": "Jamie Reviewer, CPA",
                        "conservative_max_gain": True,
                        "professional_attestation": True,
                        "auto_link": True,
                    },
                )

            self.assertEqual(200, response.status_code)
            payload = response.get_json()
            self.assertIn("$0-basis short-term assumption", payload["message"])
            self.assertEqual([], [row for row in transactions if row.trans_type == "send"])

            form_rows = get_form_8949_report_rows(transactions)
            self.assertEqual(1, len(form_rows))
            self.assertEqual("short", form_rows[0]["term"])
            self.assertEqual("", form_rows[0]["date_acquired"])
            self.assertAlmostEqual(500, form_rows[0]["proceeds"])
            self.assertAlmostEqual(0, form_rows[0]["cost_basis"])
            self.assertAlmostEqual(500, form_rows[0]["gain_loss"])
            self.assertEqual("cpa_conservative_short_term", form_rows[0]["acquisition_date_method"])

            review = next(
                row
                for row in transactions.work_order_reviews
                if row.get("decision") == "conservative_max_gain"
            )
            self.assertEqual("cpa_reviewed_position", review["resolution_status"])
            self.assertEqual("unknown_zero_for_review", review["basis_method"])
            self.assertEqual("Yes", review["calculation_applied"])
            self.assertIn("may overstate tax", review["assumption_disclosure"])

            with app.app_context():
                db.drop_all()
                db.session.remove()
                db.engine.dispose()

    def test_holdings_conservative_fallback_requires_professional_attestation_before_mutation(self):
        transactions = empty_transactions()
        send = Send("BTC", 0.5, datetime.datetime(2024, 6, 1), 1000, "wallet.csv")
        transactions.transactions = [send]

        with tempfile.TemporaryDirectory() as temp_dir:
            class RouteTestConfig(config_dict["Debug"]):
                TESTING = True
                WTF_CSRF_ENABLED = False
                INSTANCE_PATH = temp_dir
                SQLALCHEMY_DATABASE_URI = f"sqlite:///{temp_dir}/test.db"

            app = create_app(RouteTestConfig, selenium=True)
            app.config["transactions"] = transactions

            with app.test_client() as client:
                response = client.post(
                    "/holdings_accounting/sends_to_sells",
                    json={
                        "asset": ["BTC"],
                        "send_uid": send.uid,
                        "event_classification": "conservative_unknown_disposition",
                        "proceeds_method": "disposition_date_fmv",
                        "proceeds_value": "500",
                        "evidence_reference": "Wallet row 17",
                        "reviewer_name": "Jamie Reviewer, CPA",
                        "conservative_max_gain": True,
                    },
                )

            self.assertEqual(400, response.status_code)
            self.assertIn("Confirm the professional", response.get_json()["message"])
            self.assertEqual([send], transactions.transactions)
            self.assertEqual([], transactions.conversions)

            with app.app_context():
                db.drop_all()
                db.session.remove()
                db.engine.dispose()

    def test_holdings_conservative_fallback_preserves_documented_fifo_basis(self):
        transactions = empty_transactions()
        buy = Buy("BTC", 0.2, datetime.datetime(2024, 1, 1), 100, "exchange.csv")
        send = Send("BTC", 0.5, datetime.datetime(2024, 6, 1), 1000, "wallet.csv")
        transactions.transactions = [buy, send]
        transactions.set_holdings("BTC", 0)

        with tempfile.TemporaryDirectory() as temp_dir:
            class RouteTestConfig(config_dict["Debug"]):
                TESTING = True
                WTF_CSRF_ENABLED = False
                INSTANCE_PATH = temp_dir
                SQLALCHEMY_DATABASE_URI = f"sqlite:///{temp_dir}/test.db"

            app = create_app(RouteTestConfig, selenium=True)
            app.config["transactions"] = transactions

            with app.test_client() as client:
                response = client.post(
                    "/holdings_accounting/sends_to_sells",
                    json={
                        "asset": ["BTC"],
                        "send_uid": send.uid,
                        "event_classification": "conservative_unknown_disposition",
                        "proceeds_method": "disposition_date_fmv",
                        "proceeds_value": "500",
                        "evidence_reference": "Wallet row 17; price workpaper BTC-2024-06-01",
                        "reviewer_name": "Jamie Reviewer, CPA",
                        "conservative_max_gain": True,
                        "professional_attestation": True,
                        "auto_link": True,
                    },
                )

            self.assertEqual(200, response.status_code)
            rows = get_form_8949_report_rows(transactions)
            self.assertEqual(2, len(rows))
            documented = next(row for row in rows if row["buy_uid"] == buy.uid)
            conservative = next(
                row for row in rows
                if row["acquisition_date_method"] == "cpa_conservative_short_term"
            )
            self.assertAlmostEqual(0.2, documented["quantity"])
            self.assertAlmostEqual(20, documented["cost_basis"])
            self.assertAlmostEqual(0.3, conservative["quantity"])
            self.assertAlmostEqual(0, conservative["cost_basis"])
            self.assertAlmostEqual(500, sum(row["proceeds"] for row in rows))

            with app.app_context():
                db.drop_all()
                db.session.remove()
                db.engine.dispose()

    def test_auto_link_asset_normalizes_whitespace_all_time_year(self):
        transactions = empty_transactions()
        transactions.transactions = [
            Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "demo"),
        ]

        with tempfile.TemporaryDirectory() as temp_dir:
            class RouteTestConfig(config_dict["Debug"]):
                TESTING = True
                WTF_CSRF_ENABLED = False
                INSTANCE_PATH = temp_dir
                SQLALCHEMY_DATABASE_URI = f"sqlite:///{temp_dir}/test.db"

            app = create_app(RouteTestConfig, selenium=True)
            app.config["transactions"] = transactions

            for payload in (
                {"asset": ["BTC"], "algo": "fifo", "year": " All Time "},
                {"asset": ["BTC"], "algo": "fifo"},
            ):
                with app.test_client() as client, patch(
                    "app.auto_link.routes.AutoLinkService.auto_link",
                    return_value="ok",
                ) as auto_link_mock:
                    response = client.post(
                        "/auto_link/auto_link_asset",
                        json=payload,
                    )

                self.assertEqual(200, response.status_code)
                self.assertEqual("ok", response.get_json())
                self.assertIsNone(auto_link_mock.call_args.kwargs["year"])
                self.assertTrue(auto_link_mock.call_args.kwargs["rebuild"])

            with app.app_context():
                db.drop_all()
                db.session.remove()
                db.engine.dispose()

    def test_auto_link_all_fifo_normalizes_empty_and_missing_year_values(self):
        transactions = empty_transactions()
        transactions.transactions = [
            Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "demo"),
        ]

        with tempfile.TemporaryDirectory() as temp_dir:
            class RouteTestConfig(config_dict["Debug"]):
                TESTING = True
                WTF_CSRF_ENABLED = False
                INSTANCE_PATH = temp_dir
                SQLALCHEMY_DATABASE_URI = f"sqlite:///{temp_dir}/test.db"

            app = create_app(RouteTestConfig, selenium=True)
            app.config["transactions"] = transactions

            for payload in ({"year": ""}, {}):
                with app.test_client() as client, patch(
                    "app.auto_link.routes.AutoLinkService.auto_link_unlinked_sales",
                    return_value={
                        "message": "ok",
                        "links_created": 0,
                        "fixed_assets": [],
                        "failures": [],
                        "algo": "fifo",
                        "year": None,
                    },
                ) as auto_link_mock:
                    response = client.post("/auto_link/auto_link_all_fifo", json=payload)

                self.assertEqual(200, response.status_code)
                self.assertIsNone(auto_link_mock.call_args.kwargs["year"])

            with app.app_context():
                db.drop_all()
                db.session.remove()
                db.engine.dispose()

    def test_auto_link_page_uses_explicit_year_option_values(self):
        transactions = empty_transactions()
        transactions.transactions = [
            Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "demo"),
        ]

        with tempfile.TemporaryDirectory() as temp_dir:
            class RouteTestConfig(config_dict["Debug"]):
                TESTING = True
                WTF_CSRF_ENABLED = False
                INSTANCE_PATH = temp_dir
                SQLALCHEMY_DATABASE_URI = f"sqlite:///{temp_dir}/test.db"

            app = create_app(RouteTestConfig, selenium=True)
            app.config["transactions"] = transactions

            with app.test_client() as client:
                response = client.get("/auto_link/")

            self.assertEqual(200, response.status_code)
            self.assertIn(b'<option value="All Time">All Time</option>', response.data)
            self.assertIn(b'<option value="2024">2024</option>', response.data)
            self.assertNotIn(b"<option> All Time </option>", response.data)

            with app.app_context():
                db.drop_all()
                db.session.remove()
                db.engine.dispose()

    def test_bulk_holdings_sets_primary_asset_and_zeroes_others(self):
        transactions = empty_transactions()
        transactions.transactions = [
            Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "demo"),
            Buy("ETH", 2, datetime.datetime(2024, 1, 2), 50, "demo"),
            Buy("BCH", 3, datetime.datetime(2024, 1, 3), 25, "demo"),
        ]

        with tempfile.TemporaryDirectory() as temp_dir:
            class RouteTestConfig(config_dict["Debug"]):
                TESTING = True
                WTF_CSRF_ENABLED = False
                INSTANCE_PATH = temp_dir
                SQLALCHEMY_DATABASE_URI = f"sqlite:///{temp_dir}/test.db"

            app = create_app(RouteTestConfig, selenium=True)
            app.config["transactions"] = transactions

            with app.test_client() as client:
                response = client.post(
                    "/holdings_accounting/bulk_holdings",
                    json={
                        "primary_asset": "BTC",
                        "primary_quantity": "0.12345678",
                    },
                )

            self.assertEqual(200, response.status_code)
            payload = response.get_json()
            self.assertEqual("0.12345678", payload["primary_quantity"])
            self.assertEqual(0.12345678, transactions.get_holdings("BTC"))
            self.assertEqual(0, transactions.get_holdings("ETH"))
            self.assertEqual(0, transactions.get_holdings("BCH"))
            self.assertEqual(["BCH", "ETH"], payload["zeroed_assets"])

            with app.app_context():
                db.drop_all()
                db.session.remove()
                db.engine.dispose()

    def test_bulk_holdings_accepts_multiple_current_holdings_and_zeroes_rest(self):
        transactions = empty_transactions()
        transactions.transactions = [
            Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "demo"),
            Buy("ETH", 2, datetime.datetime(2024, 1, 2), 50, "demo"),
            Buy("BCH", 3, datetime.datetime(2024, 1, 3), 25, "demo"),
            Buy("LTC", 4, datetime.datetime(2024, 1, 4), 10, "demo"),
        ]

        with tempfile.TemporaryDirectory() as temp_dir:
            class RouteTestConfig(config_dict["Debug"]):
                TESTING = True
                WTF_CSRF_ENABLED = False
                INSTANCE_PATH = temp_dir
                SQLALCHEMY_DATABASE_URI = f"sqlite:///{temp_dir}/test.db"

            app = create_app(RouteTestConfig, selenium=True)
            app.config["transactions"] = transactions

            with app.test_client() as client:
                response = client.post(
                    "/holdings_accounting/bulk_holdings",
                    json={
                        "holdings": [
                            {"asset": "BTC", "quantity": "0.5"},
                            {"asset": "eth", "quantity": "1.25"},
                        ],
                    },
                )

            self.assertEqual(200, response.status_code)
            payload = response.get_json()
            self.assertEqual(
                [
                    {"asset": "BTC", "quantity": "0.5"},
                    {"asset": "ETH", "quantity": "1.25"},
                ],
                payload["declared_holdings"],
            )
            self.assertEqual(0.5, transactions.get_holdings("BTC"))
            self.assertEqual(1.25, transactions.get_holdings("ETH"))
            self.assertEqual(0, transactions.get_holdings("BCH"))
            self.assertEqual(0, transactions.get_holdings("LTC"))
            self.assertEqual(["BCH", "LTC"], payload["zeroed_assets"])

            with app.app_context():
                db.drop_all()
                db.session.remove()
                db.engine.dispose()

    def test_bulk_holdings_rejects_duplicate_rows(self):
        transactions = empty_transactions()
        transactions.transactions = [
            Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "demo"),
        ]

        with tempfile.TemporaryDirectory() as temp_dir:
            class RouteTestConfig(config_dict["Debug"]):
                TESTING = True
                WTF_CSRF_ENABLED = False
                INSTANCE_PATH = temp_dir
                SQLALCHEMY_DATABASE_URI = f"sqlite:///{temp_dir}/test.db"

            app = create_app(RouteTestConfig, selenium=True)
            app.config["transactions"] = transactions

            with app.test_client() as client:
                response = client.post(
                    "/holdings_accounting/bulk_holdings",
                    json={
                        "holdings": [
                            {"asset": "BTC", "quantity": "0.5"},
                            {"asset": "btc", "quantity": "0.25"},
                        ],
                    },
                )

            self.assertEqual(400, response.status_code)
            self.assertIn("BTC appears more than once", response.get_json()["message"])

            with app.app_context():
                db.drop_all()
                db.session.remove()
                db.engine.dispose()

    def test_data_source_summary_flags_likely_full_history_overlap(self):
        transactions = empty_transactions()
        transactions.transactions = [
            Buy("BCH", 1, datetime.datetime(2024, 1, 1), 100, "exchange_all_activity.csv"),
            Sell("BCH", 0.5, datetime.datetime(2024, 2, 1), 200, "exchange_all_activity.csv"),
            Buy("BCH", 1, datetime.datetime(2024, 1, 1), 100, "exchange_all_activity_2025.csv"),
            Sell("BCH", 0.5, datetime.datetime(2024, 2, 1), 200, "exchange_all_activity_2025.csv"),
        ]

        summary = import_routes._data_source_summary(transactions)

        self.assertEqual(1, summary["source_overlap_count"])
        self.assertEqual("Likely full-history overlap", summary["source_overlaps"][0]["status"])
        self.assertEqual(
            {"Potential overlap"},
            {source["status"] for source in summary["sources"]},
        )

    def test_audit_readiness_surfaces_source_overlap_review(self):
        transactions = empty_transactions()
        transactions.transactions = [
            Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "exchange_all_activity.csv"),
            Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "exchange_all_activity_2025.csv"),
        ]
        transactions.set_holdings("BTC", 2)

        readiness = get_audit_readiness_summary(transactions)

        self.assertFalse(readiness["is_ready"])
        self.assertEqual(1, readiness["metrics"]["source_overlaps"])
        self.assertTrue(any("overlapping source file" in warning for warning in readiness["warnings"]))
        self.assertEqual(1, len(readiness["missing_records"]["source_overlaps"]))
        self.assertTrue(any(group["key"] == "source_overlaps" for group in readiness["blocker_groups"]))
        source_group = next(group for group in readiness["blocker_groups"] if group["key"] == "source_overlaps")
        self.assertEqual("Review sources", source_group["action_label"])

    def test_audit_readiness_prioritizes_holdings_before_basis_review(self):
        transactions = empty_transactions()
        transactions.transactions = [
            Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "exchange"),
            Sell("BTC", 0.5, datetime.datetime(2024, 2, 1), 150, "exchange"),
        ]

        readiness = get_audit_readiness_summary(transactions)

        self.assertFalse(readiness["is_ready"])
        self.assertEqual("holdings", readiness["blocker_groups"][0]["key"])
        self.assertEqual("Enter holdings", readiness["primary_action"]["label"])
        self.assertIn("Declare current holdings", readiness["next_action"])
        self.assertTrue(readiness["missing_records"]["basis_summary"])

    def test_leave_basis_unresolved_keeps_export_not_ready_with_research_status(self):
        transactions = empty_transactions()
        transactions.transactions = [
            Sell("LTC", 1, datetime.datetime(2019, 1, 7), 40, "exchange"),
        ]
        transactions.set_holdings("LTC", 0)

        with tempfile.TemporaryDirectory() as temp_dir:
            class RouteTestConfig(config_dict["Debug"]):
                TESTING = True
                WTF_CSRF_ENABLED = False
                INSTANCE_PATH = temp_dir
                SQLALCHEMY_DATABASE_URI = f"sqlite:///{temp_dir}/test.db"

            app = create_app(RouteTestConfig, selenium=True)
            app.config["transactions"] = transactions

            with app.test_client() as client:
                response = client.post(
                    "/holdings_accounting/leave_basis_unresolved",
                    json={
                        "asset": ["LTC"],
                        "decision": "zero_basis_cpa_review",
                        "note": "User will investigate source records later.",
                    },
                )

            self.assertEqual(200, response.status_code)
            payload = response.get_json()
            self.assertIn("unknown basis treated as $0 for CPA review", payload["message"])
            self.assertEqual(
                "zero_basis_cpa_review",
                transactions.get_basis_review_note("LTC")["status"],
            )
            self.assertEqual(
                "User will investigate source records later.",
                transactions.get_basis_review_note("LTC")["note"],
            )

            readiness = get_audit_readiness_summary(transactions)
            self.assertFalse(readiness["is_ready"])
            self.assertEqual("Unknown basis treated as $0 for CPA review", readiness["missing_records"]["basis"][0]["status"])
            self.assertTrue(
                any("Missing basis documented for draft/CPA review" in blocker for blocker in readiness["blockers"])
            )
            basis_group = next(group for group in readiness["blocker_groups"] if group["key"] == "missing_basis")
            self.assertEqual("Documented for review", basis_group["status"])
            self.assertEqual("/holdings_accounting/?guided=1&mode=reconcile", basis_group["action_url"])

            with app.app_context():
                db.drop_all()
                db.session.remove()
                db.engine.dispose()

    def test_tax_filing_review_confirms_suggested_totals_or_marks_research(self):
        transactions = empty_transactions()
        with tempfile.TemporaryDirectory() as temp_dir:
            evidence_path = Path(temp_dir) / "2024_crypto_tax_workbook.csv"
            evidence_path.write_text(
                "Year,Reported Proceeds,Reported Cost Basis,Reported Gain Loss,Tax Paid\n"
                "2024,300,100,200,25\n",
                encoding="utf-8",
            )
            transactions.set_tax_evidence_record(
                year=2024,
                evidence_type="crypto_workbook",
                evidence_label=evidence_path.name,
                evidence_path=str(evidence_path),
            )

            class RouteTestConfig(config_dict["Debug"]):
                TESTING = True
                WTF_CSRF_ENABLED = False
                INSTANCE_PATH = temp_dir
                SQLALCHEMY_DATABASE_URI = f"sqlite:///{temp_dir}/test.db"

            app = create_app(RouteTestConfig, selenium=True)
            app.config["transactions"] = transactions

            with app.test_client() as client:
                confirm_response = client.post(
                    "/tax_filing_review/suggested_totals/confirm",
                    data={
                        "year": "2024",
                        "reported_proceeds": "300",
                        "reported_cost_basis": "100",
                        "reported_gain_loss": "200",
                        "tax_paid": "25",
                        "source_reference": evidence_path.name,
                        "notes": "Reviewed source row.",
                    },
                )
                research_response = client.post(
                    "/tax_filing_review/suggested_totals/research",
                    data={
                        "year": "2023",
                        "source_reference": "2023 filed return PDF",
                        "notes": "Needs source review.",
                    },
                )

            self.assertEqual(302, confirm_response.status_code)
            self.assertEqual(302, research_response.status_code)
            confirmed = transactions.get_tax_year_record(2024)
            self.assertEqual(300.0, confirmed["reported_proceeds"])
            self.assertEqual(25.0, confirmed["tax_paid"])
            self.assertIn("Confirmed from Gainz", confirmed["notes"])
            needs_research = transactions.get_tax_year_record(2023)
            self.assertIsNone(needs_research)
            research_items = [
                record
                for record in transactions.tax_evidence_records
                if record["year"] == 2023 and "filed totals need research" in record["evidence_label"]
            ]
            self.assertEqual(1, len(research_items))
            self.assertIn("No filed totals were recorded", research_items[0]["notes"])
            self.assertIn("Needs source review.", research_items[0]["notes"])

            with app.app_context():
                db.drop_all()
                db.session.remove()
                db.engine.dispose()

    def test_tax_evidence_scan_preset_limits_file_types_and_keywords(self):
        transactions = empty_transactions()
        with tempfile.TemporaryDirectory() as temp_dir:
            evidence_dir = Path(temp_dir) / "evidence"
            evidence_dir.mkdir()
            (evidence_dir / "coinbase_transactions_2024.csv").write_text("Date,Asset,Amount\n", encoding="utf-8")
            (evidence_dir / "2024_filed_return.pdf").write_text("synthetic return", encoding="utf-8")
            (evidence_dir / "random_notes.txt").write_text("coinbase transaction notes", encoding="utf-8")

            class RouteTestConfig(config_dict["Debug"]):
                TESTING = True
                WTF_CSRF_ENABLED = False
                INSTANCE_PATH = temp_dir
                SQLALCHEMY_DATABASE_URI = f"sqlite:///{temp_dir}/test.db"

            app = create_app(RouteTestConfig, selenium=True)
            app.config["transactions"] = transactions

            with app.test_client() as client, patch(
                "app.tax_filing_review.routes._scan_location_choices",
                return_value={"detected_taxes": str(evidence_dir)},
            ):
                response = client.post(
                    "/tax_filing_review/scan_evidence_folder?guided=1",
                    data={
                        "guided": "1",
                        "scan_location": "detected_taxes",
                        "scan_preset": "transaction_csvs",
                        "evidence_years": "2024",
                        "recursive": "1",
                    },
                )
                redirected_page = client.get(response.location)

            self.assertEqual(302, response.status_code)
            self.assertIn("saved_evidence=1", response.location)
            self.assertIn("guided=1", response.location)
            self.assertIn("scan_preset=transaction_csvs", response.location)
            self.assertIn("evidence_years=2024", response.location)
            self.assertIn("scan_completed_years=2024", response.location)
            self.assertIn(b'value="transaction_csvs" selected', redirected_page.data)
            self.assertIn(b'value="2024"', redirected_page.data)
            self.assertIn(b"2024 scan complete.", redirected_page.data)
            self.assertEqual(1, len(transactions.tax_evidence_records))
            self.assertEqual("coinbase_transactions_2024.csv", transactions.tax_evidence_records[0]["evidence_label"])
            self.assertIn("Transaction CSVs only", transactions.tax_evidence_records[0]["notes"])

            with app.app_context():
                db.drop_all()
                db.session.remove()
                db.engine.dispose()

    def test_tax_evidence_scan_skips_gainz_artifact_folders_by_default(self):
        transactions = empty_transactions()
        with tempfile.TemporaryDirectory() as temp_dir:
            evidence_dir = Path(temp_dir) / "Taxes"
            evidence_dir.mkdir()
            (evidence_dir / "2023_filed_return.pdf").write_text("synthetic return", encoding="utf-8")
            archive_dir = evidence_dir / "90_Gainz_Product_Review_Archive" / "Artifacts"
            archive_dir.mkdir(parents=True)
            (archive_dir / "2023_filed_return_duplicate.pdf").write_text("duplicate return", encoding="utf-8")
            packet_dir = evidence_dir / "gainz_audit_packet_DRAFT_2026-06-27"
            packet_dir.mkdir()
            (packet_dir / "PACKET_STATUS.md").write_text("generated packet", encoding="utf-8")

            class RouteTestConfig(config_dict["Debug"]):
                TESTING = True
                WTF_CSRF_ENABLED = False
                INSTANCE_PATH = temp_dir
                SQLALCHEMY_DATABASE_URI = f"sqlite:///{temp_dir}/test.db"

            app = create_app(RouteTestConfig, selenium=True)
            app.config["transactions"] = transactions

            with app.test_client() as client, patch(
                "app.tax_filing_review.routes._scan_location_choices",
                return_value={"detected_taxes": str(evidence_dir)},
            ):
                response = client.post(
                    "/tax_filing_review/scan_evidence_folder",
                    data={
                        "scan_location": "detected_taxes",
                        "recursive": "1",
                        "evidence_file_types": [".pdf", ".md"],
                    },
                )

            self.assertEqual(302, response.status_code)
            self.assertIn("saved_evidence=1", response.location)
            self.assertIn("skipped_evidence_folders=2", response.location)
            self.assertEqual(["2023_filed_return.pdf"], [
                record["evidence_label"] for record in transactions.tax_evidence_records
            ])

            with app.app_context():
                db.drop_all()
                db.session.remove()
                db.engine.dispose()

    def test_tax_filing_review_imports_filed_totals_csv_with_unicode_hyphen_headers(self):
        transactions = empty_transactions()
        csv_text = (
            "Tax Year,Short\u2011term Proceeds,Short\u2011term Cost Basis,Short\u2011term Gain,"
            "Long\u2011term Proceeds,Long\u2011term Cost Basis,Long\u2011term Gain\n"
            "2021,$0.00,$0.00,$0.00,$0.00,$0.00,$0.00\n"
            "2022,\"$18,338.19\",\"$18,453.11\",($114.93),\"$112,689.71\",\"$51,005.76\",\"$61,683.96\"\n"
            "missing,,,,,,\n"
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            class RouteTestConfig(config_dict["Debug"]):
                TESTING = True
                WTF_CSRF_ENABLED = False
                INSTANCE_PATH = temp_dir
                SQLALCHEMY_DATABASE_URI = f"sqlite:///{temp_dir}/test.db"

            app = create_app(RouteTestConfig, selenium=True)
            app.config["transactions"] = transactions

            with app.test_client() as client:
                response = client.post(
                    "/tax_filing_review/import_csv",
                    data={
                        "csv_file": (
                            io.BytesIO(csv_text.encode("utf-8")),
                            "filed_totals.csv",
                        ),
                    },
                    content_type="multipart/form-data",
                )

            self.assertEqual(302, response.status_code)
            self.assertIn("imported_tax_rows=2", response.location)
            self.assertIn("skipped_tax_rows=1", response.location)
            zero_record = transactions.get_tax_year_record(2021)
            self.assertEqual(0.0, zero_record["reported_gain_loss"])
            record = transactions.get_tax_year_record(2022)
            self.assertEqual(131027.90, round(record["reported_proceeds"], 2))
            self.assertEqual(69458.87, round(record["reported_cost_basis"], 2))
            self.assertEqual(61569.03, round(record["reported_gain_loss"], 2))
            self.assertIsNone(record["tax_paid"])
            self.assertEqual(
                {2021, 2022},
                {record["year"] for record in transactions.tax_evidence_records},
            )

            with app.app_context():
                db.drop_all()
                db.session.remove()
                db.engine.dispose()

    def test_export_routes_use_requested_output_folder(self):
        transactions = empty_transactions()
        buy = Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "demo")
        sell = Sell("BTC", 1, datetime.datetime(2024, 6, 1), 300, "demo")
        sell.link_transaction(buy, 1)
        transactions.transactions = [buy, sell]
        transactions.set_holdings("BTC", 0)

        with tempfile.TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir) / "tax-review-output"

            class ExportRouteTestConfig(config_dict["Debug"]):
                TESTING = True
                WTF_CSRF_ENABLED = False
                INSTANCE_PATH = temp_dir
                SQLALCHEMY_DATABASE_URI = f"sqlite:///{temp_dir}/test.db"
                EXPORT_FOLDER = str(output_dir)
                AUDIT_PACKET_FOLDER = str(output_dir)

            app = create_app(ExportRouteTestConfig, selenium=True)
            app.config["transactions"] = transactions

            with app.test_client() as client:
                export_response = client.post(
                    "/export/save",
                    json={"output_location": "audit_packets", "draft_acknowledged": True},
                )
                packet_response = client.post(
                    "/export/audit_packet",
                    json={"output_location": "audit_packets", "draft_acknowledged": True, "guided": "1"},
                )
                success_response = client.get(packet_response.get_json()["success_url"])

            export_payload = export_response.get_json()
            packet_payload = packet_response.get_json()

            self.assertEqual(200, export_response.status_code)
            self.assertEqual(200, packet_response.status_code)
            self.assertEqual(output_dir.resolve(), Path(export_payload["output_dir"]))
            self.assertEqual(output_dir.resolve(), Path(packet_payload["output_dir"]))
            self.assertIn("/export/packet_success?guided=1", packet_payload["success_url"])
            self.assertEqual(output_dir.resolve(), Path(export_payload["path"]).parent)
            self.assertEqual(output_dir.resolve(), Path(packet_payload["path"]).parent)
            self.assertTrue(Path(export_payload["path"]).exists())
            self.assertTrue((Path(packet_payload["path"]) / "03_manifests" / "audit_packet_summary.json").exists())
            self.assertTrue(Path(export_payload["path"]).name.startswith("DRAFT_"))
            self.assertIn("gainz_audit_packet_DRAFT_", Path(packet_payload["path"]).name)
            direct_workbooks = [path.resolve() for path in output_dir.glob("DRAFT_Export_*.xlsx")]
            self.assertEqual([Path(export_payload["path"]).resolve()], direct_workbooks)
            workbook = load_workbook(export_payload["path"], read_only=True)
            self.assertIn("Packet Status", workbook.sheetnames)
            self.assertEqual(
                "DRAFT - NOT FILING READY",
                workbook["Packet Status"]["A1"].value,
            )
            workbook.close()

            self.assertEqual(200, success_response.status_code)
            self.assertIn(b"Audit Packet Generated", success_response.data)
            self.assertIn(b"FOR_CPAS.md", success_response.data)
            self.assertIn(b"Copy Packet Path", success_response.data)
            self.assertIn(b"Copy CPA Summary", success_response.data)
            self.assertIn(b"Open README_FIRST", success_response.data)
            self.assertIn(b"Packet Size", success_response.data)
            self.assertIn(b"Generated", success_response.data)

            with app.app_context():
                db.drop_all()
                db.session.remove()
                db.engine.dispose()

    def test_export_route_saves_work_order_review_state(self):
        transactions = empty_transactions()
        transactions.transactions = [
            Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "source.csv")
        ]

        with tempfile.TemporaryDirectory() as temp_dir:
            class ExportRouteTestConfig(config_dict["Debug"]):
                TESTING = True
                WTF_CSRF_ENABLED = False
                INSTANCE_PATH = temp_dir
                SQLALCHEMY_DATABASE_URI = f"sqlite:///{temp_dir}/test.db"

            app = create_app(ExportRouteTestConfig, selenium=True)
            app.config["transactions"] = transactions

            readiness = get_audit_readiness_summary(transactions)
            work_items = [
                row for row in reconciliation_work_order_rows(readiness, transactions)
                if row.get("blocker_type") != "No open blockers"
            ]
            self.assertTrue(work_items)
            item_id = work_items[0]["item_id"]

            with app.test_client() as client:
                response = client.post(
                    "/export/work_order_review",
                    json={
                        "item_id": item_id,
                        "decision": "sent_to_cpa",
                        "note": "Asked CPA to review this item.",
                        "cpa_question": "Should this remain unresolved for CPA review?",
                    },
                )

            self.assertEqual(200, response.status_code)
            record = transactions.get_work_order_review(item_id)
            self.assertIsNotNone(record)
            self.assertEqual("sent_to_cpa", record["decision"])
            self.assertEqual("Asked CPA to review this item.", record["note"])
            self.assertEqual(
                "Should this remain unresolved for CPA review?",
                record["cpa_question"],
            )
            self.assertIn(
                "Updated work order item: Sent to CPA",
                transactions.saved_descriptions,
            )

            with app.app_context():
                db.drop_all()
                db.session.remove()
                db.engine.dispose()

    def test_guided_review_queue_saves_and_advances_work_order_items(self):
        transactions = empty_transactions()
        transactions.transactions = [
            Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "source.csv")
        ]

        with tempfile.TemporaryDirectory() as temp_dir:
            class ExportRouteTestConfig(config_dict["Debug"]):
                TESTING = True
                WTF_CSRF_ENABLED = False
                INSTANCE_PATH = temp_dir
                SQLALCHEMY_DATABASE_URI = f"sqlite:///{temp_dir}/test.db"

            app = create_app(ExportRouteTestConfig, selenium=True)
            app.config["transactions"] = transactions

            with app.test_client() as client:
                response = client.get("/export/review_queue?guided=1")

            self.assertEqual(200, response.status_code)
            self.assertIn(b"Step 4.2: Guided Review Queue", response.data)
            self.assertIn(b"Guided Review Queue", response.data)
            self.assertIn(b"Current holdings missing", response.data)
            self.assertIn(b"Enter current BTC holdings", response.data)
            self.assertIn(b"What amount of BTC do you currently hold", response.data)
            self.assertIn(b"Current holdings are entered", response.data)
            self.assertIn(b"Gap Investigator", response.data)
            self.assertIn(b"What Gainz Knows", response.data)
            self.assertIn(b"What Gainz Does Not Know", response.data)
            self.assertIn(b"Safe Outcomes", response.data)
            self.assertIn(b"It is valid to choose", response.data)
            self.assertIn(b"Return to Reports & Export", response.data)
            self.assertIn(b"/export/?guided=1", response.data)
            response_text = response.data.decode("utf-8")
            self.assertLess(
                response_text.index("Choose the resolution outcome"),
                response_text.index("Gap Investigator"),
            )

            readiness = get_audit_readiness_summary(transactions)
            rows = [
                row for row in reconciliation_work_order_rows(readiness, transactions)
                if row.get("blocker_type") != "No open blockers"
            ]
            self.assertTrue(rows)
            item_id = rows[0]["item_id"]

            with app.test_client() as client:
                save_response = client.post(
                    "/export/review_queue/save",
                    data={
                        "item_id": item_id,
                        "decision": "needs_research",
                        "note": "User will research source records.",
                        "cpa_question": "What source records should be checked next?",
                    },
                    follow_redirects=True,
                )

            self.assertEqual(200, save_response.status_code)
            self.assertIn(b"No Undecided Queue Items Remain", save_response.data)
            self.assertIn(b"Deferred Items", save_response.data)
            self.assertIn(b"/export/?guided=1#packet_preview", save_response.data)
            self.assertIn(b"generate or refresh the packet", save_response.data)
            record = transactions.get_work_order_review(item_id)
            self.assertEqual("needs_research", record["decision"])
            self.assertEqual("User will research source records.", record["note"])
            self.assertEqual(
                "What source records should be checked next?",
                record["cpa_question"],
            )

            with app.app_context():
                db.drop_all()
                db.session.remove()
                db.engine.dispose()

    def test_guided_review_queue_explains_missing_basis_item_clearly(self):
        transactions = empty_transactions()
        transactions.transactions = [
            Sell("BCH", 0.5, datetime.datetime(2024, 2, 1), 200, "cash_app_report.csv")
        ]
        transactions.set_holdings("BCH", 0)

        with tempfile.TemporaryDirectory() as temp_dir:
            class ExportRouteTestConfig(config_dict["Debug"]):
                TESTING = True
                WTF_CSRF_ENABLED = False
                INSTANCE_PATH = temp_dir
                SQLALCHEMY_DATABASE_URI = f"sqlite:///{temp_dir}/test.db"

            app = create_app(ExportRouteTestConfig, selenium=True)
            app.config["transactions"] = transactions

            with app.test_client() as client:
                response = client.get("/export/review_queue?guided=1")
                conservative_response = client.get(
                    "/export/review_queue?guided=1&asset=BCH&decision=conservative_max_gain"
                )

            self.assertEqual(200, response.status_code)
            self.assertIn(b"Resolve BCH missing cost basis", response.data)
            self.assertIn(b"Gainz found BCH sales, but not the BCH acquisition history", response.data)
            self.assertIn(b"cash_app_report.csv", response.data)
            self.assertIn(b"Import missing BCH acquisition records", response.data)
            self.assertIn(b"This BCH came from a fork/airdrop", response.data)
            self.assertIn(b"Already included in filed tax totals", response.data)
            self.assertIn(b"Use $0 basis for a conservative full-proceeds gain", response.data)
            self.assertIn(b"Send this BCH gap to CPA", response.data)
            self.assertIn(b"CPA review options", response.data)
            self.assertIn(b"Reconstruct basis from records", response.data)
            self.assertIn(b"Correct source classification", response.data)
            self.assertIn(b"Apply a professionally directed basis adjustment", response.data)
            self.assertNotIn(b"Keep as owner transfer", response.data)
            self.assertNotIn(b"Decide what to do with this work order item", response.data)
            self.assertEqual(200, conservative_response.status_code)
            self.assertIn(b"Professional Resolution Worksheet", conservative_response.data)
            self.assertIn(b"Conservative unknown-basis assumption", conservative_response.data)
            self.assertIn(b"generally makes the unresolved net", conservative_response.data)
            self.assertIn(b"name=\"decision\" value=\"conservative_max_gain\"", conservative_response.data)

            with app.app_context():
                db.drop_all()
                db.session.remove()
                db.engine.dispose()

    def test_guided_review_queue_splits_supported_and_unresolved_sale_values(self):
        transactions = empty_transactions()
        buy = Buy("BCH", 0.2, datetime.datetime(2021, 3, 15), 125, "coinbase.csv")
        buy.set_economics(gross_usd_total=25, fee=0.5, net_usd_total=25.5)
        sell = Sell("BCH", 0.5, datetime.datetime(2024, 2, 1), 1000, "coinbase.csv")
        sell.set_economics(gross_usd_total=500, fee=5, net_usd_total=495)
        sell.link_transaction(buy, 0.2)
        transactions.transactions = [buy, sell]
        transactions.set_holdings("BCH", 0)

        app = create_app(config_dict["Debug"], selenium=True)
        app.config.update(WTF_CSRF_ENABLED=False)
        app.config["transactions"] = transactions

        readiness = get_audit_readiness_summary(transactions)
        item = next(
            row
            for row in reconciliation_work_order_rows(readiness, transactions)
            if row.get("blocker_type") == "Missing acquisition basis"
        )

        with app.test_client() as client:
            response = client.get("/export/review_queue?guided=1")
            configured = client.post(
                "/export/review_queue/save",
                data={
                    "item_id": item["item_id"],
                    "decision": "resolved",
                    "workflow_action": "configure",
                },
            )

        self.assertEqual(200, response.status_code)
        page = response.data.decode("utf-8", errors="replace")
        self.assertIn("Supported and unresolved amounts", page)
        self.assertIn("0.20000000 BCH", page)
        self.assertIn("0.30000000 BCH", page)
        self.assertIn("$297.00 allocated net proceeds", page)
        self.assertNotIn("Professional Resolution Worksheet", page)
        configured_page = configured.data.decode("utf-8", errors="replace")
        self.assertIn("Professional Resolution Worksheet", configured_page)
        self.assertIn("Calculated suggestion: 0.30000000 BCH / 0.50000000 BCH", configured_page)
        self.assertIn(
            '<option value="allocated_source_value" selected>Calculated allocation from imported source transaction</option>',
            configured_page,
        )

    def test_cpa_resolution_reaches_ready_form_8949_and_packet_workpaper(self):
        transactions = empty_transactions()
        sell = Sell("BCH", 0.5, datetime.datetime(2024, 2, 1), 1000, "exchange.csv")
        transactions.transactions = [sell]
        transactions.set_holdings("BCH", 0)
        transactions.set_tax_year_record(
            2024,
            reported_proceeds=500,
            reported_cost_basis=125,
            reported_gain_loss=375,
            tax_paid=50,
            filing_status="Filed",
            evidence_reference="2024 filed Form 8949 and payment record",
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            class ExportRouteTestConfig(config_dict["Debug"]):
                TESTING = True
                WTF_CSRF_ENABLED = False
                INSTANCE_PATH = temp_dir
                SQLALCHEMY_DATABASE_URI = f"sqlite:///{temp_dir}/test.db"

            app = create_app(ExportRouteTestConfig, selenium=True)
            app.config["transactions"] = transactions

            before = get_audit_readiness_summary(transactions)
            missing_basis_item = next(
                row
                for row in reconciliation_work_order_rows(before, transactions)
                if row.get("blocker_type") == "Missing acquisition basis"
            )
            self.assertEqual(sell.uid, missing_basis_item["target_transaction_uid"])

            resolution_data = {
                "item_id": missing_basis_item["item_id"],
                "decision": "resolved",
                "event_classification": "cash_sale",
                "proceeds_method": "source_reported",
                "proceeds_value": "500",
                "basis_method": "documented_acquisition_cost",
                "basis_value": "125",
                "acquisition_date": "2021-03-15",
                "evidence_reference": "CPA workpaper BCH-2024-01",
                "resolution_status": "cpa_reviewed_position",
                "reviewer_name": "Jamie Reviewer",
                "reviewer_role": "cpa_ea_tax_professional",
                "professional_attestation": "yes",
                "note": "Acquisition cost reconstructed from exchange statements.",
            }
            with app.test_client() as client:
                preview = client.post(
                    "/export/review_queue/save",
                    data={**resolution_data, "workflow_action": "preview"},
                )
                self.assertEqual(200, preview.status_code)
                response = client.post(
                    "/export/review_queue/save",
                    data={
                        **resolution_data,
                        "workflow_action": "apply",
                        "preview_confirmed": "yes",
                    },
                )

            self.assertEqual(302, response.status_code)
            form_rows = get_form_8949_report_rows(transactions)
            self.assertEqual(1, len(form_rows))
            self.assertAlmostEqual(500, form_rows[0]["proceeds"])
            self.assertAlmostEqual(125, form_rows[0]["cost_basis"])
            self.assertAlmostEqual(375, form_rows[0]["gain_loss"])
            self.assertEqual(sell.uid, form_rows[0]["sell_uid"])

            after = get_audit_readiness_summary(transactions)
            self.assertEqual([], after["missing_records"]["basis"])
            self.assertTrue(after["is_ready"])
            self.assertEqual("Ready for review", after["status"])

            workpapers = cpa_resolution_workpaper_rows(after, transactions)
            self.assertEqual(1, len(workpapers))
            self.assertEqual("Yes", workpapers[0]["calculation_applied"])
            self.assertEqual("Professional direction recorded by user", workpapers[0]["resolution_status_label"])
            self.assertEqual("500.00", workpapers[0]["proceeds_value"])
            self.assertEqual("125.00", workpapers[0]["basis_value"])

            packet_dir = AuditPacketService(
                str(Path(temp_dir) / "packets"),
                str(Path(temp_dir) / "exports"),
            ).create_packet(transactions)
            workpaper_path = Path(packet_dir) / "01_reports" / "cpa_resolution_workpapers.csv"
            self.assertTrue(workpaper_path.exists())
            packet_status = (Path(packet_dir) / "PACKET_STATUS.md").read_text(encoding="utf-8")
            self.assertIn("RECONCILIATION COMPLETE - PROFESSIONAL FILING REVIEW REQUIRED", packet_status)
            self.assertNotIn("FILING-READY REVIEW PACKET", packet_status)
            self.assertIn("Professional directions recorded by user", packet_status)
            with workpaper_path.open("r", encoding="utf-8-sig", newline="") as handle:
                packet_rows = list(csv.DictReader(handle))
            self.assertEqual(1, len(packet_rows))
            self.assertEqual("Yes", packet_rows[0]["calculation_applied"])
            self.assertEqual(sell.uid, packet_rows[0]["target_transaction_uid"])

            with app.app_context():
                db.drop_all()
                db.session.remove()
                db.engine.dispose()

    def test_cpa_resolution_cannot_apply_without_professional_attestation(self):
        transactions = empty_transactions()
        sell = Sell("BCH", 0.5, datetime.datetime(2024, 2, 1), 1000, "exchange.csv")
        transactions.transactions = [sell]
        transactions.set_holdings("BCH", 0)

        with tempfile.TemporaryDirectory() as temp_dir:
            class ExportRouteTestConfig(config_dict["Debug"]):
                TESTING = True
                WTF_CSRF_ENABLED = False
                INSTANCE_PATH = temp_dir
                SQLALCHEMY_DATABASE_URI = f"sqlite:///{temp_dir}/test.db"

            app = create_app(ExportRouteTestConfig, selenium=True)
            app.config["transactions"] = transactions
            readiness = get_audit_readiness_summary(transactions)
            item = next(
                row
                for row in reconciliation_work_order_rows(readiness, transactions)
                if row.get("blocker_type") == "Missing acquisition basis"
            )

            resolution_data = {
                "item_id": item["item_id"],
                "decision": "resolved",
                "event_classification": "cash_sale",
                "proceeds_method": "source_reported",
                "proceeds_value": "500",
                "basis_method": "documented_acquisition_cost",
                "basis_value": "125",
                "acquisition_date": "2021-03-15",
                "evidence_reference": "CPA workpaper BCH-2024-01",
                "resolution_status": "cpa_reviewed_position",
                "reviewer_name": "Jamie Reviewer",
                "reviewer_role": "cpa_ea_tax_professional",
            }
            with app.test_client() as client:
                client.post(
                    "/export/review_queue/save",
                    data={**resolution_data, "workflow_action": "preview", "professional_attestation": "yes"},
                )
                response = client.post(
                    "/export/review_queue/save",
                    data={**resolution_data, "workflow_action": "apply", "preview_confirmed": "yes"},
                )

            self.assertEqual(400, response.status_code)
            self.assertIn(b"Confirm that you are recording the named professional&#39;s direction", response.data)
            self.assertEqual([], sell.links)
            self.assertEqual([sell], transactions.transactions)
            self.assertIsNone(transactions.get_work_order_review(item["item_id"]))

            with app.app_context():
                db.drop_all()
                db.session.remove()
                db.engine.dispose()

    def test_professional_resolution_previews_without_mutation_then_applies_and_reverses(self):
        transactions = empty_transactions()
        sell = Sell("BCH", 0.5, datetime.datetime(2024, 2, 1), 1000, "exchange.csv")
        sell.set_economics(
            gross_usd_total=500,
            fee=5,
            net_usd_total=495,
            economics_source="Coinbase subtotal/fee/total",
        )
        transactions.transactions = [sell]
        transactions.set_holdings("BCH", 0)

        with tempfile.TemporaryDirectory() as temp_dir:
            class ExportRouteTestConfig(config_dict["Debug"]):
                TESTING = True
                WTF_CSRF_ENABLED = False
                INSTANCE_PATH = temp_dir
                SQLALCHEMY_DATABASE_URI = f"sqlite:///{temp_dir}/test.db"

            app = create_app(ExportRouteTestConfig, selenium=True)
            app.config["transactions"] = transactions
            readiness = get_audit_readiness_summary(transactions)
            item = next(
                row
                for row in reconciliation_work_order_rows(readiness, transactions)
                if row.get("blocker_type") == "Missing acquisition basis"
            )
            resolution_data = {
                "item_id": item["item_id"],
                "decision": "resolved",
                "event_classification": "cash_sale",
                "proceeds_method": "source_reported",
                "proceeds_value": "495",
                "basis_method": "documented_acquisition_cost",
                "basis_value": "125",
                "acquisition_date": "2021-03-15",
                "resolution_status": "cpa_reviewed_position",
                "evidence_reference": "Professional workpaper BCH-2024-01",
                "reviewer_name": "Jamie Reviewer",
                "reviewer_role": "cpa_ea_tax_professional",
                "professional_attestation": "yes",
            }

            with app.test_client() as client:
                direct_apply = client.post(
                    "/export/review_queue/save",
                    data={
                        **resolution_data,
                        "workflow_action": "apply",
                        "preview_confirmed": "yes",
                    },
                )
                self.assertEqual(400, direct_apply.status_code)
                self.assertIn(b"Review the current before/after impact", direct_apply.data)
                self.assertEqual([], sell.links)

                preview = client.post(
                    "/export/review_queue/save",
                    data={**resolution_data, "workflow_action": "preview"},
                )

                preview_text = preview.data.decode("utf-8", errors="replace")
                error_position = preview_text.find("Review not saved")
                self.assertEqual(200, preview.status_code, preview_text[error_position:error_position + 500])
                self.assertIn(b"Review impact before applying", preview.data)
                self.assertIn(b"Allocated fee", preview.data)
                self.assertEqual([], sell.links)
                self.assertEqual([sell], transactions.transactions)
                self.assertIsNone(transactions.get_work_order_review(item["item_id"]))

                unconfirmed = client.post(
                    "/export/review_queue/save",
                    data={
                        **resolution_data,
                        "workflow_action": "apply",
                    },
                )
                self.assertEqual(400, unconfirmed.status_code)
                self.assertEqual([], sell.links)

                applied = client.post(
                    "/export/review_queue/save",
                    data={
                        **resolution_data,
                        "workflow_action": "apply",
                        "preview_confirmed": "yes",
                    },
                )
                self.assertEqual(302, applied.status_code)

                review = transactions.get_work_order_review(item["item_id"])
                self.assertEqual("Yes", review["calculation_applied"])
                self.assertEqual(datetime.date.today().isoformat(), review["direction_date"])
                self.assertEqual("Local Gainz user", review["direction_entered_by"])
                receipt = json.loads(review["calculation_receipt_json"])
                self.assertAlmostEqual(500, receipt["source_gross"])
                self.assertAlmostEqual(5, receipt["source_fee"])
                self.assertAlmostEqual(495, receipt["source_net"])
                self.assertAlmostEqual(370, receipt["added_gain_loss"])
                self.assertEqual(1, len(sell.links))

                reversed_response = client.post(
                    "/export/review_queue/reverse",
                    data={
                        "item_id": item["item_id"],
                        "reversal_note": "Corrected source records will be imported.",
                    },
                )
                self.assertEqual(302, reversed_response.status_code)

            self.assertEqual([], sell.links)
            self.assertEqual([sell], transactions.transactions)
            reversed_review = transactions.get_work_order_review(item["item_id"])
            self.assertEqual("needs_research", reversed_review["decision"])
            self.assertEqual("Reversed", reversed_review["calculation_applied"])
            self.assertIn("Corrected source records", reversed_review["reversal_note"])

            with app.app_context():
                db.drop_all()
                db.session.remove()
                db.engine.dispose()

    def test_cpa_conservative_max_gain_resolution_is_short_term_zero_basis_and_disclosed(self):
        transactions = empty_transactions()
        sell = Sell("BCH", 0.5, datetime.datetime(2024, 2, 1), 1000, "exchange.csv")
        transactions.transactions = [sell]
        transactions.set_holdings("BCH", 0)
        transactions.set_tax_year_record(
            2024,
            reported_proceeds=500,
            reported_cost_basis=0,
            reported_gain_loss=500,
            tax_paid=100,
            filing_status="Filed",
            evidence_reference="2024 filed Form 8949 and payment record",
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            class ExportRouteTestConfig(config_dict["Debug"]):
                TESTING = True
                WTF_CSRF_ENABLED = False
                INSTANCE_PATH = temp_dir
                SQLALCHEMY_DATABASE_URI = f"sqlite:///{temp_dir}/test.db"

            app = create_app(ExportRouteTestConfig, selenium=True)
            app.config["transactions"] = transactions
            readiness = get_audit_readiness_summary(transactions)
            item = next(
                row
                for row in reconciliation_work_order_rows(readiness, transactions)
                if row.get("blocker_type") == "Missing acquisition basis"
            )

            resolution_data = {
                "item_id": item["item_id"],
                "decision": "conservative_max_gain",
                "proceeds_method": "source_reported",
                "proceeds_value": "500",
                "evidence_reference": "Records searched; CPA workpaper BCH-2024-unknown-basis",
                "reviewer_name": "Jamie Reviewer",
                "reviewer_role": "cpa_ea_tax_professional",
                "professional_attestation": "yes",
            }
            with app.test_client() as client:
                preview = client.post(
                    "/export/review_queue/save",
                    data={**resolution_data, "workflow_action": "preview"},
                )
                self.assertEqual(200, preview.status_code)
                response = client.post(
                    "/export/review_queue/save",
                    data={
                        **resolution_data,
                        "workflow_action": "apply",
                        "preview_confirmed": "yes",
                    },
                )

            self.assertEqual(302, response.status_code)
            form_rows = get_form_8949_report_rows(transactions)
            self.assertEqual(1, len(form_rows))
            self.assertEqual("short", form_rows[0]["term"])
            self.assertEqual("", form_rows[0]["date_acquired"])
            self.assertAlmostEqual(500, form_rows[0]["proceeds"])
            self.assertAlmostEqual(0, form_rows[0]["cost_basis"])
            self.assertAlmostEqual(500, form_rows[0]["gain_loss"])

            after = get_audit_readiness_summary(transactions)
            self.assertEqual([], after["missing_records"]["basis"])
            self.assertTrue(after["is_ready"])

            workpapers = cpa_resolution_workpaper_rows(after, transactions)
            self.assertEqual(1, len(workpapers))
            self.assertEqual("Use $0 basis for a conservative full-proceeds gain", workpapers[0]["review_decision_label"])
            self.assertEqual("Unknown date - recorded short-term assumption", workpapers[0]["acquisition_date_method_label"])
            self.assertEqual("", workpapers[0]["acquisition_date"])
            self.assertIn("may overstate tax", workpapers[0]["assumption_disclosure"])

            packet_dir = AuditPacketService(
                str(Path(temp_dir) / "packets"),
                str(Path(temp_dir) / "exports"),
            ).create_packet(transactions)
            with (Path(packet_dir) / "01_reports" / "cpa_resolution_workpapers.csv").open(
                "r",
                encoding="utf-8-sig",
                newline="",
            ) as handle:
                packet_rows = list(csv.DictReader(handle))
            self.assertEqual("cpa_conservative_short_term", packet_rows[0]["acquisition_date_method"])
            self.assertIn("may overstate tax", packet_rows[0]["assumption_disclosure"])
            workbook_path = next((Path(packet_dir) / "01_reports").glob("*.xlsx"))
            workbook = load_workbook(workbook_path, data_only=False)
            workpaper_sheet = workbook["Professional Workpapers"]
            self.assertEqual("Use $0 basis for a conservative full-proceeds gain", workpaper_sheet["B5"].value)
            self.assertEqual("Unknown date - recorded short-term assumption", workpaper_sheet["N5"].value)
            self.assertIsNone(workpaper_sheet["O5"].value)
            self.assertIn("may overstate tax", workpaper_sheet["P5"].value)
            workbook.close()

            with app.app_context():
                db.drop_all()
                db.session.remove()
                db.engine.dispose()

    def test_guided_review_queue_shows_filed_tax_cross_check_for_gap_year(self):
        transactions = empty_transactions()
        buy = Buy("BTC", 0.1, datetime.datetime(2024, 1, 1), 100, "coinbase.csv")
        linked_sell = Sell("BTC", 0.1, datetime.datetime(2024, 5, 1), 300, "coinbase.csv")
        linked_sell.link_transaction(buy, 0.1)
        missing_basis_sell = Sell("BTC", 0.5, datetime.datetime(2024, 6, 1), 900, "coinbase.csv")
        transactions.transactions = [buy, linked_sell, missing_basis_sell]
        transactions.set_holdings("BTC", 0)
        transactions.set_tax_year_record(
            2024,
            reported_proceeds=1200,
            reported_cost_basis=500,
            reported_gain_loss=700,
            tax_paid=150,
            filing_status="Filed",
            evidence_reference="2024 filed Form 8949",
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            class ExportRouteTestConfig(config_dict["Debug"]):
                TESTING = True
                WTF_CSRF_ENABLED = False
                INSTANCE_PATH = temp_dir
                SQLALCHEMY_DATABASE_URI = f"sqlite:///{temp_dir}/test.db"

            app = create_app(ExportRouteTestConfig, selenium=True)
            app.config["transactions"] = transactions

            with app.test_client() as client:
                response = client.get("/export/review_queue?guided=1")

            self.assertEqual(200, response.status_code)
            self.assertIn(b"Filed tax cross-check", response.data)
            self.assertIn(b"2024 filed evidence is recorded", response.data)
            self.assertIn(b"Gainz proceeds", response.data)
            self.assertIn(b"Filed proceeds", response.data)
            self.assertIn(b"$1,200.00", response.data)
            self.assertIn(b"$150.00", response.data)
            self.assertIn(b"comparison evidence only", response.data)
            self.assertIn(b"they do not replace source records", response.data)

            with app.app_context():
                db.drop_all()
                db.session.remove()
                db.engine.dispose()

    def test_export_routes_require_draft_acknowledgement_when_not_ready(self):
        transactions = empty_transactions()

        with tempfile.TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir) / "tax-review-output"

            class ExportRouteTestConfig(config_dict["Debug"]):
                TESTING = True
                WTF_CSRF_ENABLED = False
                INSTANCE_PATH = temp_dir
                SQLALCHEMY_DATABASE_URI = f"sqlite:///{temp_dir}/test.db"
                EXPORT_FOLDER = str(output_dir)
                AUDIT_PACKET_FOLDER = str(output_dir)

            app = create_app(ExportRouteTestConfig, selenium=True)
            app.config["transactions"] = transactions

            with app.test_client() as client:
                export_response = client.post(
                    "/export/save",
                    json={"output_location": "audit_packets"},
                )
                packet_response = client.post(
                    "/export/audit_packet",
                    json={"output_location": "audit_packets"},
                )

            self.assertEqual(400, export_response.status_code)
            self.assertEqual(400, packet_response.status_code)
            self.assertTrue(export_response.get_json()["requires_draft_acknowledgement"])
            self.assertTrue(packet_response.get_json()["requires_draft_acknowledgement"])

            with app.app_context():
                db.drop_all()
                db.session.remove()
                db.engine.dispose()

    def test_packet_preview_json_returns_machine_readable_counts(self):
        transactions = empty_transactions()

        with tempfile.TemporaryDirectory() as temp_dir:
            source_path = Path(temp_dir) / "source.csv"
            source_path.write_text("demo source", encoding="utf-8")
            evidence_path = Path(temp_dir) / "2024_crypto_workbook.csv"
            evidence_path.write_text("Year,Total\n2024,1\n", encoding="utf-8")
            transactions.transactions = [
                Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, str(source_path))
            ]
            transactions.set_tax_evidence_record(
                year=2024,
                evidence_type="crypto_workbook",
                evidence_label=evidence_path.name,
                evidence_path=str(evidence_path),
            )
            output_dir = Path(temp_dir) / "preview-output"

            class PreviewRouteTestConfig(config_dict["Debug"]):
                TESTING = True
                WTF_CSRF_ENABLED = False
                INSTANCE_PATH = temp_dir
                SQLALCHEMY_DATABASE_URI = f"sqlite:///{temp_dir}/test.db"
                EXPORT_FOLDER = str(Path(temp_dir) / "default-exports")
                AUDIT_PACKET_FOLDER = str(output_dir)

            app = create_app(PreviewRouteTestConfig, selenium=True)
            app.config["transactions"] = transactions

            with app.test_client() as client:
                response = client.post(
                    "/export/packet_preview.json",
                    json={"output_location": "audit_packets"},
                )

            self.assertEqual(200, response.status_code)
            payload = response.get_json()
            preview = payload["packet_preview"]
            self.assertEqual(str(output_dir.resolve()), preview["output_folder"])
            self.assertEqual(1, preview["copied_files_count"])
            self.assertEqual(1, preview["transaction_source_files_count"])
            self.assertEqual(1, preview["reference_only_files_count"])
            self.assertEqual(0, preview["missing_tax_evidence_count"])
            self.assertEqual(
                len(payload["readiness"]["blocker_groups"]),
                preview["unresolved_blocker_group_count"],
            )
            self.assertFalse(output_dir.exists())

            with app.app_context():
                db.drop_all()
                db.session.remove()
                db.engine.dispose()

    def test_packet_preview_json_defaults_to_export_page_output_folder(self):
        transactions = empty_transactions()

        with tempfile.TemporaryDirectory() as temp_dir:
            detected_tax_folder = Path(temp_dir) / "DetectedTaxes"
            detected_tax_folder.mkdir()

            class PreviewDefaultRouteTestConfig(config_dict["Debug"]):
                TESTING = True
                WTF_CSRF_ENABLED = False
                INSTANCE_PATH = temp_dir
                SQLALCHEMY_DATABASE_URI = f"sqlite:///{temp_dir}/test.db"
                EXPORT_FOLDER = str(Path(temp_dir) / "default-exports")
                AUDIT_PACKET_FOLDER = str(Path(temp_dir) / "default-packets")

            app = create_app(PreviewDefaultRouteTestConfig, selenium=True)
            app.config["transactions"] = transactions

            with app.test_client() as client, patch(
                "app.export.routes._detected_tax_folder",
                return_value=str(detected_tax_folder),
            ):
                response = client.get("/export/packet_preview.json")

            self.assertEqual(200, response.status_code)
            preview = response.get_json()["packet_preview"]
            self.assertEqual(str(detected_tax_folder.resolve()), preview["output_folder"])
            self.assertFalse(detected_tax_folder.joinpath("gainz_audit_packet_DRAFT_YYYY-MM-DD_HH-MM-SS").exists())

            with app.app_context():
                db.drop_all()
                db.session.remove()
                db.engine.dispose()

    def test_health_check_reports_app_version(self):
        app = create_app(config_dict["Debug"], selenium=True)
        with app.test_client() as client:
            response = client.get("/healthz")

        self.assertEqual(200, response.status_code)
        self.assertEqual({"status": "ok", "version": APP_VERSION}, response.get_json())

    def test_form_8949_rows_are_built_from_links_with_prorated_fees(self):
        transactions = empty_transactions()
        buy = Buy("BTC", 2, datetime.datetime(2024, 1, 1), 100, "buy-source")
        buy.fee = 10.0
        sell = Sell("BTC", 1, datetime.datetime(2024, 6, 1), 300, "sell-source")
        sell.fee = 6.0
        sell.link_transaction(buy, 1)
        transactions.transactions = [buy, sell]

        rows = get_form_8949_report_rows(transactions, asset="BTC", term="short")
        table_rows = get_form_8949_table_data(transactions, asset="BTC", term="short")
        totals = get_form_8949_totals(transactions)

        self.assertEqual(1, len(rows))
        self.assertEqual("short", rows[0]["term"])
        self.assertEqual(294, rows[0]["proceeds"])
        self.assertEqual(105, rows[0]["cost_basis"])
        self.assertEqual(189, rows[0]["gain_loss"])
        self.assertEqual("$294.00", table_rows[0][3])
        self.assertEqual("$105.00", table_rows[0][4])
        self.assertEqual("$189.00", table_rows[0][5])
        self.assertEqual(294, totals["short"]["proceeds"])
        self.assertEqual(189, totals["total"]["gain_loss"])

    def test_audit_packet_service_creates_manifest_and_report(self):
        transactions = empty_transactions()
        buy = Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "demo_data/cash_app_sample.csv")
        sell = Sell("BTC", 1, datetime.datetime(2024, 6, 1), 300, "demo_data/cash_app_sample.csv")
        sell.link_transaction(buy, 1)
        transactions.transactions = [buy, sell]
        transactions.set_holdings("BTC", 0)
        transactions.import_warnings = ["Example warning"]
        transactions.set_tax_year_record(
            2024,
            reported_proceeds=300,
            reported_cost_basis=100,
            reported_gain_loss=200,
            tax_paid=25,
            evidence_reference="2024 filed return",
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            tax_evidence_path = Path(temp_dir) / "2024_crypto_tax_workbook.csv"
            tax_evidence_path.write_text(
                "Year,Reported Proceeds,Reported Cost Basis,Reported Gain Loss,Tax Paid\n"
                "2024,300,100,200,25\n",
                encoding="utf-8",
            )
            transactions.set_tax_evidence_record(
                year=2024,
                evidence_type="crypto_workbook",
                evidence_label=tax_evidence_path.name,
                evidence_path=str(tax_evidence_path),
            )
            packet_root = Path(temp_dir) / "packets"
            export_root = Path(temp_dir) / "exports"
            packet_path = Path(AuditPacketService(packet_root, export_root).create_packet(transactions))

            self.assertTrue((packet_path / "00_memos" / "METHODOLOGY.md").exists())
            self.assertTrue((packet_path / "README_FIRST.md").exists())
            self.assertTrue((packet_path / "PACKET_STATUS.md").exists())
            self.assertTrue((packet_path / "CPA_HANDOFF.md").exists())
            self.assertTrue((packet_path / "FOR_CPAS.md").exists())
            self.assertTrue((packet_path / "PRIVACY_AND_EVIDENCE_HANDLING.md").exists())
            self.assertTrue((packet_path / "03_manifests" / "evidence_manifest.csv").exists())
            self.assertTrue((packet_path / "01_reports" / "form_8949_short_term.csv").exists())
            self.assertTrue((packet_path / "01_reports" / "form_8949_totals.csv").exists())
            self.assertTrue((packet_path / "01_reports" / "import_economics.csv").exists())
            self.assertTrue((packet_path / "01_reports" / "holdings_reconciliation.csv").exists())
            self.assertTrue((packet_path / "01_reports" / "current_holdings_lots.csv").exists())
            self.assertTrue((packet_path / "01_reports" / "import_warnings.csv").exists())
            self.assertTrue((packet_path / "01_reports" / "source_overlap_review.csv").exists())
            self.assertTrue((packet_path / "01_reports" / "tax_filing_alignment.csv").exists())
            self.assertTrue((packet_path / "01_reports" / "tax_evidence_inventory.csv").exists())
            self.assertTrue((packet_path / "01_reports" / "tax_evidence_items.csv").exists())
            self.assertTrue((packet_path / "01_reports" / "suggested_filed_totals.csv").exists())
            self.assertTrue((packet_path / "01_reports" / "reconciliation_work_order.csv").exists())
            self.assertTrue((packet_path / "01_reports" / "reconciliation_work_order.md").exists())
            self.assertTrue((packet_path / "01_reports" / "unknown_gap_memos.csv").exists())
            self.assertTrue((packet_path / "01_reports" / "unknown_gap_memos.md").exists())
            self.assertTrue((packet_path / "03_manifests" / "tax_evidence_inventory.json").exists())
            self.assertTrue((packet_path / "03_manifests" / "suggested_filed_totals.json").exists())
            self.assertEqual(1, len(list((packet_path / "01_reports").glob("*.xlsx"))))
            self.assertEqual(1, len(list((packet_path / "02_source_files").glob("*.csv"))))

            with open(packet_path / "01_reports" / "import_warnings.csv", newline="", encoding="utf-8") as file:
                warning_rows = list(csv.DictReader(file))
            self.assertEqual("Example warning", warning_rows[0]["warning"])
            self.assertEqual("Active", warning_rows[0]["active_status"])
            self.assertEqual("Needs review", warning_rows[0]["status"])

            with open(packet_path / "01_reports" / "form_8949_totals.csv", newline="", encoding="utf-8") as file:
                totals = {row["term"]: row for row in csv.DictReader(file)}
            self.assertEqual("300.00", totals["short"]["proceeds"])
            self.assertEqual("200.00", totals["total"]["gain_loss"])

            with open(packet_path / "01_reports" / "holdings_reconciliation.csv", newline="", encoding="utf-8") as file:
                holdings_rows = list(csv.DictReader(file))
            self.assertEqual("BTC", holdings_rows[0]["asset"])
            self.assertEqual("Verified", holdings_rows[0]["status"])

            with open(packet_path / "01_reports" / "tax_filing_alignment.csv", newline="", encoding="utf-8") as file:
                tax_alignment_rows = list(csv.DictReader(file))
            self.assertEqual("2024", tax_alignment_rows[0]["year"])
            self.assertEqual("Aligned", tax_alignment_rows[0]["status"])
            self.assertEqual("25.00", tax_alignment_rows[0]["tax_paid"])

            with open(packet_path / "01_reports" / "suggested_filed_totals.csv", newline="", encoding="utf-8") as file:
                suggested_rows = list(csv.DictReader(file))
            self.assertEqual("2024", suggested_rows[0]["year"])
            self.assertEqual("High", suggested_rows[0]["confidence"])
            self.assertEqual("200.00", suggested_rows[0]["reported_gain_loss"])
            self.assertEqual("1", suggested_rows[0]["combined_suggestions_count"])

            summary = json.loads((packet_path / "03_manifests" / "audit_packet_summary.json").read_text(encoding="utf-8"))
            self.assertEqual(200, summary["form_8949_totals"]["total"]["gain_loss"])
            self.assertEqual(1, summary["import_warning_count"])
            self.assertEqual("Aligned", summary["tax_filing_alignment"]["overall_status"])
            self.assertIn("tax_evidence_inventory", summary)
            self.assertEqual(1, len(summary["suggested_filed_totals"]))

            with open(packet_path / "03_manifests" / "evidence_manifest.csv", newline="", encoding="utf-8") as file:
                manifest_rows = list(csv.DictReader(file))
            tax_evidence_rows = [row for row in manifest_rows if row["category"] == "tax_evidence"]
            self.assertEqual("REFERENCE_ONLY", tax_evidence_rows[0]["status"])
            self.assertFalse((packet_path / "02_source_files" / "tax_evidence" / tax_evidence_path.name).exists())
            self.assertTrue((packet_path / "00_memos" / "DRAFT_NOT_FILING_READY.md").exists())
            packet_status = (packet_path / "PACKET_STATUS.md").read_text(encoding="utf-8")
            readme_first = (packet_path / "README_FIRST.md").read_text(encoding="utf-8")
            self.assertNotEqual(readme_first, packet_status)
            self.assertIn("Read This First", readme_first)
            self.assertIn("Folder Map", readme_first)
            self.assertIn("unknown_gap_memos.md", readme_first)
            self.assertIn("Gainz Packet Status", packet_status)
            self.assertIn("Open Blockers", packet_status)
            self.assertIn("DRAFT - NOT FILING READY", packet_status)
            self.assertIn("Reference-only tax evidence records: 1", packet_status)
            self.assertIn("FOR_CPAS.md", packet_status)
            self.assertIn("CPA_HANDOFF.md", packet_status)
            cpa_handoff = (packet_path / "CPA_HANDOFF.md").read_text(encoding="utf-8")
            self.assertIn("How This Packet Was Generated", cpa_handoff)
            self.assertIn("Tax evidence is reference-only by default", cpa_handoff)
            for_cpas = (packet_path / "FOR_CPAS.md").read_text(encoding="utf-8")
            self.assertIn("For CPAs", for_cpas)
            self.assertIn("Suggested Review Order", for_cpas)
            self.assertIn("Reference-only tax evidence records: 1", for_cpas)
            self.assertIn("unknown_gap_memos.md", for_cpas)
            self.assertIn("Questions For The Taxpayer", for_cpas)
            self.assertIn("Can you provide source records", for_cpas)
            privacy_handling = (packet_path / "PRIVACY_AND_EVIDENCE_HANDLING.md").read_text(encoding="utf-8")
            self.assertIn("does not require a hosted account", privacy_handling)
            self.assertIn("Reference only means", privacy_handling)

            workbook_path = next((packet_path / "01_reports").glob("*.xlsx"))
            workbook = load_workbook(workbook_path, read_only=True)
            self.assertIn("Packet Status", workbook.sheetnames)
            status_sheet = workbook["Packet Status"]
            self.assertEqual("DRAFT - NOT FILING READY", status_sheet["A1"].value)
            status_values = [
                str(cell.value or "")
                for row in status_sheet.iter_rows()
                for cell in row
            ]
            self.assertIn("Material inputs and assumptions", status_values)
            workbook.close()

            with open(packet_path / "01_reports" / "reconciliation_work_order.csv", newline="", encoding="utf-8") as file:
                work_order_rows = list(csv.DictReader(file))
            self.assertTrue(any(row["blocker_type"] == "Import warning decision" for row in work_order_rows))
            self.assertIn("item_id", work_order_rows[0])
            self.assertIn("review_decision", work_order_rows[0])
            self.assertIn("cpa_question", work_order_rows[0])
            self.assertIn("what_gainz_knows", work_order_rows[0])

            with open(packet_path / "01_reports" / "unknown_gap_memos.csv", newline="", encoding="utf-8") as file:
                memo_rows = list(csv.DictReader(file))
            self.assertTrue(any(row["blocker_type"] == "Import warning decision" for row in memo_rows))
            memo_text = (packet_path / "01_reports" / "unknown_gap_memos.md").read_text(encoding="utf-8")
            self.assertIn("Unknown Gap Memos", memo_text)
            self.assertIn("preserve uncertainty", memo_text)
            self.assertEqual([], list(export_root.glob("*.xlsx")))

    def test_audit_packet_manifest_uses_reference_only_for_label_only_tax_evidence(self):
        transactions = empty_transactions()
        buy = Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "demo_data/cash_app_sample.csv")
        transactions.transactions = [buy]
        transactions.set_holdings("BTC", 1)
        transactions.set_tax_evidence_record(
            year=2024,
            evidence_type="payment_receipt",
            evidence_label="2024 payment confirmation",
            evidence_path="",
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            packet_path = Path(
                AuditPacketService(
                    Path(temp_dir) / "packets",
                    Path(temp_dir) / "exports",
                ).create_packet(transactions)
            )

            with open(packet_path / "03_manifests" / "evidence_manifest.csv", newline="", encoding="utf-8") as file:
                manifest_rows = list(csv.DictReader(file))

            tax_evidence_rows = [row for row in manifest_rows if row["category"] == "tax_evidence"]
            self.assertEqual(1, len(tax_evidence_rows))
            self.assertEqual("REFERENCE_ONLY", tax_evidence_rows[0]["status"])

    def test_audit_packet_copies_tax_evidence_only_when_explicitly_enabled(self):
        transactions = empty_transactions()
        buy = Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "demo_data/cash_app_sample.csv")
        sell = Sell("BTC", 1, datetime.datetime(2024, 6, 1), 300, "demo_data/cash_app_sample.csv")
        sell.link_transaction(buy, 1)
        transactions.transactions = [buy, sell]
        transactions.set_holdings("BTC", 0)

        with tempfile.TemporaryDirectory() as temp_dir:
            tax_evidence_path = Path(temp_dir) / "2024_filed_return.pdf"
            tax_evidence_path.write_text("synthetic filed return", encoding="utf-8")
            transactions.set_tax_evidence_record(
                year=2024,
                evidence_type="filed_return",
                evidence_label=tax_evidence_path.name,
                evidence_path=str(tax_evidence_path),
                copy_to_packet=True,
            )

            packet_path = Path(
                AuditPacketService(
                    Path(temp_dir) / "packets",
                    Path(temp_dir) / "exports",
                ).create_packet(transactions)
            )

            copied_evidence = packet_path / "02_source_files" / "tax_evidence" / tax_evidence_path.name
            self.assertTrue(copied_evidence.exists())

            with open(packet_path / "03_manifests" / "evidence_manifest.csv", newline="", encoding="utf-8") as file:
                manifest_rows = list(csv.DictReader(file))
            tax_evidence_rows = [row for row in manifest_rows if row["category"] == "tax_evidence"]
            self.assertEqual("COPIED", tax_evidence_rows[0]["status"])

    def test_packet_preview_counts_copied_and_reference_only_files(self):
        transactions = empty_transactions()
        buy = Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "demo_data/cash_app_sample.csv")
        sell = Sell("BTC", 1, datetime.datetime(2024, 6, 1), 300, "demo_data/cash_app_sample.csv")
        sell.link_transaction(buy, 1)
        transactions.transactions = [buy, sell]
        transactions.set_holdings("BTC", 0)

        with tempfile.TemporaryDirectory() as temp_dir:
            copied_path = Path(temp_dir) / "2024_filed_return.pdf"
            reference_path = Path(temp_dir) / "2024_crypto_workbook.csv"
            copied_path.write_text("filed return", encoding="utf-8")
            reference_path.write_text("crypto workbook", encoding="utf-8")
            transactions.set_tax_evidence_record(
                year=2024,
                evidence_type="filed_return",
                evidence_label=copied_path.name,
                evidence_path=str(copied_path),
                copy_to_packet=True,
            )
            transactions.set_tax_evidence_record(
                year=2024,
                evidence_type="crypto_workbook",
                evidence_label=reference_path.name,
                evidence_path=str(reference_path),
            )

            readiness = get_audit_readiness_summary(transactions)
            preview = get_packet_preview(transactions, readiness, temp_dir)

            self.assertTrue(preview["is_draft"])
            self.assertEqual(2, preview["copied_files_count"])
            self.assertEqual(1, preview["transaction_source_files_count"])
            self.assertEqual(1, preview["copied_tax_evidence_count"])
            self.assertEqual(1, preview["reference_only_files_count"])
            self.assertIn("gainz_audit_packet_DRAFT_", preview["packet_name"])

    def test_audit_packet_preserves_cleared_import_warning_review_records(self):
        transactions = empty_transactions()
        buy = Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "demo_data/cash_app_sample.csv")
        transactions.transactions = [buy]
        transactions.set_holdings("BTC", 1)
        warning = "Imported row 10 from sample.csv with $0 USD spot price."
        transactions.set_import_warning_review(
            warning,
            decision="true_zero_value_transfer",
            note="Reviewed before warning was cleared.",
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            packet_root = Path(temp_dir) / "packets"
            export_root = Path(temp_dir) / "exports"
            packet_path = Path(AuditPacketService(packet_root, export_root).create_packet(transactions))

            with open(packet_path / "01_reports" / "import_warnings.csv", newline="", encoding="utf-8") as file:
                warning_rows = list(csv.DictReader(file))

            self.assertEqual(1, len(warning_rows))
            self.assertEqual("Cleared from active warnings", warning_rows[0]["active_status"])
            self.assertEqual("Own wallet/account transfer", warning_rows[0]["decision"])
            self.assertEqual("Reviewed before warning was cleared.", warning_rows[0]["note"])

            summary = json.loads((packet_path / "03_manifests" / "audit_packet_summary.json").read_text(encoding="utf-8"))
            self.assertEqual(0, summary["import_warning_count"])
            self.assertEqual(1, summary["import_warning_review_count"])

    def test_privacy_mode_page_lists_local_storage_and_open_folder_action(self):
        transactions = empty_transactions()
        with tempfile.TemporaryDirectory() as temp_dir:
            class PrivacyRouteTestConfig(config_dict["Debug"]):
                TESTING = True
                SQLALCHEMY_DATABASE_URI = "sqlite:///:memory:"
                INSTANCE_PATH = str(Path(temp_dir) / "instance")
                UPLOAD_FOLDER = str(Path(temp_dir) / "uploads")
                EXPORT_FOLDER = str(Path(temp_dir) / "exports")
                AUDIT_PACKET_FOLDER = str(Path(temp_dir) / "packets")

            app = create_app(PrivacyRouteTestConfig, selenium=True)
            app.config["transactions"] = transactions

            with app.test_client() as client:
                response = client.get("/privacy/")

                self.assertEqual(200, response.status_code)
                self.assertIn(b"Offline / Privacy Mode", response.data)
                self.assertIn(b"What Gainz Stores Locally", response.data)
                self.assertIn(b"database.db", response.data)
                self.assertIn(b"What Gainz Does Not Upload", response.data)
                self.assertIn(b"Synced Folders", response.data)
                self.assertIn(b"Delete Local Data", response.data)

                with patch("app.privacy.routes._open_folder") as open_folder:
                    post_response = client.post("/privacy/open_data_folder")

                self.assertEqual(302, post_response.status_code)
                open_folder.assert_called_once()

    def test_quantity_parser_preserves_scientific_notation(self):
        self.assertAlmostEqual(0.00002618, parse_quantity_value("0.00002618"), places=12)
        self.assertAlmostEqual(0.00002618, parse_quantity_value("2.618e-05"), places=12)
        self.assertAlmostEqual(-0.00002618, parse_quantity_value("(2.618E-05)"), places=12)

    def test_cash_app_tiny_quantities_are_not_enlarged(self):
        transactions = empty_transactions()
        source_rows = [
            ("tiny-1", "2024-01-01 10:00:00 EST", "2.31", "88235.29411765", "0.00002618"),
            ("tiny-2", "2024-01-02 10:00:00 EST", "7.00", "89640.15879114", "0.00007809"),
            ("tiny-3", "2024-01-03 10:00:00 EST", "5.54", "97999.29241111", "0.00005653"),
        ]

        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "cash_app_tiny_quantities.csv"
            with source.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow([
                    "Transaction ID", "Date", "Transaction Type", "Currency", "Amount",
                    "Fee", "Net Amount", "Asset Type", "Asset Price", "Asset Amount",
                    "Status", "Notes",
                ])
                for transaction_id, date, amount, price, quantity in source_rows:
                    writer.writerow([
                        transaction_id, date, "Bitcoin Buy", "USD", f"(${amount})", "$0",
                        f"(${amount})", "BTC", f"${price}", quantity, "COMPLETED", "Synthetic tiny buy",
                    ])

            imported_count, skipped_count = import_transactions(str(source), transactions)

        self.assertEqual(3, imported_count)
        self.assertEqual(0, skipped_count)
        self.assertEqual(3, len(transactions.transactions))
        for transaction, expected in zip(transactions.transactions, (0.00002618, 0.00007809, 0.00005653)):
            self.assertAlmostEqual(expected, transaction.quantity, places=12)
            self.assertNotEqual("BLOCKING", transaction.input_reliability_status)
        self.assertEqual(3, len(transactions.import_receipts))
        self.assertEqual("0.00002618", transactions.import_receipts[0]["source_quantity"])

    def test_quantity_value_mismatch_blocks_tax_outputs(self):
        transactions = empty_transactions()
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "cash_app_bad_quantity.csv"
            source.write_text(
                "Transaction ID,Date,Transaction Type,Currency,Amount,Fee,Net Amount,Asset Type,Asset Price,Asset Amount\n"
                "bad-1,2024-01-01 10:00:00 EST,Bitcoin Buy,USD,($2.31),$0,($2.31),BTC,$88235.29,2.618\n",
                encoding="utf-8",
            )
            imported_count, _ = import_transactions(str(source), transactions)

        self.assertEqual(1, imported_count)
        self.assertEqual("BLOCKING", transactions.transactions[0].input_reliability_status)
        transactions.set_tax_year_record(
            2024,
            reported_proceeds=2.31,
            reported_cost_basis=2.31,
            reported_gain_loss=0,
            tax_paid=0,
            filing_status="Filed",
            evidence_reference="2024 filed return and payment record",
        )
        readiness = get_audit_readiness_summary(transactions)
        self.assertEqual("Inputs not reliable", readiness["status"])
        self.assertEqual("Suppressed", readiness["metrics"]["form_8949_proceeds"])
        self.assertEqual(1, readiness["metrics"]["input_reliability_failures"])
        self.assertEqual(0, readiness["metrics"]["import_economics_warnings"])
        alignment = get_tax_filing_alignment_summary(transactions)
        self.assertEqual("Inputs not reliable", alignment["rows"][0]["status"])
        self.assertEqual("Suppressed", alignment["rows"][0]["calculated_proceeds_display"])

    def test_coinbase_raw_dual_leg_import_requires_preview_then_preserves_legs(self):
        transactions = empty_transactions()
        headers = [
            "Timestamp",
            "Transaction Type",
            "Transaction ID",
            "Asset Acquired",
            "Quantity Acquired (Bought, Received, etc)",
            "Cost Basis (incl. fees and/or spread) (USD)",
            "Asset Disposed (Sold, Sent, etc)",
            "Quantity Disposed",
            "Proceeds (excl. fees and/or spread) (USD)",
            "Notes",
        ]
        rows = [
            ["2024-01-01 12:00:00 UTC", "Buy", "cb-1", "BTC", "0.05", "2000", "USD", "2000", "0", "Buy BTC"],
            ["2024-06-01 12:00:00 UTC", "Sell", "cb-2", "USD", "1500", "0", "BTC", "0.025", "1500", "Sell BTC"],
            ["2024-07-01 12:00:00 UTC", "Convert", "cb-3", "BTC", "0.02", "1000", "ETH", "0.5", "1300", "Convert ETH to BTC"],
            ["2024-08-01 12:00:00 UTC", "Deposit", "cb-4", "USD", "250", "250", "", "", "0", "Fiat-only row"],
        ]

        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "Coinbase_raw_transactions.csv"
            with source.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(headers)
                writer.writerows(rows)

            analysis = analyze_csv_import(str(source))
            self.assertEqual("coinbase_raw", analysis["detected_format"])
            self.assertEqual(4, analysis["import_preview"]["source_rows"])
            self.assertEqual(4, analysis["import_preview"]["output_rows"])
            self.assertEqual(1, analysis["import_preview"]["skipped_source_rows"])
            self.assertEqual(2800.0, analysis["import_preview"]["source_reported_proceeds"])
            self.assertEqual(3000.0, analysis["import_preview"]["source_reported_basis"])

            preview_result = ImportService(temp_dir).import_upload(FileUpload(source), transactions)
            self.assertTrue(preview_result["native_preview_required"])
            self.assertEqual([], transactions.transactions)

            import_result = ImportService(temp_dir).import_native_file(
                preview_result["file_path"],
                transactions,
            )

        self.assertEqual(4, import_result["imported_count"])
        self.assertEqual(["buy", "sell", "sell", "buy"], [row.trans_type for row in transactions.transactions])
        self.assertEqual(["acquired", "disposed", "disposed", "acquired"], [row.source_leg for row in transactions.transactions])
        self.assertEqual(["cb-1", "cb-2", "cb-3", "cb-3"], [row.source_transaction_id for row in transactions.transactions])
        skipped_receipts = [row for row in transactions.import_receipts if row["outcome"] == "Skipped"]
        self.assertEqual(1, len(skipped_receipts))
        self.assertEqual("cb-4", skipped_receipts[0]["source_transaction_id"])
        self.assertIn("No supported crypto", skipped_receipts[0]["reason"])

    def test_current_coinbase_raw_date_and_time_preview_payload_commits_atomically(self):
        transactions = empty_transactions()
        headers = [
            "Transaction ID",
            "Transaction Type",
            "Date & time",
            "Asset Acquired",
            "Quantity Acquired (Bought, Received, etc)",
            "Cost Basis (incl. fees and/or spread) (USD)",
            "Data Source",
            "Asset Disposed (Sold, Sent, etc)",
            "Quantity Disposed",
            "Proceeds (excl. fees and/or spread) (USD)",
        ]

        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "coinbase_current_raw.csv"
            with source.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(headers)
                writer.writerow([
                    "raw-1", "Buy", "2024-01-02 03:04:05 UTC", "BTC", "0.01",
                    "425.50", "Coinbase", "USD", "425.50", "0",
                ])

            service = ImportService(temp_dir)
            preview = service.import_upload(FileUpload(source), transactions)
            result = service.import_native_payload(
                preview["_prepared_payload_path"],
                transactions,
            )

        self.assertTrue(preview["native_preview_required"])
        self.assertEqual(1, preview["import_preview"]["output_rows"])
        self.assertEqual(1, result["imported_count"])
        self.assertTrue(result["transactional_commit"])
        self.assertEqual("raw-1", transactions.transactions[0].source_transaction_id)
        self.assertEqual(datetime.datetime(2024, 1, 2, 3, 4, 5, tzinfo=datetime.timezone.utc), transactions.transactions[0].time_stamp)
        self.assertEqual(1, len(transactions.saved_descriptions))

    def test_native_ledger_live_import_preserves_wallet_direction_and_crypto_fee(self):
        transactions = empty_transactions()
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "ledger_live_operations.csv"
            with source.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow([
                    "Operation Date", "Currency Ticker", "Operation Type", "Operation Amount",
                    "Operation Fees", "Operation Hash", "Account Name", "Account xpub",
                ])
                writer.writerow([
                    "2024-01-01T12:00:00Z", "BTC", "IN", "0.5", "0", "ledger-in", "BTC wallet", "synthetic-xpub",
                ])
                writer.writerow([
                    "2024-02-01T12:00:00Z", "BTC", "OUT", "-0.1", "0.00002", "ledger-out", "BTC wallet", "synthetic-xpub",
                ])

            service = ImportService(temp_dir)
            preview = service.import_upload(FileUpload(source), transactions)
            result = service.import_native_payload(
                preview["_prepared_payload_path"],
                transactions,
            )

        self.assertEqual("ledger_live", preview["detected_format"])
        self.assertEqual({"Receive": 1, "Send": 1}, preview["import_preview"]["row_counts_by_type"])
        self.assertEqual(2, result["imported_count"])
        self.assertEqual(["receive", "send"], [row.trans_type for row in transactions.transactions])
        outgoing = transactions.transactions[1]
        self.assertEqual("BTC", outgoing.fee_currency)
        self.assertAlmostEqual(0.00002, outgoing.source_fee_amount, places=10)
        self.assertIsNone(outgoing.fee)
        self.assertEqual("PASSED_WALLET_MOVEMENT", outgoing.input_reliability_status)
        self.assertEqual([], result["warnings"])

    def test_cash_app_withdrawal_integrity_uses_net_value_and_preserves_equation(self):
        transactions = empty_transactions()
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "cash_app_withdrawal_equation.csv"
            source.write_text(
                "Transaction ID,Date,Transaction Type,Currency,Amount,Fee,Net Amount,Asset Type,Asset Price,Asset Amount,Status,Notes\n"
                "send-1,2024-01-01 10:00:00 EST,Bitcoin Withdrawal,USD,($105.00),($5.00),($100.00),BTC,$10000.00,0.01,COMPLETED,Synthetic withdrawal\n",
                encoding="utf-8",
            )
            result = ImportService(temp_dir)._transactional_import(str(source), transactions)

        self.assertEqual(1, result["imported_count"])
        transaction = transactions.transactions[0]
        self.assertEqual("send", transaction.trans_type)
        self.assertEqual(105.0, transaction.gross_usd_total)
        self.assertEqual(5.0, transaction.fee)
        self.assertEqual(100.0, transaction.net_usd_total)
        self.assertEqual(100.0, transaction.source_usd_total)
        self.assertEqual("PASSED", transaction.input_reliability_status)
        self.assertNotIn("INPUT RELIABILITY BLOCKER", transaction.economics_warning)
        self.assertEqual(100.0, get_import_economics_rows(transactions)[0]["net_tax_usd"])

    def test_failed_native_commit_rolls_back_without_persistent_row_warnings(self):
        transactions = empty_transactions()
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "coinbase_bad_commit.csv"
            source.write_text(
                "Transaction ID,Transaction Type,Date & time,Asset Acquired,Quantity Acquired (Bought; Received; etc),Cost Basis (incl. fees and/or spread) (USD),Data Source,Asset Disposed (Sold; Sent; etc),Quantity Disposed,Proceeds (excl. fees and/or spread) (USD)\n"
                "bad-date,Buy,not-a-date,BTC,0.01,100,Coinbase,USD,100,0\n",
                encoding="utf-8",
            )
            # Semicolon punctuation normalizes to the same official header names.
            service = ImportService(temp_dir)
            preview = service.import_upload(FileUpload(source), transactions)
            result = service.import_native_payload(
                preview["_prepared_payload_path"],
                transactions,
            )

        self.assertTrue(result["source_failure"])
        self.assertTrue(result["rollback_complete"])
        self.assertEqual(0, result["persistent_warnings_added"])
        self.assertEqual([], transactions.transactions)
        self.assertEqual([], transactions.import_warnings)
        self.assertEqual([], transactions.import_receipts)
        self.assertEqual([], transactions.saved_descriptions)

    def test_failed_mapped_import_rolls_back_without_persistent_row_warnings(self):
        transactions = empty_transactions()
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "mapped_wallet.csv"
            source.write_text(
                "When,Direction,Coin,Units,USD Value\n"
                "2024-01-01T00:00:00Z,UNKNOWN,BTC,0.1,100\n",
                encoding="utf-8",
            )
            result = ImportService(temp_dir).import_mapped_file(
                str(source),
                transactions,
                header_row=1,
                column_mapping={
                    "date": "When",
                    "transaction_type": "Direction",
                    "asset_type": "Coin",
                    "asset_amount": "Units",
                    "fiat_amount": "USD Value",
                },
            )

        self.assertTrue(result["source_failure"])
        self.assertTrue(result["rollback_complete"])
        self.assertEqual(0, result["persistent_warnings_added"])
        self.assertEqual([], transactions.transactions)
        self.assertEqual([], transactions.import_warnings)
        self.assertEqual([], transactions.import_receipts)
        self.assertEqual([], transactions.saved_descriptions)

    def test_coinbase_readable_total_is_authoritative_and_fee_sign_does_not_warn(self):
        transactions = empty_transactions()
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "coinbase_readable.csv"
            source.write_text(
                "Timestamp,Transaction Type,Asset,Quantity Transacted,Spot Price Currency,"
                "Spot Price at Transaction,Subtotal,Total (inclusive of fees and/or spread),"
                "Fees and/or Spread,Notes\n"
                "2024-01-01T00:00:00Z,Buy,ETH,2,USD,100,200,205,-5,Synthetic buy\n"
                "2024-06-01T00:00:00Z,Sell,ETH,-0.5,USD,300,150,147,-3,Synthetic sell\n",
                encoding="utf-8",
            )
            result = ImportService(temp_dir).import_upload(FileUpload(source), transactions)

        self.assertEqual(2, result["imported_count"])
        self.assertEqual([], result["warnings"])
        buy, sell = transactions.transactions
        self.assertEqual(200.0, buy.gross_usd_total)
        self.assertEqual(205.0, buy.net_usd_total)
        self.assertEqual(205.0, buy.tax_usd_total)
        self.assertEqual(5.0, buy.fee)
        self.assertEqual(150.0, sell.gross_usd_total)
        self.assertEqual(147.0, sell.net_usd_total)
        self.assertEqual(147.0, sell.tax_usd_total)
        self.assertEqual(3.0, sell.fee)
        self.assertIn("Coinbase total inclusive", sell.economics_source)

    def test_unlinked_sale_quantity_never_appears_as_zero_basis_gain(self):
        transactions = empty_transactions()
        buy = Buy("BTC", 0.4, datetime.datetime(2024, 1, 1), 100, "synthetic.csv")
        sell = Sell("BTC", 1.0, datetime.datetime(2024, 6, 1), 200, "synthetic.csv")
        sell.link_transaction(buy, 0.4)
        transactions.transactions = [buy, sell]

        with tempfile.TemporaryDirectory() as temp_dir:
            export_path = transactions.export_to_excel(output_dir=temp_dir)
            workbook = load_workbook(export_path, read_only=True, data_only=True)
            try:
                gain_sheet = workbook["2024 BTC Gains"]
                values = [
                    cell
                    for row in gain_sheet.iter_rows(values_only=True)
                    for cell in row
                ]
            finally:
                workbook.close()

        self.assertEqual(1, len(get_form_8949_report_rows(transactions)))
        self.assertAlmostEqual(0.4, get_form_8949_report_rows(transactions)[0]["quantity"])
        self.assertNotIn("N/A", values)
        self.assertEqual(1, values.count(sell.id))

    def test_evidence_content_years_and_account_transcript_semantics(self):
        transactions = empty_transactions()
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "filed_totals_workbook.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Tax Year", "Filed Proceeds", "Generated Date"])
            sheet.append([2022, 100])
            sheet.append([2024, 2026, "2026-07-20"])
            workbook.save(workbook_path)

            detected_years = infer_tax_evidence_years_from_file(workbook_path)

        self.assertEqual([2022, 2024], detected_years)
        transactions.set_tax_evidence_record(
            year=2024,
            evidence_type="account_transcript",
            evidence_label="2024 IRS Record of Account",
            evidence_path="",
        )
        inventory = get_tax_evidence_inventory_summary(transactions)
        row = next(item for item in inventory["rows"] if item["year"] == 2024)
        self.assertEqual("2024 IRS Record of Account", row["filed_return_evidence"])
        self.assertEqual("Missing", row["payment_evidence"])

    def test_tax_evidence_classification_and_open_year_are_conservative(self):
        self.assertEqual("crypto_workbook", classify_tax_evidence("Crypto Taxes Paid.csv"))
        self.assertEqual("account_transcript", classify_tax_evidence("IRS Account Transcript 2024.pdf"))

        transactions = empty_transactions()
        current_year = datetime.datetime.now().year
        transactions.set_tax_evidence_record(
            year=current_year,
            evidence_type="estimate",
            evidence_label=f"{current_year} estimate",
            evidence_path="",
        )
        inventory = get_tax_evidence_inventory_summary(transactions)
        self.assertEqual("Open tax year - filing not due", inventory["rows"][0]["status"])
        self.assertEqual([], inventory["review_rows"])

    def test_unsafe_equal_extraction_is_low_confidence_and_generated_manifests_are_ignored(self):
        transactions = empty_transactions()
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "2024_crypto_tax_workbook.csv"
            source.write_text(
                "Year,Proceeds,Cost Basis,Gain Loss\n2024,0.20,0.20,0.20\n",
                encoding="utf-8",
            )
            manifest = Path(temp_dir) / "2024_audit_packet_manifest.csv"
            manifest.write_text(
                "Year,Proceeds,Cost Basis,Gain Loss\n2024,100,40,60\n",
                encoding="utf-8",
            )
            transactions.set_tax_evidence_record(
                year=2024,
                evidence_type="crypto_workbook",
                evidence_label=source.name,
                evidence_path=str(source),
            )
            transactions.set_tax_evidence_record(
                year=2024,
                evidence_type="crypto_workbook",
                evidence_label=manifest.name,
                evidence_path=str(manifest),
            )

            suggestions = get_suggested_filed_totals(transactions)

        self.assertEqual(1, len(suggestions))
        self.assertEqual("Low", suggestions[0]["confidence"])
        self.assertIn("reported_proceeds", suggestions[0]["field_provenance"])

    def test_packet_handles_target_transaction_uid_and_declares_contract_version(self):
        transactions = empty_transactions()
        sell = Sell("BTC", 0.5, datetime.datetime(2024, 6, 1), 30000, "synthetic-source.csv")
        transactions.transactions = [sell]
        transactions.set_holdings("BTC", 0)

        with tempfile.TemporaryDirectory() as temp_dir:
            packet_path = Path(
                AuditPacketService(
                    Path(temp_dir) / "packets",
                    Path(temp_dir) / "exports",
                ).create_packet(transactions)
            )
            with (packet_path / "01_reports" / "missing_basis_review.csv").open(
                newline="", encoding="utf-8"
            ) as file:
                rows = list(csv.DictReader(file))
            summary = json.loads(
                (packet_path / "03_manifests" / "audit_packet_summary.json").read_text(encoding="utf-8")
            )

        self.assertEqual(sell.uid, rows[0]["target_transaction_uid"])
        self.assertEqual("1.0", summary["packet_contract_version"])

    def test_packet_failure_is_atomic_and_leaves_only_failure_receipt(self):
        transactions = empty_transactions()
        transactions.transactions = [
            Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "synthetic-source.csv")
        ]
        transactions.set_holdings("BTC", 1)

        with tempfile.TemporaryDirectory() as temp_dir:
            packet_root = Path(temp_dir) / "packets"
            service = AuditPacketService(packet_root, Path(temp_dir) / "exports")
            with patch.object(
                service,
                "_write_reconciliation_work_order",
                side_effect=RuntimeError("synthetic packet failure"),
            ):
                with self.assertRaises(RuntimeError):
                    service.create_packet(transactions)

            published_directories = [path for path in packet_root.iterdir() if path.is_dir()]
            failure_receipts = list(packet_root.glob("FAILED_INCOMPLETE_*.txt"))

        self.assertEqual([], published_directories)
        self.assertEqual(1, len(failure_receipts))

    def test_unreliable_inputs_generate_suppression_memo_instead_of_form_8949(self):
        transactions = empty_transactions()
        buy = Buy("BTC", 1, datetime.datetime(2024, 1, 1), 100, "synthetic-source.csv")
        buy.set_economics(
            gross_usd_total=100,
            net_usd_total=100,
            source_quantity_text="0.00001",
            source_usd_total=100,
            implied_usd_total=10000000,
            value_variance_usd=9999900,
            value_tolerance_usd=10,
            input_reliability_status="BLOCKING",
            economics_warning="INPUT RELIABILITY BLOCKER: synthetic mismatch",
        )
        sell = Sell("BTC", 1, datetime.datetime(2024, 6, 1), 150, "synthetic-source.csv")
        buy.link_transaction(sell, 1)
        transactions.transactions = [buy, sell]
        transactions.set_holdings("BTC", 0)

        with tempfile.TemporaryDirectory() as temp_dir:
            packet_path = Path(
                AuditPacketService(
                    Path(temp_dir) / "packets",
                    Path(temp_dir) / "exports",
                ).create_packet(transactions)
            )

            self.assertTrue((packet_path / "01_reports" / "FORM_8949_SUPPRESSED.md").is_file())
            self.assertFalse((packet_path / "01_reports" / "form_8949_totals.csv").exists())
            summary = json.loads(
                (packet_path / "03_manifests" / "audit_packet_summary.json").read_text(encoding="utf-8")
            )
            workbook_path = next((packet_path / "01_reports").glob("*.xlsx"))
            workbook = load_workbook(workbook_path, read_only=True, data_only=True)
            try:
                self.assertIn("Calculations Suppressed", workbook.sheetnames)
                self.assertFalse(any(name.endswith(" Gains") for name in workbook.sheetnames))
                self.assertFalse(any(" 8949 " in f" {name} " for name in workbook.sheetnames))
                self.assertFalse(any(name.endswith(" Sales") for name in workbook.sheetnames))
            finally:
                workbook.close()
            with (packet_path / "03_manifests" / "evidence_manifest.csv").open(
                newline="", encoding="utf-8"
            ) as file:
                generated_rows = [
                    row for row in csv.DictReader(file) if row["status"] == "GENERATED"
                ]
            generated_source_exists = Path(generated_rows[0]["source_path"]).is_file()

        self.assertTrue(summary["form_8949_totals"]["suppressed"])
        self.assertEqual(str(packet_path), summary["packet_path"])
        self.assertNotIn(".building-", json.dumps(summary))
        self.assertEqual(1, len(generated_rows))
        self.assertTrue(generated_source_exists)
        self.assertNotIn(".building-", generated_rows[0]["source_path"])


if __name__ == "__main__":
    unittest.main()
